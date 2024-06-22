from django.shortcuts import render
from multiprocessing import context
from os import system
from django.shortcuts import render, redirect
import datetime
from brand_app.models import *
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from brand_app.models import mediums
from django.http import JsonResponse
import secrets
import sweetify
from django.http import HttpResponseBadRequest


# Create your views here.
# Generate random brand code like first letter capital 01 ,02 like for Kohinoor - K01 
def generate_brand_code(brand_name):
    # Get the count of existing brands with the same first letter
    existing_brands_count = brands.objects.filter(brand_name__istartswith=brand_name[0]).count()

    # Format the count with leading zeros and append the first letter as prefix
    brand_code = f"{brand_name[0].upper()}{existing_brands_count + 1:02}"
    return brand_code

from datetime import datetime
@login_required(login_url="")
def brand_setting(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        brand_name = request.POST.get('brand_name')
        # sub_brand=request.POST.get('sub_brand')
        # brand_size = request.POST.get('brand_size')
        brands_added_on = request.POST.get('brands_added_on')
        # pages_per_form = request.POST.get('pages_per_form')

        if brands.objects.filter(brand_name__iexact=brand_name).exists():
            error_message = f"A brand with the name '{brand_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            # return HttpResponseBadRequest(error_message)
        else:
            # brand_code = generate_brand_code()
            brand_code = generate_brand_code(brand_name)
            brands_data = brands(brand_name=brand_name,brand_code=brand_code, brands_added_on=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'))
            brands_data.save()
            sweetify.success(request, "Brand Created Successfully.", timer=6000, button='OK')

    data = brands.objects.all().order_by('-id')
    return render(request, 'brands/brand.html', {'data': data})

@login_required(login_url="") 
def delete_brand(request, id):
    cus = brands.objects.get(id=id)
    cus.delete()
    return redirect('/setting-brands/')

@login_required(login_url="")
def edit_brands(request, id):
    data = brands.objects.get(id=id)
    if request.method == 'POST':
            data.brand_name = request.POST.get('brand_name')
       
            # data.brand_size = request.POST.get('brand_size')
            # data.pages_per_form = request.POST.get('pages_per_form')
            data.save()
           
            return redirect('/setting-brands/')
    
    print("datanew", data.brand_name)
    context = {'data':data,}
    return render(request, 'brands/edit_brand.html', context)




def subbrand_name(request):
    categories = Category.objects.all()
    if request.method == 'POST':
        brand = request.POST.get('brand')
        subbrand = request.POST.get('subbrand')
        subtitle =request.POST.get('subtitle')
        brand_size = request.POST.get('brand_size')
        pages_per_form= request.POST.get('pages_per_form')
        binding_form_size = request.POST.get('binding_form_size')
        printing_form_size= request.POST.get('printing_form_size')
        brand_combine= request.POST.get('brand_combine')

        category_1 = request.POST.get('category_1')
        mainDiscount_1 = request.POST.get('mainDiscount_1')
        extraDiscount_1 = request.POST.get('extraDiscount_1')
        additionalDiscount_1 = request.POST.get('additionalDiscount_1')

        if request.POST.get('category_2')!= '':
            category_2 = request.POST.get('category_2')
            mainDiscount_2 = request.POST.get('mainDiscount_2')
            extraDiscount_2 = request.POST.get('extraDiscount_2')
            additionalDiscount_2 = request.POST.get('additionalDiscount_2')


        if request.POST.get('category_3')!= '':
            category_3 = request.POST.get('category_3')
            mainDiscount_3 = request.POST.get('mainDiscount_3')
            extraDiscount_3 = request.POST.get('extraDiscount_3')
            additionalDiscount_3 = request.POST.get('additionalDiscount_3')


        if request.POST.get('category_4')!= '':
            category_4 = request.POST.get('category_4')
            mainDiscount_4 = request.POST.get('mainDiscount_4')
            extraDiscount_4 = request.POST.get('extraDiscount_4')
            additionalDiscount_4 = request.POST.get('additionalDiscount_4')

        if request.POST.get('category_5')!= '':
            category_5 = request.POST.get('category_5')
            mainDiscount_5 = request.POST.get('mainDiscount_5')
            extraDiscount_5 = request.POST.get('extraDiscount_5')
            additionalDiscount_5 = request.POST.get('additionalDiscount_5')

        if request.POST.get('category_6')!= '':
            category_6 = request.POST.get('category_6')
            mainDiscount_6 = request.POST.get('mainDiscount_6')
            extraDiscount_6 = request.POST.get('extraDiscount_6')
            additionalDiscount_6 = request.POST.get('additionalDiscount_6')

        if request.POST.get('category_7')!= '':
            category_7 = request.POST.get('category_7')
            mainDiscount_7 = request.POST.get('mainDiscount_7')
            extraDiscount_7 = request.POST.get('extraDiscount_7')
            additionalDiscount_7 = request.POST.get('additionalDiscount_7')


        if request.POST.get('category_8')!= '':
            category_8 = request.POST.get('category_8')
            mainDiscount_8 = request.POST.get('mainDiscount_8')
            extraDiscount_8 = request.POST.get('extraDiscount_8')
            additionalDiscount_8 = request.POST.get('additionalDiscount_8')

        if request.POST.get('category_9')!= '':
            category_9 = request.POST.get('category_9')
            mainDiscount_9 = request.POST.get('mainDiscount_9')
            extraDiscount_9 = request.POST.get('extraDiscount_9')
            additionalDiscount_9 = request.POST.get('additionalDiscount_9')

        if request.POST.get('category_10') != '':
            category_10 = request.POST.get('category_10')
            mainDiscount_10 = request.POST.get('mainDiscount_10')
            extraDiscount_10 = request.POST.get('extraDiscount_10')
            additionalDiscount_10 = request.POST.get('additionalDiscount_10')
        
        # Create a new Subbrand object
        subbrand_data = Subbrand(brand_combine=brand_combine,subbrand=subbrand,brand=brand,subtitle=subtitle,brand_size=brand_size,pages_per_form=pages_per_form,printing_form_size=printing_form_size,binding_form_size=binding_form_size,
                                 category_1=category_1, mainDiscount_1=mainDiscount_1, extraDiscount_1=extraDiscount_1, additionalDiscount_1=additionalDiscount_1,
                                 category_2=category_2, mainDiscount_2=mainDiscount_2, extraDiscount_2=extraDiscount_2, additionalDiscount_2=additionalDiscount_2,
                                 category_3=category_3, mainDiscount_3=mainDiscount_3, extraDiscount_3=extraDiscount_3, additionalDiscount_3=additionalDiscount_3,
                                 category_4=category_4, mainDiscount_4=mainDiscount_4, extraDiscount_4=extraDiscount_4, additionalDiscount_4=additionalDiscount_4,
                                 category_5=category_5, mainDiscount_5=mainDiscount_5, extraDiscount_5=extraDiscount_5, additionalDiscount_5=additionalDiscount_5,
                                 category_6=category_6, mainDiscount_6=mainDiscount_6, extraDiscount_6=extraDiscount_6, additionalDiscount_6=additionalDiscount_6,
                                 category_7=category_7, mainDiscount_7=mainDiscount_7, extraDiscount_7=extraDiscount_7, additionalDiscount_7=additionalDiscount_7,
                                 category_8=category_8, mainDiscount_8=mainDiscount_8, extraDiscount_8=extraDiscount_8, additionalDiscount_8=additionalDiscount_8,
                                 category_9=category_9, mainDiscount_9=mainDiscount_9, extraDiscount_9=extraDiscount_9, additionalDiscount_9=additionalDiscount_9,
                                 category_10=category_10, mainDiscount_10=mainDiscount_10, extraDiscount_10=extraDiscount_10, additionalDiscount_10=additionalDiscount_10
                                 )
        subbrand_data.save()  # Call save() as a method to save the object

        # You may want to use a library like sweetify for messages here

    data = Subbrand.objects.all()
    data50=brands.objects.all()
    return render(request, 'brands/subbrand.html', {'data': data,'data50':data50, 'categories':categories})

def edit_subbrand(request,id):
    data = Subbrand.objects.get(id=id)
    if request.method == 'POST':
        data.subbrand = request.POST.get('subbrand')
        data.brand =request.POST.get('brand')
        data.subtitle =request.POST.get('subtitle')
        data.brand_size = request.POST.get('brand_size')
       
        data.binding_form_size = request.POST.get('binding_form_size')
        data.printing_form_size= request.POST.get('printing_form_size')
        

        data.category_1 = request.POST.get('category_1')
        data.mainDiscount_1 = request.POST.get('mainDiscount_1')
        data.extraDiscount_1 = request.POST.get('extraDiscount_1')
        data.additionalDiscount_1 = request.POST.get('additionalDiscount_1')

        data.category_2 = request.POST.get('category_2')
        data.mainDiscount_2 = request.POST.get('mainDiscount_2')
        data.extraDiscount_2 = request.POST.get('extraDiscount_2')
        data.additionalDiscount_2 = request.POST.get('additionalDiscount_2')

        data.category_3 = request.POST.get('category_3')
        data.mainDiscount_3 = request.POST.get('mainDiscount_3')
        data.extraDiscount_3 = request.POST.get('extraDiscount_3')
        data.additionalDiscount_3 = request.POST.get('additionalDiscount_3')

        data.category_4 = request.POST.get('category_4')
        data.mainDiscount_4 = request.POST.get('mainDiscount_4')
        data.extraDiscount_4 = request.POST.get('extraDiscount_4')
        data.additionalDiscount_4 = request.POST.get('additionalDiscount_4')

        data.category_5 = request.POST.get('category_5')
        data.mainDiscount_5 = request.POST.get('mainDiscount_5')
        data.extraDiscount_5 = request.POST.get('extraDiscount_5')
        data.additionalDiscount_5 = request.POST.get('additionalDiscount_5')

        data.category_6 = request.POST.get('category_6')
        data.mainDiscount_6 = request.POST.get('mainDiscount_6')
        data.extraDiscount_6 = request.POST.get('extraDiscount_6')
        data.additionalDiscount_6 = request.POST.get('additionalDiscount_6')

        data.category_7 = request.POST.get('category_7')
        data.mainDiscount_7 = request.POST.get('mainDiscount_7')
        data.extraDiscount_7 = request.POST.get('extraDiscount_7')
        data.additionalDiscount_7 = request.POST.get('additionalDiscount_7')

        data.category_8 = request.POST.get('category_8')
        data.mainDiscount_8 = request.POST.get('mainDiscount_8')
        data.extraDiscount_8 = request.POST.get('extraDiscount_8')
        data.additionalDiscount_8 = request.POST.get('additionalDiscount_8')

        data.category_9 = request.POST.get('category_9')
        data.mainDiscount_9 = request.POST.get('mainDiscount_9')
        data.extraDiscount_9 = request.POST.get('extraDiscount_9')
        data.additionalDiscount_9 = request.POST.get('additionalDiscount_9')

        data.category_10= request.POST.get('category_10')
        data.mainDiscount_10 = request.POST.get('mainDiscount_10')
        data.extraDiscount_10 = request.POST.get('extraDiscount_10')
        data.additionalDiscount_10 = request.POST.get('additionalDiscount_10')
        data.save()
           
        return redirect('/subbrand_name/')
    data1 = brands.objects.values('brand_name').distinct().order_by('brand_name')
    categories = Category.objects.all()
    context = {'data':data,'data1':data1,'categories':categories}
    return render(request, 'brands/edit_subbrand.html', context)


def view_subbrand(request,id):
    data = Subbrand.objects.filter(id=id).first()
    # print('9999999999999999999',data.pages_per_form)
    total_main_discount = 0
    total_extra_discount = 0
    total_additional_discount = 0

    if data:
        # Calculate total main discount
        total_main_discount = sum([float(getattr(data, f"mainDiscount_{i}", 0) or 0) for i in range(1, 11)])

        # Calculate total extra discount
        total_extra_discount = sum([float(getattr(data, f"extraDiscount_{i}", 0) or 0) for i in range(1, 11)])

        # Calculate total additional discount
        total_additional_discount = sum([float(getattr(data, f"additionalDiscount_{i}", 0) or 0) for i in range(1, 11)])
    else:
        total_main_discount = total_extra_discount = total_additional_discount = 0
    context = {
        'data':data,
        'total_main_discount':total_main_discount,
        'total_extra_discount':total_extra_discount,
        'total_additional_discount':total_additional_discount
    }
    return render(request,'brands/subbrand_view.html',context)


def delete_subbrand(request, id):
    cus1 = Subbrand.objects.get(id=id)
    cus1.delete()
    return redirect('/subbrand_name/')

from datetime import datetime
@login_required(login_url="")
def class_setting(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        class_name = request.POST.get('class_name')
        class_book = request.POST.get('class_book')
        classes_added_on = request.POST.get('classes_added_on')
        subject = request.POST.get('subject')

        # Check if the entered class is within the range 5 to 12
        try:
            class_number = int(class_name.split()[0])
            if class_number < 5 or class_number > 12:
                error_message = "Please enter a class between 5 and 12."
                sweetify.error(request, error_message, timer=8000, button='OK')
                data = classes.objects.all().order_by('-id')
                return render(request, 'classes/classes.html', {'data': data})
        except ValueError:
            error_message = "Please enter a valid class number (e.g., 5 std)."
            sweetify.error(request, error_message, timer=8000, button='OK')
            data = classes.objects.all().order_by('-id')
            return render(request, 'classes/classes.html', {'data': data})

        # Check if the class already exists
        if classes.objects.filter(class_name=class_name).exists():
            error_message = f"A class with the name '{class_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            class_data = classes(classes_added_on=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),
                                 class_name=class_name, class_book=class_book, subject=subject)
            class_data.save()
            sweetify.success(request, "Class Created Successfully.", timer=6000)

    data = classes.objects.all().order_by('-id')
    return render(request, 'classes/classes.html', {'data': data})



@login_required(login_url="")
def delete_class(request, id):
    cus = classes.objects.get(id=id)
    cus.delete()
    return redirect('/setting-classes/')

@login_required(login_url="")
def edit_classes(request, id):
    data = classes.objects.get(id=id)
    if request.method == 'POST':
        data.class_name = request.POST.get('class_name')
        data.class_book = request.POST.get('class_book')
        data.medium = request.POST.get('medium')
        if classes.objects.filter(class_name__iexact=data.class_name).exists():
            error_message = f"A class with the name '{data.class_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK',)
            return redirect('/setting-classes/')
        else:
            data.save()
            sweetify.success(request, "Class Created Successfully.", timer=6000)
            return redirect('/setting-classes/')
    
    context = {'data':data,}
    return render(request, 'classes/edit_classes.html', context)




from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font


def excel_export_view(request):
    # Create a new Excel workbook and add a worksheet to it
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add headings to the worksheet
    headings = ["Code", "Brand", "Class", "Subject", "Medium", "Pages", "Binding Size", "Opening Stock Of Book","Printing Rate","Binding Rate","Single Colour Rate","Double Colour Rate","Four Colour Rate","Opening Stock Of Cover","Opening Stock Of Topi"]
    ws.append(headings)
    font = Font(size=14, bold=True, color='000000')  # Adjust size and color as needed
    for cell in ws[1]:
        cell.font = font

    column_widths = [30, 15, 25, 15, 20,30, 20,40 , 20 ,40,40,40,40,40]  # Adjust the widths as needed
    for i, width in enumerate(column_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = width    

    # Create a response with Excel MIME type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Book details.xlsx'

    # Save the workbook to the response
    wb.save(response)
    return response



from datetime import datetime
from .resources import PersonResource
from tablib import Dataset
def books_setting(request):
    # vv = books.objects.all()
    # vv.delete()
    brand1 = Subbrand.objects.all()
    print("dfghjkl;",brand1)
    
    if request.method == 'POST':
        if 'myfile' in request.FILES:
            person_resource = PersonResource()
            dataset = Dataset()
            new_persons = request.FILES['myfile']
            imported_data = dataset.load(new_persons.read(), format='xlsx')

            print("Imported Data:", imported_data)  # Print to check the structure of imported_data

            audit_year_list = []
            for data in imported_data:
                 # Adjust the index based on the number of elements you expect in each row
                    code = data[0]
                    value = books(
                        code = data[0],books_name=data[1],books_class=data[2],subject=data[3],medium=data[4],no_of_pages=data[5],stock=data[6],binding_rate=data[7],onec_rate=data[8],twoc_rate=data[9],fourc_rate=data[10],cover=data[11],topi=data[12],shrink_pack_size=data[13],bitti_pack_size=data[14],books_added_on=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
                        
                    )
                    for i in brand1:
                        if i.brand == value.books_name:
                            no_pages = int(value.no_of_pages)
                            print("777777777777777",no_pages)
                            pages_form = int(i.pages_per_form)
                            book_forms = no_pages / pages_form
                            value.per_book_forms = book_forms
                            value.code = code
                            value.printing_form_size = no_pages / int(i.printing_form_size)
                            print("zxghjk", value.printing_form_size)
                            value.binding_form_size = no_pages / int(i.binding_form_size)
                            print("uuuuuuuuuuuuuu", value.binding_form_size)

                    value.save()
                
   
        else:
            id = request.POST.get('id')
            books_name1 =request.POST.get('books_name')
            
            books_pages =request.POST.get('books_pages')
            books_class =request.POST.get('books_class')
            subject =request.POST.get('subject')
            no_of_pages =request.POST.get('no_of_pages')
            books_added_on = request.POST.get('books_added_on')
            book_size =request.POST.get('book_size')
            per_book_forms =request.POST.get('per_book_forms')
            printing_form_size=request.POST.get('printing_form_size')
            code=request.POST.get('code')
            medium=request.POST.get('medium')
            stock= request.POST.get('stock')
            shrink_pack_size= request.POST.get('shrink_pack_size')
            bitti_pack_size= request.POST.get('bitti_pack_size')
            cover = request.POST.get('cover')
            topi = request.POST.get('topi')
            form1=request.POST.get('form1')
            form2=request.POST.get('form2')
            form3=request.POST.get('form3')
            form4=request.POST.get('form4')
            form5=request.POST.get('form5')
            form6=request.POST.get('form6')
            form7=request.POST.get('form7')
            form8=request.POST.get('form8')
            form9=request.POST.get('form9')
            form10=request.POST.get('form10')
            form11=request.POST.get('form11')
            form12=request.POST.get('form12')
            onec_rate=request.POST.get('onec_rate')
            twoc_rate=request.POST.get('twoc_rate')
            fourc_rate=request.POST.get('fourc_rate')
            binding_rate =request.POST.get('binding_rate')
            color =request.POST.get('color')
            selling_price= request.POST.get('selling_price')
            main_discount = request.POST.get('main_discount')
            print("asdfghjklkjhgfdssdfghjkjhgfdfgh",main_discount)
            extra_discount =request.POST.get('extra_discount')
            subbrand_printing_form_size =request.POST.get('subbrand_printing_form_size')
            subbrand_binding_form_size=request.POST.get('subbrand_binding_form_size')

            print('bvvbvbcbvcbvcvbcbvcb',books_name1)

            # ass_data = books.objects.create(code= code,books_name=books_name1,books_class=books_class,subject=subject,medium=medium)
            # ass_data.save()

            for i in brand1:
                aa = i.brand + '*' + i.subbrand + '*' + i.subtitle
                print('sdddddddddddddddddddddddd',aa)
                if aa == books_name1:
                    no_pages=int(no_of_pages)
                    print("666666666666666666",no_pages)
                    # pages_form=int(i.pages_per_form)
                    # print("666666666666666666",pages_form)
                    # book_forms=no_pages/pages_form
                    # print("666666666666666666",book_forms)

                    # per_book_forms=book_forms
                    # print("666666666666666666",per_book_forms)
                    printing_form_size=no_pages/int(i.printing_form_size)
                    print("ghjkl",printing_form_size)
                    binding_form_size = no_pages/int(i.binding_form_size)
                    print("dfghjk",binding_form_size)
                    if books.objects.filter(code= code,books_name=books_name1,books_class=books_class,subject=subject,medium=medium).exists():
                        error_message = f"A Books with the name '{code}','{books_name1}','{books_class}','{subject}','{medium}' already exists."
                        sweetify.error(request, error_message, timer=8000, button='OK')
                        # return HttpResponseBadRequest(error_message)
                    else:
                        books_data = books(books_added_on=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),main_discount=main_discount,extra_discount=extra_discount,selling_price=selling_price,medium=medium,books_name=books_name1,books_pages=books_pages,books_class=books_class,subject=subject,no_of_pages=no_of_pages,book_size=book_size,code=code,per_book_forms=per_book_forms,binding_form_size=binding_form_size,stock=stock,shrink_pack_size=shrink_pack_size,bitti_pack_size=bitti_pack_size,cover=cover,topi=topi,form1=form1,form2=form2,form3=form3,form4=form4,form5=form5,form6=form6,form7=form7,form8=form8,form9=form9,form10=form10,form11=form11,form12=form12,binding_rate=binding_rate,onec_rate=onec_rate,twoc_rate=twoc_rate,fourc_rate=fourc_rate,color=color,printing_form_size=printing_form_size,subbrand_binding_form_size=subbrand_binding_form_size,subbrand_printing_form_size=subbrand_printing_form_size)
                        books_data.save()
                        print("books_data",books_data.code)
                        sweetify.success(request, "Books created successfully.", timer=6000,button='OK')
        

    # data5=Subbrand.objects.values('printing_form_size').distinct().order_by('printing_form_size')
    # print("iiiiiiiiiiiiiiiiiiiiiii",data5)
    data = books.objects.all().order_by('-id')
    print("data",data)
    ddd = Subbrand.objects.all()
    data1 = Subbrand.objects.values('brand').distinct().order_by('brand')
    
   
    data2 = classes.objects.values('class_name').distinct().order_by('class_name')
    print("45678",data2)
    data3 = mediums.objects.values('medium').distinct().order_by('medium')
    data4 = brands.objects.values('brand_size').distinct().order_by('brand_size')
    per_book_forms=""
    for i in data:
        if i.no_of_pages:

            no_pages=int(i.no_of_pages)
            for j in ddd:
                if i.books_name==j.brand:
                    books_name=i.books_name
                    brand_name=j.brand
                    # pages_form=int(j.pages_per_form)
                    # book_forms=no_pages/pages_form
                    # i.per_book_forms=book_forms
                    
                    # per_book_forms=i.per_book_forms
                    binding_form_size = no_pages/int(j.binding_form_size)
                    printing_form_size = no_pages/int(j.printing_form_size)
                    print("rtyuioyddfghjvcv",printing_form_size)
                    
    
   

    

    context={
         'data':data,
         'data1':data1,
         'data2':data2,
         'data3':data3,
         'data4':data4,
         
         
         'brand1':brand1,
         
    }
    return render(request,'books/books.html',context)




def Courselist(request,id):
    print(id,"==========")
	# courses_name = request.POST['courses_name']
    list1 = []
    filtered_course_list = []
    user_courses_name = brands.objects.filter(brand_name = id)
    print("brand_name",user_courses_name)
    for brand_name in user_courses_name:
        list1.append({
			"c_course_type": brand_name.brand_name,
			})
    for x in list1:
        if x['c_course_type'] not in filtered_course_list:
            if x['c_course_type'] != '':
                filtered_course_list.append(x['c_course_type'])
            filtered_course_list.sort()
    return JsonResponse({"data":filtered_course_list})

def Feeslist(request,id):
    print(id,"==========")
	# courses_name = request.POST['courses_name']
    list1 = []
    filtered_fees_list = []
    fees_name = classes.objects.filter(class_name = id)
    print("class_name",fees_name)
    for class_name in fees_name:
        list1.append({
			"c_fees": class_name.class_name
			})
    for x in list1:
        if x['c_fees'] not in filtered_fees_list:
            if x['c_fees'] != '':
                filtered_fees_list.append(x['c_fees'])
            filtered_fees_list.sort()
            print(filtered_fees_list,'asa')
    return JsonResponse({"data":filtered_fees_list})

def Discountlist(request,id):
    print(id,"==========")
	# courses_name = request.POST['courses_name']
    list1 = []
    filtered_discount_list = []
    discount_name = mediums.objects.filter(medium = id)
    print("course_type",discount_name)
    for medium in discount_name:
        list1.append({
			"c_discount": medium.medium
			})
    for x in list1:
        if x['c_discount'] not in filtered_discount_list:
            if x['c_discount'] != '':
                filtered_discount_list.append(x['c_discount'])
            filtered_discount_list.sort()
            print(filtered_discount_list,'asa')
    return JsonResponse({"data":filtered_discount_list})

def Brandsize(request,id):
    print(id,"==========")
	# courses_name = request.POST['courses_name']
    list1 = []
    filtered_size_list = []
    user_brans_size = brands.objects.filter(brand_size = id)
    print("brand_name",user_brans_size)
    for brand_size in user_brans_size:
        list1.append({
			"c_brand_size": brand_size.brand_size,
			})
    for x in list1:
        if x['c_brand_size'] not in filtered_size_list:
            if x['c_brand_size'] != '':
                filtered_size_list.append(x['c_brand_size'])
            filtered_size_list.sort()
    return JsonResponse({"data":filtered_size_list})

@login_required(login_url="")
def delete_books(request, id):
    cus = books.objects.get(id=id)
    cus.delete()
    return redirect('/setting-books/')

@login_required(login_url="")
def edit_books(request, id):
    data = books.objects.get(id=id)
    if request.method == 'POST':
        data.code = request.POST.get('code')
        data.books_name = request.POST.get('books_name')
        data.books_class = request.POST.get('books_class')
        data.shrink_pack_size =  request.POST.get('shrink_pack_size')
        data.medium = request.POST.get('medium')
        data.subject = request.POST.get('subject')
        data.color = request.POST.get('color')
        data.bitti_pack_size = request.POST.get('bitti_pack_size')
        data.binding_rate = request.POST.get('binding_rate')
        data.onec_rate = request.POST.get('onec_rate')
        data.twoc_rate = request.POST.get('twoc_rate')
        data.fourc_rate = request.POST.get('fourc_rate')
        # data.printing_rate1 = request.POST.get('printing_rate1')
        data.binding_form_size=request.POST.get('binding_form_size')
        print("567890",data.binding_form_size)
        # data.printing_form_size=request.POST.get('prining_form_size')
        # print("567890-",data.printing_form_size)
        data.no_of_pages = request.POST.get('no_of_pages')

 
        data.cover = request.POST.get('cover')
        data.topi = request.POST.get('topi')
        data.form1=request.POST.get('form1')
        data.form2=request.POST.get('form2')
        data.form3=request.POST.get('form3')
        data.form4=request.POST.get('form4')
        data.form5=request.POST.get('form5')
        data.form6=request.POST.get('form6')
        data.form7=request.POST.get('form7')
        data.form8=request.POST.get('form8')
        data.form9=request.POST.get('form9')
        data.form10=request.POST.get('form10')
        data.form11=request.POST.get('form11')
        data.form12=request.POST.get('form12')
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/setting-books/')
    data3 = mediums.objects.values('medium').distinct().order_by('medium')
    data2 = classes.objects.values('class_name').distinct().order_by('class_name')
    data1 = Subbrand.objects.all()
    
    context = {'data':data,'data3':data3 ,'data2':data2,'data1':data1,}
    return render(request, 'books/edit_books.html', context)    
      

from datetime import datetime
@login_required(login_url="")
def form_setting(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        rmname = request.POST.get('rmname')
        rmalias = request.POST.get('rmalias')
        print_name = request.POST.get('print_name')
        qunit = request.POST.get('qunit')
        opening_stock_value = request.POST.get('opening_stock_value')
        opening_stock_rm = request.POST.get('opening_stock_rm')
        gsm = request.POST.get('gsm')
        p_width = request.POST.get('p_width')
        p_length = request.POST.get('p_length')
        unit1 = request.POST.get('unit1')
        qunit1 = request.POST.get('qunit1')
        unit2 = request.POST.get('unit2')
        qunit2 = request.POST.get('qunit2')
        cunit4 = request.POST.get('cunit4')
        unit3 = request.POST.get('unit3')
        print("333333333333333",unit3)
        qunit3 = request.POST.get('qunit3')
        print("yyyyyyyyyyyyyy",qunit3)
        unit1_conversion = request.POST.get('unit1_conversion')
        cunit1 = request.POST.get('cunit1')
        unit2_conversion = request.POST.get('unit2_conversion')
        cunit2 = request.POST.get('cunit2')
        Waste = request.POST.get('Waste')
        waste_percent = request.POST.get('waste_percent')
        # waste_percent=float(waste_percent)/100
        raw_material_center=request.POST.get('raw_material_center')
        unit4=request.POST.get('unit4')
        unit3_conversion = request.POST.get('unit3_conversion')
        cunit3= request.POST.get('cunit3')
        print("yyyyyyyyyyyyyy",cunit3)
        opening_stock1= request.POST.get('opening_stock1')
        
        conversion_factor1= request.POST.get('conversion_factor1')
        
        




        # Rest of your code
        books_data = forms(
            forms_added_on=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),
            rmname=rmname,
            rmalias=rmalias,
            print_name=print_name,
            qunit=qunit,
            opening_stock_value=opening_stock_value,
            gsm=gsm,
            p_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),
            p_width=p_width,
            p_length=p_length,
            unit1=unit1,
            qunit1=qunit1,
            unit2=unit2,
            qunit2=qunit2,
            unit3=unit3,
            qunit3=qunit3,
            opening_stock_rm=opening_stock_rm,
            unit1_conversion=unit1_conversion,
            cunit1=cunit1,
            unit2_conversion=unit2_conversion,
            cunit2=cunit2,
            Waste=Waste,
            unit4=unit4,
            cunit3=cunit3,
            cunit4=cunit4,
            unit3_conversion=unit3_conversion,
            waste_percent=waste_percent,
            raw_material_center=raw_material_center,
            opening_stock1=opening_stock1,
            
            conversion_factor1=conversion_factor1,
            

        )
        books_data.save()

    data = forms.objects.all()
    data1=master_unit.objects.all()
    center = material_centre.objects.all()
    return render(request, 'form/forms.html', {'data': data, 'center': center,'data1':data1,})

@login_required(login_url="")
def edit_form(request,id):
    data=forms.objects.get(id=id)
    if request.method=='POST':
        id=request.POST.get('id')
        data.rmname =request.POST.get('rmname')
        data.rmalias =request.POST.get('rmalias')
        data.print_name =request.POST.get('print_name')
        data.opening_stock_value =request.POST.get('opening_stock_value')
        data.opening_stock_rm =request.POST.get('opening_stock_rm')
        data.unit1_conversion = request.POST.get('unit1_conversion')
        data.unit2_conversion = request.POST.get('unit2_conversion')
        data.unit2 = request.POST.get('unit2')
        data.unit3 = request.POST.get('unit3')
        
        data.p_width = request.POST.get('p_width')
        data.p_length = request.POST.get('p_length')
        data.p_raw_material =request.POST.get('p_raw_material')
        data.purchase_quantity=request.POST.get('purchase_quantity')
        data.gsm_paper=request.POST.get('gsm_paper')
        data.gsm=request.POST.get('gsm')
        data.Waste = request.POST.get('Waste')
        data.waste_percent = request.POST.get('waste_percent')
        data.qunit=request.POST.get('qunit')
        data.cunit1=request.POST.get('cunit1')
        data.cunit2=request.POST.get('cunit2')
        data.qunit2=request.POST.get('qunit2')
        data.qunit3=request.POST.get('qunit3')
        data.raw_material_center=request.POST.get('raw_material_center')
        data.unit4=request.POST.get('unit4')
        data.unit3_conversion = request.POST.get('unit3_conversion')
        data.cunit3= request.POST.get('cunit3')
        data.cunit4= request.POST.get('cunit4')
        data.cunit4= request.POST.get('cunit4')
        data.conversion_factor1= request.POST.get('conversion_factor1')
        data.opening_stock1 = request.POST.get('opening_stock1')

        
        data.save()
        messages.success(request,'Updated Successfully...!')
        return redirect('/forms-setting/')
    data1= material_centre.objects.all()
    data2=master_unit.objects.all()
    context={'data':data,'data1':data1,'data2':data2,}
    return render (request,'form/edit_form.html',context)



@login_required(login_url="")
def delete_form(request, id):
    cus = forms.objects.get(id=id)
    cus.delete()
    return redirect('/forms-setting/')

from datetime import datetime
@login_required(login_url="")
def medium_setting(request):
    if request.method =='POST':
        id = request.POST.get('id')
        medium = request.POST.get('medium')
        medium_added_on = request.POST.get('medium_added_on') 
        if mediums.objects.filter(medium__iexact=medium).exists():
            error_message = f"A medium with the name '{medium}' already exists."
            sweetify.error(request, error_message, timer=2000, button='OK')
            # return HttpResponseBadRequest(error_message)
        else:
            medium_data = mediums(medium_added_on=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'), medium=medium)
            medium_data.save()
            sweetify.success(request, "Medium Created Successfully.", timer=6000)

    data = mediums.objects.all().order_by('-id')
    return render (request,'classes/medium.html',{'data':data,})


def edit_medium(request, id):
    data = mediums.objects.get(id=id)
    if request.method == 'POST':
       
        data.medium = request.POST.get('medium')
        
        # Check if the class name already exists in the database
        existing_medium = mediums.objects.filter(medium__iexact=data.medium).exclude(id=id).first()
        if existing_medium:
            messages.error(request, "medium with the same name already exists!")
            return redirect('/medium-setting/')
        else:
            # Update the record
           
            # data.medium = medium
            data.save()
            messages.success(request, "Updated Successfully...!!")
            return redirect('/medium-setting/')
    
    context = {'data':data,}
    return render(request, 'classes/edit_medium.html', context)

def delete_medium(request, id):
    mid = mediums.objects.get(id=id)
    mid.delete()
    return redirect('/medium-setting/')




@login_required(login_url="")
def company(request):
    if request.method =='POST':
        id = request.POST.get('id')
        company_type=request.POST.get('company_type')
        company_name = request.POST.get('company_name')
        company_contact = request.POST.get('company_contact')
        company_email = request.POST.get('company_email')
        gstin_no =request.POST.get('gstin_no')
        company_address = request.POST.get('company_address')
        company_created_by = request.POST.get('company_created_by')
        state = request.POST.get('state')
        country = request.POST.get('country')
        pincode = request.POST.get('pincode') 
        website = request.POST.get('website')
        if Company.objects.filter(company_name__iexact=company_name).exists():
            error_message = f"A Company with the name='{company_name}' already exists."
            sweetify.error(request, error_message, timer=2000, button='OK')
            # return HttpResponseBadRequest(error_message)
        else:
            
            vender_data = Company(company_type=company_type,company_name=company_name,company_contact=company_contact,company_email=company_email,gstin_no=gstin_no,company_address=company_address,state=state,pincode=pincode,country=country,website=website,company_created_by=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'))
            vender_data.save()           
            sweetify.success(request, "Company Created Successfully.", timer=6000 ,button='OK')
                    
    data = Company.objects.all().order_by('-id')
    grp = Group.objects.all()
    context = {
        'data':data,
        'grp':grp,
    }
    return render (request,'company/company.html',context)

@login_required(login_url="")
def edit_company(request, id):
    data = Company.objects.get(id=id)
    if request.method == 'POST':
        data.company_type=request.POST.get('company_type')
        data.company_name = request.POST.get('company_name')
        data.company_contact = request.POST.get('company_contact')
        data.company_email = request.POST.get('company_email')
        data.gstin_no = request.POST.get('gstin_no')
        data.company_address = request.POST.get('company_address')
        data.state = request.POST.get('state')
        data.country = request.POST.get('country')
        data.pincode = request.POST.get('pincode')
        data.website = request.POST.get('website')
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/company/')
    
    context = {'data':data,}
    return render(request, 'company/edit_company.html', context) 

@login_required(login_url="")
def delete_company(request, id):
    cus = Company.objects.get(id=id)
    cus.delete()
    return redirect('/company/')

# def unit_conversion(request):
#     pass


@login_required(login_url="")
def binding_vender(request):
    if request.method =='POST':
        id = request.POST.get('id')
        organization_name1 = request.POST.get('organization_name1')
        vender_contact1 = request.POST.get('vender_contact1')
        vender_email1 = request.POST.get('vender_email1')
        vender_gstin_no1 =request.POST.get('vender_gstin_no1')
        vender_address1 = request.POST.get('vender_address1')
        vender_created_by1 = request.POST.get('vender_created_by1')
        vender_group1 = request.POST.get('vender_group1')
        if Bindingvender.objects.filter(organization_name1=organization_name1).exists():
            error_message = f"A venders with the name='{organization_name1}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            vender_data = Bindingvender(vender_created_by1=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),organization_name1=organization_name1,vender_contact1=vender_contact1,vender_email1=vender_email1,vender_gstin_no1=vender_gstin_no1,vender_address1=vender_address1,vender_group1=vender_group1)
            vender_data.save()           
            sweetify.success(request, "Venders Created Successfully.", timer=6000,button='OK')
                    
    data = Bindingvender.objects.all().order_by('-id')
    grp = Group.objects.all()
    context = {
        'data':data,
        'grp':grp,
    }
    return render (request,'bindingvenders.html',context)




@login_required(login_url="")
def gathering_vender(request):
    if request.method =='POST':
        id = request.POST.get('id')
        gorganization_name1 = request.POST.get('gorganization_name1')
        gvender_contact1 = request.POST.get('gvender_contact1')
        gvender_email1 = request.POST.get('gvender_email1')
        gvender_gstin_no1 =request.POST.get('gvender_gstin_no1')
        gvender_address1 = request.POST.get('gvender_address1')
        gvender_created_by1 = request.POST.get('gvender_created_by1')
        gvender_group1 = request.POST.get('gvender_group1')
        if Gatheringvender.objects.filter(gorganization_name1=gorganization_name1).exists():
            error_message = f"A venders with the name='{gorganization_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            vender_data = Gatheringvender(gvender_created_by1=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),gorganization_name1=gorganization_name1,gvender_contact1=gvender_contact1,gvender_email1=gvender_email1,gvender_gstin_no1=gvender_gstin_no1,gvender_address1=gvender_address1,gvender_group1=gvender_group1)
            vender_data.save()           
            sweetify.success(request, "Venders Created Successfully.", timer=6000,button='OK')
                    
    data = Gatheringvender.objects.all().order_by('-id')
    grp = Group.objects.all()
    context = {
        'data':data,
        'grp':grp,
    }
    return render (request,'gatheringvenders.html',context)



@login_required(login_url="")
def edit_gatheringvenders(request, id):
    data = Gatheringvender.objects.get(id=id)
    if request.method == 'POST':
        data.gorganization_name1 = request.POST.get('gorganization_name1')
        data.gvender_contact1 = request.POST.get('gvender_contact1')
        data.gvender_email1 = request.POST.get('gvender_email1')
        data.gvender_gstin_no1 = request.POST.get('gvender_gstin_no1')
        data.gvender_address1 = request.POST.get('gvender_address1')
        data.gvender_group1 = request.POST.get('gvender_group1')
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/gathering_vender/')
    
    context = {'data':data,}
    return render(request, 'edit_gatheringvenders.html', context) 



@login_required(login_url="")
def cutting_vender(request):
    if request.method =='POST':
        id = request.POST.get('id')
        corganization_name1 = request.POST.get('corganization_name1')
        cvender_contact1 = request.POST.get('cvender_contact1')
        cvender_email1 = request.POST.get('cvender_email1')
        cvender_gstin_no1 =request.POST.get('cvender_gstin_no1')
        cvender_address1 = request.POST.get('cvender_address1')
        cvender_created_by1 = request.POST.get('cvender_created_by1')
        cvender_group1 = request.POST.get('cvender_group1')
        if Cuttingvender.objects.filter(corganization_name1=corganization_name1).exists():
            error_message = f"A venders with the name='{corganization_name1}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            vender_data = Cuttingvender(
                cvender_created_by1=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),
                corganization_name1=corganization_name1,
                cvender_contact1=cvender_contact1,
                cvender_email1=cvender_email1,
                cvender_gstin_no1=cvender_gstin_no1,
                cvender_address1=cvender_address1,
                cvender_group1=cvender_group1
             )
            vender_data.save()          
            sweetify.success(request, "Venders Created Successfully.", timer=6000,button='OK')
                    
    data = Cuttingvender.objects.all().order_by('-id')
    grp = Group.objects.all()
    context = {
        'data':data,
        'grp':grp,
    }
    return render (request,'vendors/cuttingvenders.html',context)




@login_required(login_url="")
def packing_vender(request):
    if request.method =='POST':
        id = request.POST.get('id')
        porganization_name1 = request.POST.get('porganization_name1')
        pvender_contact1 = request.POST.get('pvender_contact1')
        pvender_email1 = request.POST.get('pvender_email1')
        pvender_gstin_no1 =request.POST.get('pvender_gstin_no1')
        pvender_address1 = request.POST.get('pvender_address1')
        pvender_created_by1 = request.POST.get('pvender_created_by1')
        pvender_group1 = request.POST.get('pvender_group1')
        if Packingvender.objects.filter(porganization_name1=porganization_name1).exists():
            error_message = f"A venders with the name='{porganization_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            vender_data = Packingvender(
                pvender_created_by1=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),
                porganization_name1=porganization_name1,
                pvender_contact1=pvender_contact1,
                pvender_email1=pvender_email1,
                pvender_gstin_no1=pvender_gstin_no1,
                pvender_address1=pvender_address1,
                pvender_group1=pvender_group1
             )
            vender_data.save()          
            sweetify.success(request, "Venders Created Successfully.", timer=6000,button='OK')
                    
    data = Packingvender.objects.all().order_by('-id')
    grp = Group.objects.all()
    context = {
        'data':data,
        'grp':grp,
    }
    return render (request,'vendors/packingvenders.html',context)






@login_required(login_url="")
def edit_bindingvender(request, id):
    data = Bindingvender.objects.get(id=id)
    if request.method == 'POST':
        data.organization_name1 = request.POST.get('organization_name1')
        data.vender_contact1 = request.POST.get('vender_contact1')
        data.vender_email1 = request.POST.get('vender_email1')
        data.vender_gstin_no1 = request.POST.get('vender_gstin_no1')
        data.vender_address1 = request.POST.get('vender_address1')
        data.vender_group1 = request.POST.get('vender_group1')
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/binding_vender/')
    
    context = {'data':data,}
    return render(request, 'edit_bindingvender.html', context) 

@login_required(login_url="")
def delete_bindingvender(request, id):
    cus = Bindingvender.objects.get(id=id)
    cus.delete()
    return redirect('/binding_vender/')

@login_required(login_url="")
def delete_packingvenders(request, id):
    cus = Packingvender.objects.get(id=id)
    cus.delete()
    return redirect('/packing_vender/')



@login_required(login_url="")
def edit_packingvenders(request, id):
    data = Packingvender.objects.get(id=id)
    if request.method == 'POST':
        data.porganization_name1 = request.POST.get('porganization_name1')
        data.pvender_contact1 = request.POST.get('pvender_contact1')
        data.pvender_email1 = request.POST.get('pvender_email1')
        data.pvender_gstin_no1 = request.POST.get('pvender_gstin_no1')
        data.pvender_address1 = request.POST.get('pvender_address1')
        data.pvender_group1 = request.POST.get('pvender_group1')

        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/packing_vender/')
    
    context = {'data':data,}
    return render(request, 'vendors/edit_packingvender.html', context) 


@login_required(login_url="")
def edit_cuttingvenders(request, id):
    data = Cuttingvender.objects.get(id=id)
    if request.method == 'POST':
        data.corganization_name1 = request.POST.get('corganization_name1')
        data.cvender_contact1 = request.POST.get('cvender_contact1')
        data.cvender_email1 = request.POST.get('cvender_email1')
        data.cvender_gstin_no1 = request.POST.get('cvender_gstin_no1')
        data.cvender_address1 = request.POST.get('cvender_address1')
        data.cvender_group1 = request.POST.get('cvender_group1')
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/cutting_vender/')

    
    context = {'data':data,}
    return render(request, 'vendors/edit_cuttingvender.html', context) 


@login_required(login_url="")
def delete_cuttingvenders(request, id):
    cus = Cuttingvender.objects.get(id=id)
    cus.delete()
    return redirect('/cutting_vender/')

@login_required(login_url="")
def delete_gatheringvenders(request, id):
    cus = Gatheringvender.objects.get(id=id)
    cus.delete()
    return redirect('/gathering_vender/')


@login_required(login_url="")
def cover_vender(request):
    if request.method =='POST':
        id = request.POST.get('id')
        c_organization_name = request.POST.get('c_organization_name')
        c_vender_contact = request.POST.get('c_vender_contact')
        c_vender_email = request.POST.get('c_vender_email')
        c_vender_gstin_no =request.POST.get('c_vender_gstin_no')
        c_vender_address = request.POST.get('c_vender_address')
        c_vender_created_by = request.POST.get('c_vender_created_by')
        c_vender_group = request.POST.get('c_vender_group')
        if Covervender.objects.filter(c_organization_name=c_organization_name).exists():
            error_message = f"A venders with the name='{c_organization_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            vender_data = Covervender(c_vender_created_by=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),c_organization_name=c_organization_name,c_vender_contact=c_vender_contact,c_vender_email=c_vender_email,c_vender_gstin_no=c_vender_gstin_no,c_vender_address=c_vender_address,c_vender_group=c_vender_group)
            vender_data.save()           
            sweetify.success(request, "Venders Created Successfully.", timer=6000,button='OK')
                    
    data = Covervender.objects.all().order_by('-id')
    grp = Group.objects.all()
    context = {
        'data':data,
        'grp':grp,
    }
    return render (request,'cover_vender.html',context)

@login_required(login_url="")
def edit_cover_vender(request, id):
    data = Covervender.objects.get(id=id)
    if request.method == 'POST':
        data.c_organization_name = request.POST.get('c_organization_name')
        data.c_vender_contact = request.POST.get('c_vender_contact')
        data.c_vender_email = request.POST.get('c_vender_email')
        data.c_vender_gstin_no = request.POST.get('c_vender_gstin_no')
        data.c_vender_address = request.POST.get('c_vender_address')
        data.c_vender_group = request.POST.get('c_vender_group')
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/cover_vender/')
    
    context = {'data':data,}
    return render(request, 'edit_cover.html', context) 

@login_required(login_url="")
def delete_cover_vender(request, id):
    cus = Covervender.objects.get(id=id)
    cus.delete()
    return redirect('/cover_vender/')    
    

from datetime import datetime
@login_required(login_url="")
def vender_setting(request):
    if request.method =='POST':
        id = request.POST.get('id')
        organization_name = request.POST.get('organization_name')
        vender_contact = request.POST.get('vender_contact')
        vender_email = request.POST.get('vender_email')
        vender_gstin_no =request.POST.get('vender_gstin_no')
        vender_address = request.POST.get('vender_address')
        vender_created_by = request.POST.get('vender_created_by')
        vender_group = request.POST.get('vender_group')
        if venders.objects.filter(organization_name=organization_name).exists():
            error_message = f"A venders with the name='{organization_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            vender_data = venders(vender_created_by=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),organization_name=organization_name,vender_contact=vender_contact,vender_email=vender_email,vender_gstin_no=vender_gstin_no,vender_address=vender_address,vender_group=vender_group)
            vender_data.save()           
            sweetify.success(request, "Vendor Created Successfully.", timer=6000,button='OK')
                    
    data = venders.objects.all().order_by('-id')
    grp = Group.objects.all()
    data1=Main_group.objects.all()
    return render (request,'venders.html',{'data':data,'grp':grp,'data1':data1})

@login_required(login_url="")
def edit_venders(request, id):
    data = venders.objects.get(id=id)
    if request.method == 'POST':
        data.organization_name = request.POST.get('organization_name')
        data.vender_contact = request.POST.get('vender_contact')
        data.vender_email = request.POST.get('vender_email')
        data.vender_gstin_no = request.POST.get('vender_gstin_no')
        data.vender_address = request.POST.get('vender_address')
        data.vender_group=request.POST.get('vender_group')
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/venders-setting/')
    data1=Main_group.objects.all()
    context = {'data':data,'data1':data1,}
    return render(request, 'edit_venders.html', context) 

@login_required(login_url="")
def delete_venders(request, id):
    cus = venders.objects.get(id=id)
    cus.delete()
    return redirect('/venders-setting/')






@login_required(login_url="")
def centre_setting(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        centre_name =request.POST.get('centre_name')
        alias =request.POST.get('alias')
        print_name = request.POST.get('print_name')
        group=request.POST.get('group')
        if material_centre.objects.filter(centre_name__iexact=centre_name).exists():
            error_message = f"A Centre with the name '{centre_name}' already exists."
            sweetify.error(request, error_message)
            # return HttpResponseBadRequest()
        else:
            centre_code = secrets.token_hex(3).upper()
            centre_data = material_centre(centre_name=centre_name,alias=alias, print_name=print_name, group=group)
            centre_data.save()
            sweetify.success(request, "Centre Created Successfully.", timer=6000, button='OK')
    data = material_centre.objects.all().order_by('-id')
    vend = venders.objects.all()
    group1=Group.objects.all()
    # center=forms.objects.all()

    context = {
    'vend':vend,
    'data':data,
    'group1':group1
    
    }
    return render(request, 'material_centre/material_centre.html',context)


@login_required(login_url="") 
def delete_centre(request, id):
    cus = material_centre.objects.get(id=id)
    cus.delete()
    return redirect('/centre-setting/')     



@login_required(login_url="")
def edit_centre(request, id):
    data = material_centre.objects.get(id=id)
    if request.method == 'POST':
        data.centre_name = request.POST.get('centre_name')
        data.alias = request.POST.get('alias')
        data.print_name = request.POST.get('print_name')
        data.group = request.POST.get('group')
        # if material_centre.objects.filter(centre_name__iexact=data.centre_name).exists():
        #     error_message = f"A centre with the name '{data.centre_name}' already exists."
        #     # return HttpResponseBadRequest(error_message)
        # else:
        #     data.save()
        #     messages.success(request, "Updated Successfully...!!")
        data.save()
        return redirect('/centre-setting/')
    
    context = {'data':data,}
    return render(request, 'material_centre/edit_centre.html', context)

from datetime import datetime
@login_required(login_url="")
def distributor(request):
    if request.method =='POST':
        id = request.POST.get('id')
        distributor_name = request.POST.get('distributor_name')
        distributor_contact = request.POST.get('distributor_contact')
        distributor_email = request.POST.get('distributor_email')
        distributor_gstin_no =request.POST.get('distributor_gstin_no')
        distributor_address = request.POST.get('distributor_address')
        distributor_created_by = request.POST.get('distributor_created_by')
        distributor_group = request.POST.get('distributor_group')
        
        if Distributors.objects.filter(distributor_name=distributor_name).exists():
          error_message = f"A Distributor with the name='{distributor_name}' already exists."
          sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            distributor_data = Distributors(distributor_created_by=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),distributor_name=distributor_name,distributor_contact=distributor_contact,distributor_email=distributor_email,distributor_gstin_no=distributor_gstin_no,distributor_address=distributor_address,distributor_group=distributor_group)
            distributor_data.save()           
            sweetify.success(request, "Distributor Created Successfully.", timer=6000, button='OK')
                    
    data = Distributors.objects.all().order_by('-id')
    grp = Group.objects.all()
    return render (request,'distributor.html',{'data':data,'grp':grp})


@login_required(login_url="") 
def delete_distributor(request, id):
    cus = Distributors.objects.get(id=id)
    cus.delete()
    return redirect('/distributor/')     



@login_required(login_url="")
def edit_distributor(request, id):
    data = Distributors.objects.get(id=id)
    if request.method == 'POST':
        # data.group_name = request.POST.get('group_name')
        data.distributor_name = request.POST.get('distributor_name')
        data.distributor_contact  = request.POST.get('distributor_contact')
        print('mobile',data.distributor_contact)
        data.distributor_email  = request.POST.get('distributor_email')
        print('email',data.distributor_email)
        data.distributor_gstin_no= request.POST.get('distributor_gstin_no')
        data.distributor_address = request.POST.get('distributor_address')
        # data.distributor_created_by = request.POST.get('distributor_created_by')
        # print('date',data.distributor_created_by)
        data.distributor_group  = request.POST.get('distributor_group')
        # if Group.objects.filter(group_type=data.group_type).exists():
        #     error_message = f"A centre with the name '{data.group_type}' already exists."
        #     return HttpResponseBadRequest(error_message)
        # else:
        #     data.save()
        #     messages.success(request, "Updated Successfully...!!")
        #     return redirect('/group/')
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/distributor/')
    context = {'data':data,}
    return render(request, 'edit_distributor.html', context)




def group(request):
    if request.method =='POST':
        id = request.POST.get('id')
        group_type = request.POST.get('group_type')
        main_discount = request.POST.get('main_discount')
        extra_discount = request.POST.get('extra_discount')
        if Group.objects.filter(group_type=group_type).exists():
            error_message = f"A Group with the type='{group_type}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            distributor_data = Group(group_type=group_type,main_discount=main_discount,extra_discount=extra_discount)
            distributor_data.save()           
            sweetify.success(request, "Group Created Successfully.", timer=6000 ,button='OK')
                    
    data = Group.objects.all()

    context= {
        'data':data,
    }
    return render(request,'group/group.html',context) 



@login_required(login_url="") 
def delete_group(request, id):
    cus = Group.objects.get(id=id)
    cus.delete()
    return redirect('/group/')     



@login_required(login_url="")
def edit_group(request, id):
    data = Group.objects.get(id=id)
    if request.method == 'POST':
        # data.group_name = request.POST.get('group_name')
        data.group_type = request.POST.get('group_type')
        data.main_discount = request.POST.get('main_discount')
        data.extra_discount = request.POST.get('extra_discount')
        # if Group.objects.filter(group_type=data.group_type).exists():
        #     error_message = f"A centre with the name '{data.group_type}' already exists."
        #     return HttpResponseBadRequest(error_message)
        # else:
        #     data.save()
        #     messages.success(request, "Updated Successfully...!!")
        #     return redirect('/group/')
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/group/')
    context = {'data':data,}
    return render(request, 'group/edit_group.html', context)






from brand_app.models import *
from django.utils import timezone
from django.db.models import Q
from order_app.models import *
from datetime import datetime
def printing_bookwise(request):
     total_received_forms = ""
     total_wastage_forms= ""
     total_quantities = {}
     
     if request.method == 'POST':
         p_vender_name=request.POST.get("p_vender_name")
         print('name',p_vender_name)
       
         p_brand_name_1=request.POST.get("p_brand_name_1")
         p_class_name_1=request.POST.get("p_class_name_1")
         p_medium_1=request.POST.get("p_medium_1")
         p_subject_1=request.POST.get("p_subject_1")
         p_quantity_1=request.POST.get("p_quantity_1")
         p_rim_1=request.POST.get("p_rim_1")
         p_pages_1=request.POST.get("p_pages_1")
         printing_received_form_1=request.POST.get("p_printing_received_form_1")
         printing_received_form_2=request.POST.get("p_printing_received_form_2")
         printing_received_form_3=request.POST.get("p_printing_received_form_3")
         printing_received_form_4=request.POST.get("p_printing_received_form_4")
         printing_received_form_5=request.POST.get("p_printing_received_form_5")
         printing_received_form_6=request.POST.get("p_printing_received_form_6")
         printing_received_form_7=request.POST.get("p_printing_received_form_7")
         printing_received_form_8=request.POST.get("p_printing_received_form_8")
         printing_received_form_9=request.POST.get("p_printing_received_form_9")
         printing_received_form_10=request.POST.get("p_printing_received_form_10")
         print_wastage_form_1=request.POST.get("p_print_wastage_form_1")
         print_wastage_form_2=request.POST.get("p_print_wastage_form_2")
         print_wastage_form_3=request.POST.get("p_print_wastage_form_3")
         print_wastage_form_4=request.POST.get("p_print_wastage_form_4")
         print_wastage_form_5=request.POST.get("p_print_wastage_form_5")
         print_wastage_form_6=request.POST.get("p_print_wastage_form_6")
         print_wastage_form_7=request.POST.get("p_print_wastage_form_7")
         print_wastage_form_8=request.POST.get("p_print_wastage_form_8")
         print_wastage_form_9=request.POST.get("p_print_wastage_form_9")
         print_wastage_form_10=request.POST.get("p_print_wastage_form_10")
         p_single_1=request.POST.get("p_single_1")
         p_double_1=request.POST.get("p_double_1")
         p_four_1=request.POST.get("p_four_1")
         p_code=request.POST.get("p_code")
         p_print_date = datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
         total_received_forms = request.POST.get('total_received_forms')
         print('ljgffg',total_received_forms)
         
          # Calculate the total sum of p_printing_received_form_1 to p_printing_received_form_10
         received_form_values = [
            float(request.POST.get(f'p_printing_received_form_{i}', '0')) for i in range(1, 11)
         ]
         total_received_forms = sum(received_form_values)
         print('total',total_received_forms)

         
         existing_entries = printing.objects.filter(
            p_brand_name_1=p_brand_name_1,
            p_class_name_1=p_class_name_1,
            p_medium_1=p_medium_1,
            p_subject_1=p_subject_1,
            # m1_no=m1_no,
            # a1_name=a1_name,
            p_pages_1=p_pages_1,
            # p_print_date = p_print_date,
         )


         # Extract the values of additional boxes
         additional_box_values = [
            int(request.POST.get(f'p_additional_box_{i}', '0'))
            for i in range(1, 11)
         ]

         # Calculate the total received forms, including the additional boxes
         received_form_values = [
            int(request.POST.get(f'p_printing_received_form_{i}', '0')) + additional_box_values[i - 1]
            for i in range(1, 11)
         ]
         total_received_forms = sum(received_form_values)

         new_entry_needed = True
         for entry in existing_entries:
            if entry.p_vender_name == p_vender_name:

                # Update existing entry
                for i in range(1, 11):
                    entry_field_name = f'p_printing_received_form_{i}'
                    new_value = int(getattr(entry, entry_field_name, 0)) + received_form_values[i - 1]
                    setattr(entry, entry_field_name, new_value)
                entry.p_vender_name= p_vender_name
                entry.p_code=p_code
                # entry.p_printing_received_form_1 =  int(entry.p_printing_received_form_1) + int(printing_received_form_1)
                # entry.p_printing_received_form_2 = int(entry.p_printing_received_form_2) + int(printing_received_form_2)
                # entry.p_printing_received_form_3 = int(entry.p_printing_received_form_3) + int(printing_received_form_3)
                # entry.p_printing_received_form_4 = int(entry.p_printing_received_form_4) + int(printing_received_form_4)
                # entry.p_printing_received_form_5 = int(entry.p_printing_received_form_5) + int(printing_received_form_5)
                # entry.p_printing_received_form_6 = int(entry.p_printing_received_form_6) + int(printing_received_form_6)
                # entry.p_printing_received_form_7 = int(entry.p_printing_received_form_7) + int(printing_received_form_7)
                # entry.p_printing_received_form_8 = int(entry.p_printing_received_form_8) + int(printing_received_form_8)
                # entry.p_printing_received_form_9 = int(entry.p_printing_received_form_9) + int(printing_received_form_9)
                # entry.p_printing_received_form_10 = int(entry.p_printing_received_form_10) + int(printing_received_form_10)
                # existing_entry.p_print_wastage_form_1 = int(existing_entry.p_print_wastage_form_1) + int(print_wastage_form_1)
                # existing_entry.p_print_wastage_form_2 = int(existing_entry.p_print_wastage_form_2) + int(print_wastage_form_2)
                # existing_entry.p_print_wastage_form_3 = int(existing_entry.p_print_wastage_form_3) + int(print_wastage_form_3)
                # existing_entry.p_print_wastage_form_4 = int(existing_entry.p_print_wastage_form_4) + int(print_wastage_form_4)
                # existing_entry.p_print_wastage_form_5 = int(existing_entry.p_print_wastage_form_5) + int(print_wastage_form_5)
                # existing_entry.p_print_wastage_form_6 = int(existing_entry.p_print_wastage_form_6) + int(print_wastage_form_6)
                # existing_entry.p_print_wastage_form_7 = int(existing_entry.p_print_wastage_form_7) + int(print_wastage_form_7)
                # existing_entry.p_print_wastage_form_8 = int(existing_entry.p_print_wastage_form_8) + int(print_wastage_form_8)
                # existing_entry.p_print_wastage_form_9 = int(existing_entry.p_print_wastage_form_9) + int(print_wastage_form_9)
                # existing_entry.p_print_wastage_form_10 = int(existing_entry.p_print_wastage_form_10) + int(print_wastage_form_10)
                entry.total_received_forms= float(entry.total_received_forms) + float(total_received_forms)
                entry.p_print_date = p_print_date
                
                
                entry.save()
                new_entry_needed = False
                break

           
        
         if new_entry_needed:
        
            # Create new entry

        
            data1=printing(p_vender_name= p_vender_name, p_brand_name_1=p_brand_name_1,p_class_name_1=p_class_name_1,p_medium_1=p_medium_1,p_subject_1=p_subject_1,p_printing_received_form_1=received_form_values[0],p_printing_received_form_2=received_form_values[1],p_printing_received_form_3=received_form_values[2],p_printing_received_form_4=received_form_values[3],p_print_date=p_print_date,
                       p_print_wastage_form_1= print_wastage_form_1, p_print_wastage_form_2 = print_wastage_form_2, p_print_wastage_form_3 = print_wastage_form_3, p_print_wastage_form_4 = print_wastage_form_4,p_single_1=p_single_1,p_double_1=p_double_1,p_four_1=p_four_1,p_quantity_1= p_quantity_1 , p_rim_1 = p_rim_1 , p_pages_1= p_pages_1,
                       p_printing_received_form_5 = received_form_values[4], p_printing_received_form_6 = received_form_values[5], p_printing_received_form_7 = received_form_values[6], p_printing_received_form_8= received_form_values[7], p_printing_received_form_9 = received_form_values[8], p_printing_received_form_10 = received_form_values[9],
                       p_print_wastage_form_5 = print_wastage_form_5, p_print_wastage_form_6 = print_wastage_form_6, p_print_wastage_form_7 = print_wastage_form_7, p_print_wastage_form_8= print_wastage_form_8, p_print_wastage_form_9 = print_wastage_form_9, p_print_wastage_form_10 = print_wastage_form_10,
                       total_received_forms = total_received_forms,p_code=p_code)
            data1.save()
            print("fffg",data1)
     vend=venders.objects.all()
     book = books.objects.all()
     obj = books.objects.values('code').distinct().order_by('code')
     obj1 = books.objects.values('books_name').distinct().order_by('books_name')
     obj2 = books.objects.values('books_class').distinct().order_by('books_class')
     obj3 = books.objects.values('medium').distinct().order_by('medium')
     obj4 = books.objects.values('subject').distinct().order_by('subject')
     obj5 = books.objects.values('no_of_pages').distinct().order_by('no_of_pages') 
     quan = orders.objects.all()
     print('orders',quan)
   
     total_quantity = 0

     for order in quan:
        for i in range(1, 51):
            field_prefix = f'brand_name_{i}'  # Change 'brand_name' to the appropriate field prefix

            # Use getattr() to dynamically access the field by its name for the current order
            field_value = getattr(order, field_prefix, None)
            if field_value is not None:
                print(f'{field_prefix}: {field_value}')
                
                # Similarly, access other fields like class_name, subject, medium, pages, and quantity
                class_field = f'class_name_{i}'
                subject_field = f'subject_{i}'
                medium_field = f'medium_{i}'
                pages_field = f'pages_{i}'
                quantity_field = f'quantity_{i}'
                
                class_value = getattr(order, class_field, None)
                subject_value = getattr(order, subject_field, None)
                medium_value = getattr(order, medium_field, None)
                pages_value = getattr(order, pages_field, None)
                quantity_value = int(getattr(order, quantity_field, 0))
                # Add the quantity to the total
                total_quantity += quantity_value
                # Storing the total quantity for the combination in the dictionary
                combination = (field_value, class_value, subject_value, medium_value,pages_value)
                if combination in total_quantities:
                    total_quantities[combination] += quantity_value
                else:
                    total_quantities[combination] = quantity_value

                
                print(f'{class_field}: {class_value}')
                print(f'{subject_field}: {subject_value}')
                print(f'{medium_field}: {medium_value}')
                print(f'{pages_field}: {pages_value}')
                print(f'{quantity_field}: {quantity_value}')
                
                print('-----------------------')
     # Print the calculated total quantity for the desired brand combination
     print(f"Total Quantity for 'kohinoor' combination: {total_quantity}")
     for combination, total_quantity in total_quantities.items():
        brand, class_name, subject, medium , pages = combination
        print(f"Brand: {brand}, Class: {class_name}, Subject: {subject}, Medium: {medium}, Pages: {pages}")
        print(f"Total Quantity: {total_quantity}")
        print("-----------------------")

        brands = {
        'Kohinoor': {'pages_per_form': 64},
        'Spark': {'pages_per_form': 128},
        'Vidya Mitra': {'pages_per_form': 32},
        'Abhiyashika': {'pages_per_form': 16},
        }   
        # forms_data = {}
        form1 = []
        forms_data = {}
        total_form = 0

        

         # Iterate through the total_quantities dictionary
        for combination, total_quantity in total_quantities.items():
            brand, class_name, subject, medium, pages = combination[:5]
            # Check if the brand is in the predefined brands dictionary
            try:
                pages = float(pages)  # Convert to float
            except ValueError:
            # Handle the case where conversion fails
                print(f"Invalid 'pages' value for combination: {combination}")
                continue
            if brand in brands:
                pages_per_form = brands[brand]['pages_per_form']
                
                # Calculate the number of forms based on the pages and quantity
                forms = (pages / pages_per_form * total_quantity)
                forms_data[combination] = {'forms': forms}
                
                # total_set = pages / 32
                # forms_per_set = forms / total_set
                # forms_data[combination]['total_set'] = total_set
                # forms_data[combination]['forms_per_set'] = forms_per_set
                
                form1.append(forms_data)

            # total_forms = sum(form1)
            print("67890",forms)

     data11=printing.objects.all().order_by('-id')
     for i in data11:
         
        print("gggggggggg",i.p_brand_name_1) 
   
     context={
             "data11":data11,
             "vend":vend,
             "book":book,
             "total_received_forms":total_received_forms,
             'quan':quan,
             'obj1':obj1,
             'obj2':obj2,
             'obj3':obj3,
             'obj4':obj4,
             'obj5':obj5,
             'obj':obj,
             'total_quantities':total_quantities,
             'forms_data':forms_data,
            }
     return render(request, 'brands/printing_bookwise.html', context)






def edit_printing_data(request, id):
     data = printing.objects.get(id=id)
     if request.method == 'POST':
        data.p_vender_name=request.POST.get("p_vender_name")
        data.p_vender_address =request.POST.get("p_vender_address")
        data.p_vender_mob=request.POST.get("p_vender_mob")
        data.p_brand_name_1=request.POST.get("p_brand_name_1")
        data.p_class_name_1=request.POST.get("p_class_name_1")
        data.p_medium_1=request.POST.get("p_medium_1")
        
        data.p_subject_1=request.POST.get("p_subject_1")
        data.p_quantity_1=request.POST.get("p_quantity_1")
        data.p_rim_1=request.POST.get("p_rim_1")
        data.p_pages_1=request.POST.get("p_pages_1")
        data.p_printing_received_form_1=request.POST.get("p_printing_received_form_1")
        data.p_printing_received_form_2=request.POST.get("p_printing_received_form_2")
        data.p_printing_received_form_3=request.POST.get("p_printing_received_form_3")
        data.p_printing_received_form_4=request.POST.get("p_printing_received_form_4")
        data.p_printing_received_form_5=request.POST.get("p_printing_received_form_5")
        data.p_printing_received_form_6=request.POST.get("p_printing_received_form_6")
        data.p_printing_received_form_7=request.POST.get("p_printing_received_form_7")
        data.p_printing_received_form_8=request.POST.get("p_printing_received_form_8")
        data.p_printing_received_form_9=request.POST.get("p_printing_received_form_9")
        data.p_printing_received_form_10=request.POST.get("p_printing_received_form_10")
        data.p_print_wastage_form_1=request.POST.get("p_print_wastage_form_1")
        data.p_print_wastage_form_2=request.POST.get("p_print_wastage_form_2")
        data.p_print_wastage_form_3=request.POST.get("p_print_wastage_form_3")
        data.p_print_wastage_form_4=request.POST.get("p_print_wastage_form_4")
        data.p_print_wastage_form_5=request.POST.get("p_print_wastage_form_5")
        data.p_print_wastage_form_6=request.POST.get("p_print_wastage_form_6")
        data.p_print_wastage_form_7=request.POST.get("p_print_wastage_form_7")
        data.p_print_wastage_form_8=request.POST.get("p_print_wastage_form_8")
        data.p_print_wastage_form_9=request.POST.get("p_print_wastage_form_9")
        data.p_print_wastage_form_10=request.POST.get("p_print_wastage_form_10")
        data.p_single_1=request.POST.get("p_single_1")
        data.p_double_1=request.POST.get("p_double_1")
        data.p_four_1=request.POST.get("p_four_1")

        data.p_print_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
         
        # data.p_print_date=request.POST.get('p_print_date')
        total_received_forms = ""
        # Convert gather fields to float (modify as needed based on your model)
        received_form_values = [
            int(getattr(data, f'p_printing_received_form_{i}') or 0) if getattr(data, f'p_printing_received_form_{i}') is not None else 0 for i in range(1, 11)
         ]
        data.total_received_forms = sum(received_form_values)
        print('total',total_received_forms)
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/printing-data/')
    
     data2 = venders.objects.all()  # Correct the model name based on your actual model
    
     context = {'data': data, 'data2': data2}
     return render(request, "brands/edit_printing_data.html", context)


def delete_printing(request, id):
    cus = printing.objects.get(id=id)
    cus.delete()
    return redirect('/printing-data/')


# def print_view(request,id):
#     data = printing.objects.get(id=id)
#     print('data', data)
    
#     context = {
#         'data': data,
#     }    
#     return render(request, 'brands/print_view.html',context)   




def print_view(request,id):
    data = printing.objects.get(id=id)
   
   
    if request.method == 'POST':
        data.tra_from_1=request.POST.get("tra_from_1")
        data.to_1=request.POST.get("to_1")
        print("1234",data.to_1)
        data.transfer1=request.POST.get("transfer1")
        data.transfer2=request.POST.get("transfer2")
        data.transfer3=request.POST.get("transfer3")
        data.transfer4=request.POST.get("transfer4")
        data.transfer5=request.POST.get("transfer5")
        # data.transfer5=request.POST.get("tra_from_6")
        data.transfer6=request.POST.get("transfer6")
        data.transfer7=request.POST.get("transfer7")
        data.transfer8=request.POST.get("transfer8")
        data.transfer9=request.POST.get("transfer9")
        data.transfer10=request.POST.get("transfer10")
        # data.transfer_pending1=request.POST.get("transfer_pending1")
        # data.transfer_pending2=request.POST.get("transfer_pending2")
        # data.transfer_pending3=request.POST.get("transfer_pending3")
        # data.transfer_pending4=request.POST.get("transfer_pending4")
        # data.transfer_pending5=request.POST.get("transfer_pending5")
        # data.transfer_pending6=request.POST.get("transfer_pending6")
        # data.transfer_pending7=request.POST.get("transfer_pending7")
        # data.transfer_pending8=request.POST.get("transfer_pending8")
        # data.transfer_pending9=request.POST.get("transfer_pending9")
        # data.transfer_pending10=request.POST.get("transfer_pending10")
        data.p_stock=request.POST.get("p_stock")
        data.p_vender_name=request.POST.get("p_vender_name")
        data.p_code=request.POST.get("p_code")
        data.p_vender_address =request.POST.get("p_vender_address")
        data.p_vender_mob=request.POST.get("p_vender_mob")
        data.p_brand_name_1=request.POST.get("p_brand_name_1")
        data.p_class_name_1=request.POST.get("p_class_name_1")
        data.p_medium_1=request.POST.get("p_medium_1")
        
        data.p_subject_1=request.POST.get("p_subject_1")
        data.Material_center = request.POST.get('Material_center')


        data.p_quantity_1=request.POST.get("p_quantity_1")
        data.p_rim_1=request.POST.get("p_rim_1")
        data.p_pages_1=request.POST.get("p_pages_1")
        data.p_printing_received_form_1=request.POST.get("p_printing_received_form_1")
        data.p_printing_received_form_2=request.POST.get("p_printing_received_form_2")
        data.p_printing_received_form_3=request.POST.get("p_printing_received_form_3")
        data.p_printing_received_form_4=request.POST.get("p_printing_received_form_4")
        data.p_printing_received_form_5=request.POST.get("p_printing_received_form_5")
        data.p_printing_received_form_6=request.POST.get("p_printing_received_form_6")
        data.p_printing_received_form_7=request.POST.get("p_printing_received_form_7")
        data.p_printing_received_form_8=request.POST.get("p_printing_received_form_8")
        data.p_printing_received_form_9=request.POST.get("p_printing_received_form_9")
        data.p_printing_received_form_10=request.POST.get("p_printing_received_form_10")
        data.p_print_wastage_form_1=request.POST.get("p_print_wastage_form_1")
        data.p_print_wastage_form_2=request.POST.get("p_print_wastage_form_2")
        data.p_print_wastage_form_3=request.POST.get("p_print_wastage_form_3")
        data.p_print_wastage_form_4=request.POST.get("p_print_wastage_form_4")
        data.p_print_wastage_form_5=request.POST.get("p_print_wastage_form_5")
        data.p_print_wastage_form_6=request.POST.get("p_print_wastage_form_6")
        data.p_print_wastage_form_7=request.POST.get("p_print_wastage_form_7")
        data.p_print_wastage_form_8=request.POST.get("p_print_wastage_form_8")
        data.p_print_wastage_form_9=request.POST.get("p_print_wastage_form_9")
        data.p_print_wastage_form_10=request.POST.get("p_print_wastage_form_10")
        data.p_single_1=request.POST.get("p_single_1")
        data.p_double_1=request.POST.get("p_double_1")
        data.p_four_1=request.POST.get("p_four_1")
        
        data.p_print_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.transfer=request.POST.get("transfer")
        data.pending=request.POST.get("transfer_pending")
        transfer = int(data.transfer) if data.transfer else 0
        data.total_recived_forms=request.POST.get('total_recived_forms')
         
        # data.p_print_date=request.POST.get('p_print_date')
        total_received_forms = ""
        # Convert gather fields to float (modify as needed based on your model)
        received_form_values = [
            int(getattr(data, f'p_printing_received_form_{i}') or 0) if getattr(data, f'p_printing_received_form_{i}') is not None else 0 for i in range(1, 11)
         ]
        data.total_received_forms = sum(received_form_values)
        print('total',total_received_forms)
        data.tra_from_1 = data.tra_from_1
        print('fghjk',data.tra_from_1)
        data.to_1 = data.to_1
        print('fghjk',data.to_1)
        data.transfer1 = data.transfer1
        print('222222222222',data.transfer1)
        data.Material_center=data.Material_center
        # data.transfer2 = data.transfer2
        # data.transfer3 = data.transfer3
        # data.transfer4 = data.transfer4
        # data.transfer5 = data.transfer5
        # data.transfer6 = data.transfer6
        # data.transfer7 = data.transfer7
        # data.transfer8 = data.transfer8
        # data.transfer9 = data.transfer9
        # data.transfer10 = data.transfer10
        data.transfer_pending1 = int(data.p_printing_received_form_1 or '0') - int(data.transfer1 or '0')
        print("333333333",data.transfer_pending1)
        data.transfer_pending2 = int(data.p_printing_received_form_2 or '0') - int(data.transfer2 or '0')
        data.transfer_pending3 = int(data.p_printing_received_form_3 or '0') - int(data.transfer3 or '0')
        data.transfer_pending4 = int(data.p_printing_received_form_4 or '0') - int(data.transfer4 or '0')
        data.transfer_pending5 = int(data.p_printing_received_form_5 or '0') - int(data.transfer5 or '0')
        data.transfer_pending6 = int(data.p_printing_received_form_6 or '0') - int(data.transfer6 or '0')
        data.transfer_pending7 = int(data.p_printing_received_form_7 or '0') - int(data.transfer7 or '0')
        data.transfer_pending8 = int(data.p_printing_received_form_8 or '0') - int(data.transfer8 or '0')
        data.transfer_pending9 = int(data.p_printing_received_form_9 or '0') - int(data.transfer9 or '0')
        data.transfer_pending10 = int(data.p_printing_received_form_10 or '0') - int(data.transfer10 or '0')
        print('111111111111111',data.transfer_pending10)
        # data.transfer_pending1 = int(data.p_printing_received_form_4) - int(data.transfer4)
        print('fghjk',data.transfer1)
        # data.pending=int(data.total_received_forms) - int(transfer)
        # print("ppppppppppeeending",data.pending)
        # data.transfer_pending = data.pending
        transfer = ""
        # Convert gather fields to float (modify as needed based on your model)
        transfer_form_values = [
            int(getattr(data, f'transfer{i}') or 0) if getattr(data, f'transfer{i}') is not None else 0 for i in range(1, 11)
        ]

        data.transfer = sum(transfer_form_values)
        print('total',transfer)
        transfer_pending = ""
        # Convert gather fields to float (modify as needed based on your model)
        transfer_form_values = [
            int(getattr(data, f'transfer_pending{i}') or 0) if getattr(data, f'transfer_pending{i}') is not None else 0 for i in range(1, 11)
        ]

        data.transfer_pending = sum(transfer_form_values)
        print('total',data.transfer_pending)
        data.save()
        
       
        
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("print_approval") == "approval_button":
                printing.objects.filter(Q(id=id) & Q(print_approved_1=False)).update(print_approved_1=True)
                return redirect('/printing-data/')
                

        
        # return redirect('/printing-data/')
    data.total_received_form_data = data.total_received_forms
    data.total_transfer_form = data.transfer
    
    data.transfer_total_form1= data.transfer1 
    data.transfer_total_form2= data.transfer2
    data.transfer_total_form3= data.transfer3 
    data.transfer_total_form4= data.transfer4 
    data.transfer_total_form5= data.transfer5 
    data.transfer_total_form6= data.transfer6 
    data.transfer_total_form7= data.transfer7 
    data.transfer_total_form8= data.transfer8 
    data.transfer_total_form9= data.transfer9 
    data.transfer_total_form10= data.transfer10
    data.save()
    
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3=Bindingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data5':data5}
  
    
        
    return render(request, 'brands/print_view.html',context)





from django.http import HttpResponse,HttpResponseRedirect
from django.template import loader
from django.urls import reverse
def print_transfer(request):
    # data1 =printing.objects.all().values()
    data1 = printing.objects.filter(print_approved_1=True)
    print("data1",data1)
   
    template=loader.get_template('brands/print_transfer.html')
    context={
      'data1':data1,
    }
    return HttpResponse(template.render(context,request))    





# from django.db.models import Q ,Sum
# from django.utils import timezone
# from datetime import datetime
# def gathering(request):
#     if request.method == 'POST':
#         v1_name = request.POST.get("v1_name")
#         brand_name = request.POST.get("brand_name")
#         class_name = request.POST.get("class_name")
#         medium = request.POST.get("medium")
#         subject = request.POST.get("subject")
#         gather1 = request.POST.get("gather1")
#         gather2 = request.POST.get("gather2")
#         gather3 = request.POST.get("gather3")
#         gather4 = request.POST.get("gather4")
#         gather5 = request.POST.get("gather5")
#         gather6 = request.POST.get("gather6")
#         gather7 = request.POST.get("gather7")
#         gather8 = request.POST.get("gather8")
#         gather9 = request.POST.get("gather9")
#         gather10 = request.POST.get("gather10")
#         page = request.POST.get("page")
#         total_gather=request.POST.get("total_gather")
#         current_date = datetime.now()
#         formatted_date = current_date.strftime('%d-%m-%Y %I:%M:%S %p')
#         # gather_fields = [gather1, gather2, gather3, gather4, gather5, gather6, gather7, gather8, gather9, gather10]
#         # gather_values = [float(request.POST.get(f"gather{i}", 0)) if request.POST.get(f"gather{i}") else 0 for i in range(1, 11)]
#         # total_gather = sum(gather_values)
#         # print("3456789",total_gather)


#         gather_values = []
#         for i in range(1, 11):
#             value = request.POST.get(f"gather{i}")
#             if value is not None:
#                 gather_values.append(int(value))
#             else:
#                 gather_values.append(0)
        
#         # Calculate total gather
#         total_gather = sum(gather_values)
#         print("3456789", total_gather)

#         existing_entries = Gathering.objects.filter(
#             brand_name=brand_name,
#             class_name=class_name,
#             medium=medium,
#             subject=subject,
#             # m1_no=m1_no,
#             # a1_name=a1_name,
#             page=page,
#             # date = formatted_date,

#         )

#         new_entry_needed = True
#         for entry in existing_entries:
#             if entry.v1_name == v1_name:
#                 # Update existing entry
#                 entry.v1_name = v1_name

#                 for i in range(10):
#                     if getattr(entry, f"gather{i+1}") is not None:
#                         setattr(entry, f"gather{i+1}", int(getattr(entry, f"gather{i+1}")) + gather_values[i])
#                     else:
#                         setattr(entry, f"gather{i+1}", gather_values[i])
#                     # setattr(entry, f"gather{i+1}", int(getattr(entry, f"gather{i+1}")) + gather_values[i])
#                 # entry.total_gather += total_gather
                
#                 # entry.gather1 = float(gather1 )+ float(entry.gather1)
#                 # entry.gather2 = float(gather2) + float(entry.gather2)
#                 # entry.gather3 = float(gather3) + float(entry.gather3) 
#                 # entry.gather4 = float(gather4)  + float(entry.gather4)
#                 # entry.gather5 = float(gather5) + float(entry.gather5)
#                 # entry.gather6 = float(gather6) + float(entry.gather6)
#                 # entry.gather7 = float(gather7) + float(entry.gather7)
#                 # entry.gather8 = float(gather8)  + float(entry.gather8)
#                 # entry.gather9 = float(gather9)  + float(entry.gather9)
#                 # entry.gather10 = float(gather10) + float(entry.gather10)
#                 entry.total_gather= int(total_gather) + int(entry.total_gather)
#                 print('jhgfgh',entry.total_gather,total_gather)
#                 entry.date = formatted_date
#                 entry.save()
#                 new_entry_needed = False
#                 break
        
#         if new_entry_needed:
        
#             # Create new entry
#             data1 = Gathering(
#                 brand_name=brand_name,
#                 class_name=class_name,
#                 medium=medium,
#                 subject=subject,
#                 gather1=gather1,
#                 gather2=gather2,
#                 gather3=gather3,
#                 gather4=gather4,
#                 gather5=gather5,
#                 gather6=gather6,
#                 gather7=gather7,
#                 gather8=gather8,
#                 gather9=gather9,
#                 gather10=gather10,
#                 total_gather=total_gather,
#                 date=formatted_date,
#                 v1_name=v1_name,
#                 # m1_no=m1_no,
#                 # a1_name=a1_name,
#                 page=page,
#             )
#             data1.save()

#             print("12322",data1)    

#     vend = venders.objects.all()
#     book = books.objects.all()
#     data11 = Gathering.objects.all().order_by('-id')
#     # gather_values = [float(request.POST.get(f"gather{i}", 0)) for i in range(1, 11)]
#     # total_gather = sum(gather_values)
         

    
#     context = {
#         "data11": data11,
#         "vend": vend,
#         "book": book,
       
        
#     }
#     return render(request, 'brands/gather.html', context)



from django.db.models import Q ,Sum
from django.utils import timezone
from datetime import datetime
def gathering(request):
    if request.method == 'POST':
        v1_name = request.POST.get("v1_name")
        brand_name = request.POST.get("brand_name")
        print("brand name",brand_name)
        class_name = request.POST.get("class_name")
        medium = request.POST.get("medium")
        subject = request.POST.get("subject")
        g_quantity = request.POST.get("g_quantity")
        print("gathering Quantity",g_quantity)
        # gather1 = request.POST.get("gather1")
        # gather2 = request.POST.get("gather2")
        # gather3 = request.POST.get("gather3")
        # gather4 = request.POST.get("gather4")
        # gather5 = request.POST.get("gather5")
        # gather6 = request.POST.get("gather6")
        # gather7 = request.POST.get("gather7")
        # gather8 = request.POST.get("gather8")
        # gather9 = request.POST.get("gather9")
        # gather10 = request.POST.get("gather10")
        page = request.POST.get("page")
        # total_gather=request.POST.get("total_gather")
        code_g=request.POST.get("code_g")
        current_date = datetime.now()
        formatted_date = current_date.strftime('%d-%m-%Y %I:%M:%S %p')
        


        gather_values = []
        for i in range(1, 11):
            value = request.POST.get(f"gather{i}")
            if value is not None:
                gather_values.append(int(value))
            else:
                gather_values.append(0)
        
        # Calculate total gather
        total_gather = sum(gather_values)
        print("3456789", total_gather)

        existing_entries = Gathering.objects.filter(
            brand_name=brand_name,
            class_name=class_name,
            medium=medium,
            subject=subject,
            
            page=page,
            # date = formatted_date,

        )

        new_entry_needed = True
        for entry in existing_entries:
            if entry.v1_name == v1_name:
                # Update existing entry
                entry.v1_name = v1_name
                entry.code_g = code_g

                # for i in range(10):
                #     if getattr(entry, f"gather{i+1}") is not None:
                #         setattr(entry, f"gather{i+1}", int(getattr(entry, f"gather{i+1}")) + gather_values[i])
                #     else:
                #         setattr(entry, f"gather{i+1}", gather_values[i])
                 
                # entry.total_gather= int(total_gather) + int(entry.total_gather)
                # print('jhgfgh',entry.total_gather,total_gather)
                entry.date = formatted_date
                entry.save()
                new_entry_needed = False
                break
        
        if new_entry_needed:
        
            # Create new entry
            data1 = Gathering(
                brand_name=brand_name,
                class_name=class_name,
                medium=medium,
                subject=subject,
                g_quantity=g_quantity,
                
                # gather1=gather1,
                # gather2=gather2,
                # gather3=gather3,
                # gather4=gather4,
                # gather5=gather5,
                # gather6=gather6,
                # gather7=gather7,
                # gather8=gather8,
                # gather9=gather9,
                # gather10=gather10,
                # total_gather=total_gather,
                date=formatted_date,
                v1_name=v1_name,
                code_g=code_g,
                # a1_name=a1_name,
                page=page,
            )
            data1.save()

            print("12322",data1)    

    vend = venders.objects.all()
    book = books.objects.all()
    data11 = Gathering.objects.all().order_by('-id')
    # gather_values = [float(request.POST.get(f"gather{i}", 0)) for i in range(1, 11)]
    # total_gather = sum(gather_values)
    obj = books.objects.values('code').distinct().order_by('code')
    obj1 = books.objects.values('books_name').distinct().order_by('books_name')
    obj2 = books.objects.values('books_class').distinct().order_by('books_class')  
    data99=Gatheringvender.objects.all() 
    data3=Gather_ledger.objects.all() 

    
    context = {
        "data11": data11,
        "vend": vend,
        "book": book,
        "obj": obj,
        "obj1":obj1,
        "obj2":obj2,
        "data99":data99,
        "data3":data3,
       
        
    }
    return render(request, 'brands/gather.html', context)





def delete_gather(request, id):
    cus = Gathering.objects.get(id=id)
    cus.delete()
    return redirect('/gathering/')    





def edit_gather(request, id):
    data = Gathering.objects.get(id=id)
    if request.method == 'POST':
        data.v1_name = request.POST.get('v1_name')
        data.m1_no = request.POST.get('m1_no')
        data.a1_name = request.POST.get('a1_name')
        data.brand_name = request.POST.get('brand_name')
        data.class_name = request.POST.get('class_name')
        data.medium = request.POST.get('medium')
        data.subject = request.POST.get('subject')
        data.g_quantity = request.POST.get('g_quantity')
        data.gather1 = request.POST.get('gather1')
        data.gather2 = request.POST.get('gather2')
        data.gather3 = request.POST.get('gather3')
        data.gather4 = request.POST.get('gather4')
        data.gather5 = request.POST.get('gather5')
        data.gather6 = request.POST.get('gather6')
        data.gather7 = request.POST.get('gather7')
        data.gather8 = request.POST.get('gather8')
        data.gather9 = request.POST.get('gather9')
        data.gather10 = request.POST.get('gather10')
        current_date = datetime.now()
        data.date=current_date.strftime('%d-%m-%Y %I:%M:%S %p')
        
        # Convert gather fields to float (modify as needed based on your model)
        gather_values = [int(request.POST.get(f"gather{i}", 0)) if request.POST.get(f"gather{i}") else 0 for i in range(1, 11)]
        
        total_gather = sum(gather_values)
        data.total_gather = total_gather
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        return redirect('/gathering/')
    
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3=Gather_ledger.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3}
    return render(request, "brands/edit_gather.html", context)

# def gather_view(request,id):
#     data = Gathering.objects.get(id=id)
#     print('data', data)
    
#     context = {
#         'data': data,
#     }    
#     return render(request, 'brands/gather_view.html',context)    



def gather_view(request,id):
    data = Gathering.objects.get(id=id)
    if request.method == 'POST':
        data.tra_from_2=request.POST.get("tra_from_2")
        data.to_2=request.POST.get("to_2")
        print("1234",data.to_2)
        # data.voucher_2=request.POST.get("voucher_2")
        data.code_g = request.POST.get('code_g')
        data.v1_name = request.POST.get('v1_name')
        data.m1_no = request.POST.get('m1_no')
        data.a1_name = request.POST.get('a1_name')
        data.brand_name = request.POST.get('brand_name')
        data.class_name = request.POST.get('class_name')
        data.medium = request.POST.get('medium')
        data.subject = request.POST.get('subject')
        data.Material_center = request.POST.get('Material_center')
        # data.gather1 = request.POST.get('gather1')
        # data.gather2 = request.POST.get('gather2')
        # data.gather3 = request.POST.get('gather3')
        # data.gather4 = request.POST.get('gather4')
        # data.gather5 = request.POST.get('gather5')
        # data.gather6 = request.POST.get('gather6')
        # data.gather7 = request.POST.get('gather7')
        # data.gather8 = request.POST.get('gather8')
        # data.gather9 = request.POST.get('gather9')
        # data.gather10 = request.POST.get('gather10')
        data.g_quantity=request.POST.get('g_quantity')
        data.page = request.POST.get('page')
        data.date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.transfer=request.POST.get('transfer')
        data.pending=request.POST.get("transfer_pending")
        transfer = int(data.transfer) if data.transfer else 0
        
        # Convert gather fields to float (modify as needed based on your model)
        gather_values =[
        
            int(getattr(data, f'gather{i}') or 0) if getattr(data, f'gather{i}') is not None else 0 for i in range(1, 11)
        ]
        total_gather = sum(gather_values)
        data.total_gather = total_gather
        data.tra_from_2 = data.tra_from_2
        print('fghjk',data.tra_from_2)
        data.Material_center=data.Material_center
        data.to_2 = data.to_2
        print('fghjk',data.to_2)
        data.pending=int(data.g_quantity) - int(transfer)
        print("ppppppppppeeending",data.pending)
        data.transfer_pending = data.pending
        data.save()
        messages.success(request, "Updated Successfully...!!")
        if request.method == 'POST':
            if request.POST.get("gather_approved_1") == "approval_button":
                Gathering.objects.filter(Q(id=id) & Q( gather_approved_1=False)).update( gather_approved_1=True)
                return redirect('/gathering/')
              
        # return redirect('/gathering/')
    data.total_remaining = data.transfer_pending
    data.total_transfer_form=data.transfer
    data.save()
    
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3=Bindingvender.objects.all()
    data4=Gather_ledger.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data4':data4,'data5':data5}
    return render(request, 'brands/gather_view.html',context)



def gather_transfer(request):
    # data1 =printing.objects.all().values()
    data1 = Gathering.objects.filter(gather_approved_1=True).exclude(brand_name="Spark")
    print("data1",data1)
   
    template=loader.get_template('brands/gather_transfer.html')
    context={
      'data1':data1,
    }
    return HttpResponse(template.render(context,request))     




def Cutting_info(request):
    if request.method == 'POST':
        c_v1_name = request.POST.get("c_v1_name")
        c_brand_name = request.POST.get("c_brand_name")
        c_class_name = request.POST.get("c_class_name")
        c_medium = request.POST.get("c_medium")
        c_subject = request.POST.get("c_subject")
        c_quantity_5 = (request.POST.get("c_quantity_5"))  # Convert to int
        c_code = request.POST.get("c_code")
        complete_book = request.POST.get("complete_book")
        
        # Check if the book with the given code exists
        books_queryset = books.objects.filter(code=c_code)
        if books_queryset.exists():
            # Choose the first book from the queryset (or implement your own logic)
            book = books_queryset.first()

            # Update the book's stock
            book.stock += c_quantity_5
            book.save()
            complete_book = book.stock
            print("bookstock", complete_book)
        else:
            # Handle the case where no book with the given code exists
            # You can add your own logic here, such as creating a new book entry
            pass

        existing_entry = Cuttting.objects.filter(
            c_v1_name=c_v1_name,
            c_brand_name=c_brand_name,
            c_class_name=c_class_name,
            c_medium=c_medium,
            c_subject=c_subject,
            c_code=c_code
        ).first()

        if existing_entry:
            # Entry exists, update it
            existing_entry.c_quantity_5 += c_quantity_5
            existing_entry.c_date = datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
            existing_entry.save()
        else:
            # Create a new entry in the "cutting" module
            new_entry = Cuttting.objects.create(
                c_v1_name=c_v1_name,
                c_brand_name=c_brand_name,
                c_class_name=c_class_name,
                c_medium=c_medium,
                c_subject=c_subject,
                c_quantity_5=c_quantity_5,
                c_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),
                c_code=c_code,
                complete_book=complete_book,
            )

    vend = venders.objects.all()
    books_list = books.objects.all()
    obj1 = books.objects.values('books_name').distinct().order_by('books_name')
    obj2 = books.objects.values('books_class').distinct().order_by('books_class')
    obj = books.objects.values('code').distinct().order_by('code')
    data98 = Bindingvender.objects.all()

    data11 = Cuttting.objects.all().order_by('-id')
    data = Gathering.objects.all()
    context = {
        'vend': vend,
        'book': books_list,
        'data11': data11,
        'obj1': obj1,
        'obj2': obj2,
        'obj': obj,
        'data98': data98,
    }

    return render(request, 'brands/cutting.html', context)



def delete_cutting(request, id):
    cus = Cuttting.objects.get(id=id)
    cus.delete()
    return redirect('/cutting_info/')   


# def edit_cutting(request, id):
#     data = Cuttting.objects.get(id=id)
#     if request.method == 'POST':
#         data.c_v1_name = request.POST.get('c_v1_name')
#         data.c_m1_no = request.POST.get('c_m1_no')
#         data.c_a1_name = request.POST.get('c_a1_name')
#         data.c_brand_name = request.POST.get('c_brand_name')
#         data.c_class_name = request.POST.get('c_class_name')
#         data.c_medium = request.POST.get('c_medium')
#         data.c_subject = request.POST.get('c_subject')
        
#         data.c_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
#         data.c_quantity_5=request.POST.get('c_quantity_5')
        
#         # Convert gather fields to float (modify as needed based on your model)
#         # gather_values = [float(request.POST.get(f"gather{i}", 0)) if request.POST.get(f"gather{i}") else 0 for i in range(1, 11)]
        
#         # total_gather = sum(gather_values)
#         # data.total_gather = total_gather
        
#         data.save()
#         messages.success(request, "Updated Successfully...!!")
#         return redirect('/cutting_info/')
    
#     data2 = venders.objects.all()  # Correct the model name based on your actual model
    
#     context = {'data': data, 'data2': data2}
#     return render(request, "brands/edit_cutting.html", context)  



def edit_cutting(request, id):
    data = Cuttting.objects.get(id=id)
    if request.method == 'POST':
        data.c_code = request.POST.get('c_code')
        data.tra_from_c = request.POST.get('tra_from_c')
        data.to_c = request.POST.get('to_c')
        data.voucher_c = request.POST.get('voucher_c')
        data.c_brand_name = request.POST.get('c_brand_name')
        data.c_v1_name = request.POST.get('c_v1_name')
        data.c_m1_no = request.POST.get('c_m1_no')
        data.c_a1_name = request.POST.get('c_a1_name')
        data.c_brand_name = request.POST.get('c_brand_name')
        data.c_class_name = request.POST.get('c_class_name')
        data.c_medium = request.POST.get('c_medium')
        data.c_subject = request.POST.get('c_subject')
        data.Material_center = request.POST.get('Material_center')
        
        data.c_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.c_quantity_5=request.POST.get('c_quantity_5')

        data.transfer=request.POST.get("transfer")
        data.pending=request.POST.get("transfer_pending")

        transfer = int(data.transfer) if data.transfer else 0
        
        # Convert gather fields to float (modify as needed based on your model)
        # gather_values = [float(request.POST.get(f"gather{i}", 0)) if request.POST.get(f"gather{i}") else 0 for i in range(1, 11)]
        
        # total_gather = sum(gather_values)
        # data.total_gather = total_gather
        data.pending=int(data.c_quantity_5) - int(transfer)
        print("ppppppppppeeending",data.pending)
        data.transfer_pending = data.pending



        
        # Convert gather fields to float (modify as needed based on your model)
        # gather_values = [float(request.POST.get(f"gather{i}", 0)) if request.POST.get(f"gather{i}") else 0 for i in range(1, 11)]
        
        # total_gather = sum(gather_values)
        # data.total_gather = total_gather
        data.tra_from_c = data.tra_from_c
        data.Material_center=data.Material_center
        print('fghjk',data.tra_from_c)
        data.to_c = data.to_c
        print('fghjk',data.to_c)


        
        
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("machine_cutting_approval") == "approval_button":
                Cuttting.objects.filter(Q(id=id) & Q(machine_cutting_approved_1=False)).update(machine_cutting_approved_1=True)
                return redirect('/cutting_info/')
            # else:
            #     pass    
        
        
        # return redirect('/cutting_info/')
    data.total_transfer_form = data.transfer
    data.total_remaining = data.transfer_pending
    data.save()
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3 = Packingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data5':data5}
    return render(request, "brands/edit_cutting.html", context)




def cutting_machine_transfer_packing(request):
    data1 = Cuttting.objects.filter(machine_cutting_approved_1=True).exclude(Q(c_code__startswith='S') | Q(c_code__startswith='s'))
    print("data1",data1)
   
    template=loader.get_template('brands/cutting_machine_transfer_packing.html')
    context={
      'data1':data1,
    }
    return HttpResponse(template.render(context,request))     
    

# def  machine_binding(request):
#     if request.method == 'POST':
#         b_v1_name = request.POST.get("b_v1_name")
       
#         b_brand_name = request.POST.get("b_brand_name")
#         b_class_name = request.POST.get("b_class_name")
#         b_medium = request.POST.get("b_medium")
#         b_subject = request.POST.get("b_subject")
        
        
#         # page = request.POST.get("page")
#         total_binding=request.POST.get("total_binding")
#         # current_date = timezone.now().date()
#         formatted_date = datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
#         print('lklklkll',formatted_date)
#         # gather_fields = [gather1, gather2, gather3, gather4, gather5, gather6, gather7, gather8, gather9, gather10]
#         # gather_values = [float(request.POST.get(f"gather{i}", 0)) if request.POST.get(f"gather{i}") else 0 for i in range(1, 11)]
#         # total_gather = sum(gather_values)
#         # print("3456789",total_gather)

#         existing_entries = Binding.objects.filter(
#             b_v1_name=b_v1_name,
           
#             b_brand_name=b_brand_name,
#             b_class_name=b_class_name,
#             b_medium=b_medium,
#             b_subject=b_subject,
#         )

#         if existing_entries.exists():
#             # Update existing entries
#             for entry in existing_entries:
#                 entry.total_binding= float(total_binding) + float(entry.total_binding)
#                 entry.b_date = formatted_date
#                 entry.save()
#         else:

        
#             # Create new entry
#             data1 = Binding(
#                 b_brand_name=b_brand_name,
#                 b_class_name=b_class_name,
#                 b_medium=b_medium,
#                 b_subject=b_subject,
                
#                 total_binding=total_binding,
#                 b_date=formatted_date,
#                 b_v1_name=b_v1_name,
               
#                 # page=page
#             )
#             data1.save()

#     vend = venders.objects.all()
#     book = books.objects.all()
#     obj1 = books.objects.values('books_name').distinct().order_by('books_name')
#     obj2 = books.objects.values('books_class').distinct().order_by('books_class')


#     data11 = Binding.objects.all()
#     data =Gathering.objects.all()
#     context={
#         'vend':vend,
#         'data':data,
#         'book':book,
#         'data11':data11,
#         'obj1':obj1,
#         'obj2':obj2,
#     }

#     return render(request, 'brands/machine_binding.html',context)



def  machine_binding(request):
    if request.method == 'POST':
        b_v1_name = request.POST.get("b_v1_name")
        b_code = request.POST.get("b_code")
        b_brand_name = request.POST.get("b_brand_name")
        b_class_name = request.POST.get("b_class_name")
        b_medium = request.POST.get("b_medium")
        b_subject = request.POST.get("b_subject")
        
        
        # page = request.POST.get("page")
        total_binding=request.POST.get("total_binding")
        # current_date = timezone.now().date()
        formatted_date = datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        print('lklklkll',formatted_date)

        if b_brand_name.lower() != 'spark':
            
            
          
        
            existing_entries = Binding.objects.filter(
               
                b_brand_name=b_brand_name,
                b_class_name=b_class_name,
                b_medium=b_medium,
                b_subject=b_subject,
            )

            new_entry_needed = True
            for entry in existing_entries:
                if entry.b_v1_name == b_v1_name:
                    entry. b_v1_name= b_v1_name
                    entry.b_code=b_code
                    entry.total_binding= int(total_binding) + int(entry.total_binding)
                    entry.b_date = formatted_date
                    entry.save()
                    new_entry_needed = False
                    break
            
            if new_entry_needed:
            

            
                # Create new entry
             data1 = Binding(
                    b_brand_name=b_brand_name,
                    b_class_name=b_class_name,
                    b_medium=b_medium,
                    b_subject=b_subject,
                    
                    total_binding=total_binding,
                    b_date=formatted_date,
                    b_v1_name=b_v1_name,
                    b_code=b_code,
                   
                    # page=page
                )
             data1.save()
        else:
            error_message = f"Brand name '{b_brand_name}' is not allowed for entry creation."
            sweetify.error(request, error_message, timer=8000, button='OK')

    vend = Bindingvender.objects.all()
    book = books.objects.all()
    obj1 = books.objects.values('books_name').distinct().order_by('books_name')
    obj2 = books.objects.values('books_class').distinct().order_by('books_class')
    obj = books.objects.values('code').distinct().order_by('code')


    data11 = Binding.objects.all().order_by('-id')
    data =Gathering.objects.all()
    context={
        'vend':vend,
        'data':data,
        'book':book,
        'data11':data11,
        'obj1':obj1,
        'obj2':obj2,
        'obj':obj,
     }

    return render(request, 'brands/machine_binding.html',context)


def delete_machine(request, id):
    cus = Binding.objects.get(id=id)
    cus.delete()
    return redirect('/machine_binding/')    


# def edit_machine(request, id):
#     data = Binding.objects.get(id=id)
#     if request.method == 'POST':
#         data.b_v1_name = request.POST.get('b_v1_name')
#         data.b_m1_no = request.POST.get('b_m1_no')
#         data.b_a1_name = request.POST.get('b_a1_name')
#         data.b_brand_name = request.POST.get('b_brand_name')
#         data.b_class_name = request.POST.get('b_class_name')
#         data.b_medium = request.POST.get('b_medium')
#         data.b_subject = request.POST.get('b_subject')
        
#         data.b_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
#         data.total_binding=request.POST.get('total_binding')
        
#         # Convert gather fields to float (modify as needed based on your model)
#         # gather_values = [float(request.POST.get(f"gather{i}", 0)) if request.POST.get(f"gather{i}") else 0 for i in range(1, 11)]
        
#         # total_gather = sum(gather_values)
#         # data.total_gather = total_gather
        
#         data.save()
#         messages.success(request, "Updated Successfully...!!")
#         return redirect('/machine_binding/')
    
#     data2 = venders.objects.all()  # Correct the model name based on your actual model
    
#     context = {'data': data, 'data2': data2}
#     return render(request, "brands/edit_machine.html", context)


def edit_machine(request, id):
    data = Binding.objects.get(id=id)
    if request.method == 'POST':
        data.tra_from_b=request.POST.get("tra_from_b")
        data.to_b=request.POST.get("to_b")
        print("1234",data.to_b)
        data.b_code=request.POST.get("b_code")
        data.voucher_b=request.POST.get("voucher_b")
        data.b_v1_name = request.POST.get('b_v1_name')
        data.b_m1_no = request.POST.get('b_m1_no')
        data.b_a1_name = request.POST.get('b_a1_name')
        data.b_brand_name = request.POST.get('b_brand_name')
        data.b_class_name = request.POST.get('b_class_name')
        data.b_medium = request.POST.get('b_medium')
        data.b_subject = request.POST.get('b_subject')
        data.Material_center = request.POST.get('Material_center')
        
        data.b_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.total_binding=request.POST.get('total_binding')
        data.transfer=request.POST.get("transfer")
        data.pending=request.POST.get("transfer_pending")
        transfer = int(data.transfer) if data.transfer else 0
        # Convert gather fields to float (modify as needed based on your model)
        # gather_values = [float(request.POST.get(f"gather{i}", 0)) if request.POST.get(f"gather{i}") else 0 for i in range(1, 11)]
        
        # total_gather = sum(gather_values)
        # data.total_gather = total_gather
        data.tra_from_b = data.tra_from_b
        print('fghjk',data.tra_from_b)
        data.to_b = data.to_b
        print('fghjk',data.to_b)
        data.Material_center=data.Material_center
        
        
        data.pending=int(data.total_binding) - int(transfer)
        print("ppppppppppeeending",data.pending)
        data.transfer_pending = data.pending
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("binding_approval") == "approval_button":
                Binding.objects.filter(Q(id=id) & Q(binding_approved_1=False)).update(binding_approved_1=True)
                return redirect('/machine_binding/')
               

        
        
        # return redirect('/machine_binding/')
    data.total_transfer_form = data.transfer
    data.total_remaining = data.transfer_pending
    data.save()
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3=Bindingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data,'data3':data3, 'data2': data2,'data5':data5}
    return render(request, "brands/edit_machine.html", context)



def machine_transfer(request):
    # data1 =printing.objects.all().values()
    data1 = Binding.objects.filter(binding_approved_1=True).exclude(Q(b_code__startswith='S') | Q(b_code__startswith='s'))
    print("data1",data1)
   
    template=loader.get_template('brands/machine_transfer.html')
    context={
      'data1':data1,
    }
    return HttpResponse(template.render(context,request))     


# def Manual_info(request):
#     if request.method == 'POST':
#         b_vender_name = request.POST.get("b_vender_name")
#         m_brand_name1 = request.POST.get("m_brand_name1")
#         m_class_name1 = request.POST.get("m_class_name1")
#         m_medium1 = request.POST.get("m_medium1")
#         m_subject1 = request.POST.get("m_subject1")
#         m_striching1 = request.POST.get("m_striching1")
#         m_pasting1 = request.POST.get("m_pasting1")
#         m_cutting1 = request.POST.get("m_cutting1")
#         b_print_date = datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
#         # formatted_date = b_print_date.strftime('%d-%m-%Y')

#         existing_entries = Manual.objects.filter(
#             m_brand_name1=m_brand_name1,
#             m_class_name1=m_class_name1,
#             m_medium1=m_medium1,
#             m_subject1=m_subject1,
#         )

#         new_entry_needed = True
#         for entry in existing_entries:
#             if entry.b_vender_name == b_vender_name:
#                 # Update existing entry
#                 entry.b_vender_name = b_vender_name
#                 entry.m_striching1 = float(m_striching1) + float(entry.m_striching1)
#                 entry.m_pasting1 = float(m_pasting1) + float(entry.m_pasting1)
#                 entry.m_cutting1 = float(m_cutting1) + float(entry.m_cutting1)
#                 entry.b_print_date = b_print_date
#                 entry.save()
#                 new_entry_needed = False
#                 break
        
#         if new_entry_needed:
#             # Create new entry
#             data1 = Manual(
#                 m_brand_name1=m_brand_name1,
#                 m_class_name1=m_class_name1,
#                 m_medium1=m_medium1,
#                 m_subject1=m_subject1,
#                 m_striching1=m_striching1,
#                 m_pasting1=m_pasting1,
#                 m_cutting1=m_cutting1,
#                 b_vender_name=b_vender_name,
#                 b_print_date=b_print_date,
#             )
#             data1.save()

#     vend = venders.objects.all()
#     book = books.objects.all()
#     obj1 = books.objects.values('books_name').distinct().order_by('books_name')
#     obj2 = books.objects.values('books_class').distinct().order_by('books_class')

#     data11 = Manual.objects.all()
    
#     context = {
#         "data11": data11,
#         "vend": vend,
#         "book": book,
#         "obj1":obj1,
#         "obj2":obj2,
#     }
#     return render(request, 'brands/manual_binding.html', context)



def Manual_info(request):
    if request.method == 'POST':
        b_vender_name = request.POST.get("b_vender_name")
        m_brand_name1 = request.POST.get("m_brand_name1")
        m_class_name1 = request.POST.get("m_class_name1")
        m_medium1 = request.POST.get("m_medium1")
        m_subject1 = request.POST.get("m_subject1")
        m_striching1 = request.POST.get("m_striching1")
        # m_pasting1 = request.POST.get("m_pasting1")
        # m_cutting1 = request.POST.get("m_cutting1")
        m_code = request.POST.get("m_code")
        b_print_date = datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        # formatted_date = b_print_date.strftime('%d-%m-%Y')
        if m_brand_name1.lower() == 'spark':
                existing_entries = Manual.objects.filter(
                    
                    m_brand_name1=m_brand_name1,
                    m_class_name1=m_class_name1,
                    m_medium1=m_medium1,
                    m_subject1=m_subject1,
                )

                new_entry_needed = True
                for entry in existing_entries:
                    if entry.b_vender_name == b_vender_name:
                        # Update existing entry
                        entry.b_vender_name = b_vender_name
                        entry.m_code = m_code
                        entry.m_striching1 = int(m_striching1) + int(entry.m_striching1)
                        # entry.m_pasting1 = float(m_pasting1) + float(entry.m_pasting1)
                        # entry.m_cutting1 = float(m_cutting1) + float(entry.m_cutting1)
                        entry.b_print_date = b_print_date
                        entry.save()
                        new_entry_needed = False
                        break
                
                if new_entry_needed:
                    # Create new entry
                 data1 = Manual(
                        m_brand_name1=m_brand_name1,
                        m_class_name1=m_class_name1,
                        m_medium1=m_medium1,
                        m_subject1=m_subject1,
                        m_striching1=m_striching1,
                        # m_pasting1=m_pasting1,
                        # m_cutting1=m_cutting1,
                        b_vender_name=b_vender_name,
                        b_print_date=b_print_date,
                        m_code=m_code,
                 )
                 data1.save()
        else:
            error_message = f"Brand name '{m_brand_name1}' is not allowed for entry creation."
            sweetify.error(request, error_message, timer=8000, button='OK')     

    vend = Bindingvender.objects.all()
    book = books.objects.all()
    obj1 = books.objects.values('books_name').distinct().order_by('books_name')
    obj = books.objects.values('code').distinct().order_by('code')
    obj2 = books.objects.values('books_class').distinct().order_by('books_class')

    data11 = Manual.objects.all().order_by('-id')
    
    context = {
        "data11": data11,
        "vend": vend,
        "book": book,
        "obj1":obj1,
        "obj2":obj2,
        "obj":obj,
    }
    return render(request, 'brands/manual_binding.html', context)



def delete_manual(request, id):
    cus = Manual.objects.get(id=id)
    cus.delete()
    return redirect('/manual-info/')    


# def edit_manual(request, id):
#     data = Manual.objects.get(id=id)
#     if request.method == 'POST':
#         data.b_vender_name = request.POST.get('b_vender_name')
#         data.b_mobile_no = request.POST.get('b_mobile_no')
#         data.b_adress_name = request.POST.get('b_adress_name')
#         data.m_brand_name1 = request.POST.get('m_brand_name1')
#         data.m_class_name1 = request.POST.get('m_class_name1')
#         data.m_medium1 = request.POST.get('m_medium1')
#         data.m_subject1 = request.POST.get('m_subject1')
#         data.m_striching1 = request.POST.get('m_striching1')
#         data.m_pasting1 = request.POST.get('m_pasting1')
#         data.m_cutting1 = request.POST.get('m_cutting1')
       
#         data.b_print_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        
#         # Convert gather fields to float (modify as needed based on your model)
        
        
#         data.save()
#         messages.success(request, "Updated Successfully...!!")
#         return redirect('/manual-info/')
    
#     data2 = venders.objects.all()  # Correct the model name based on your actual model
    
#     context = {'data': data, 'data2': data2}
#     return render(request, "brands/edit_manual.html", context)



def edit_manual(request, id):
    data = Manual.objects.get(id=id)
    if request.method == 'POST':
        
        data.tra_from_m = request.POST.get('tra_from_m')
        data.to_m= request.POST.get('to_m')
        data.voucher_m = request.POST.get('voucher_m')
        data.m_code = request.POST.get('m_code')
        data.b_vender_name = request.POST.get('b_vender_name')
        data.b_mobile_no = request.POST.get('b_mobile_no')
        data.b_adress_name = request.POST.get('b_adress_name')
        data.m_brand_name1 = request.POST.get('m_brand_name1')
        data.m_class_name1 = request.POST.get('m_class_name1')
        data.m_medium1 = request.POST.get('m_medium1')
        data.m_subject1 = request.POST.get('m_subject1')
        data.m_striching1 = request.POST.get('m_striching1')
        data.transfer=request.POST.get("transfer")
        data.pending=request.POST.get("transfer_pending")
        data.Material_center = request.POST.get('Material_center')
        transfer = int(data.transfer) if data.transfer else 0
        # data.m_pasting1 = request.POST.get('m_pasting1')
        # data.m_cutting1 = request.POST.get('m_cutting1')
       
        data.b_print_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        
        # Convert gather fields to float (modify as needed based on your model)
        data.tra_from_m = data.tra_from_m
        data.Material_center =data.Material_center
        print('fghjk',data.tra_from_m)
        data.to_m = data.to_m
        print('fghjk',data.to_m)
        data.pending=int(data.m_striching1) - int(transfer)
        print("ppppppppppeeending",data.pending)
        data.transfer_pending = data.pending
        
        
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("stiching_approval") == "approval_button":
                Manual.objects.filter(Q(id=id) & Q(stiching_approved_1=False)).update(stiching_approved_1=True)
                return redirect('/manual-info/')
               
        
       
    data.total_transfer_form = data.transfer
    data.total_remaining = data.transfer_pending
    data.save()   
    
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3=Bindingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data5':data5}
    return render(request, "brands/edit_manual.html", context)




def gather_transfer_manual(request):
    # data1 =printing.objects.all().values()
    brand_name = "Spark"  # Aap jo brand_name chahte hain, usko yahaan par set karein
    
    # Data ko filter karein jahaan brand_name Spark ho
    data1 = Gathering.objects.filter(gather_approved_1=True, brand_name=brand_name)
    print("data1",data1)
   
    template=loader.get_template('brands/gather_transfer_manual.html')
    context={
      'data1':data1,
    }
    return HttpResponse(template.render(context,request)) 


    
from itertools import chain
    
def vouchers_1(request):
     if request.method =='POST':
    
         series = request.POST.get('series')
         date = request.POST.get('date')
         voucher_no = request.POST.get('voucher_no')
         voucher_type =request.POST.get('voucher_type')
         print("voucher type",voucher_type)

         account=request.POST.get('account')
         debit =request.POST.get('debit')
         credit =request.POST.get('credit')
         short_narration=request.POST.get('short_narration')
         dc=request.POST.get('dc')
          
         account1=request.POST.get('account1')
         debit1 =request.POST.get('debit1')
         credit1 =request.POST.get('credit1')
         short_narration1=request.POST.get('short_narration1')
         dc1=request.POST.get('dc1')

         account2=request.POST.get('account2')
         debit2 =request.POST.get('debit2')
         credit2 =request.POST.get('credit2')
         short_narration2=request.POST.get('short_narration2')
         dc2=request.POST.get('dc2')

         account3=request.POST.get('account3')
         debit3 =request.POST.get('debit3')
         credit3 =request.POST.get('credit3')
         short_narration3=request.POST.get('short_narration3')
         dc3=request.POST.get('dc3')

         account4=request.POST.get('account4')
         debit4 =request.POST.get('debit4')
         credit4 =request.POST.get('credit4')
         short_narration4=request.POST.get('short_narration4')
         dc4=request.POST.get('dc4')

         account5=request.POST.get('account5')
         debit5 =request.POST.get('debit5')
         credit5 =request.POST.get('credit5')
         short_narration5=request.POST.get('short_narration5')
         dc5=request.POST.get('dc5')

         account6=request.POST.get('account6')
         debit6 =request.POST.get('debit6')
         credit6 =request.POST.get('credit6')
         short_narration6=request.POST.get('short_narration6')
         dc6=request.POST.get('dc6')

         account7=request.POST.get('account7')
         debit7 =request.POST.get('debit7')
         credit7 =request.POST.get('credit7')
         short_narration7=request.POST.get('short_narration7')
         dc7=request.POST.get('dc7')

         account8=request.POST.get('account8')
         debit8 =request.POST.get('debit8')
         credit8 =request.POST.get('credit8')
         short_narration8=request.POST.get('short_narration8')
         dc8=request.POST.get('dc8')

         account9=request.POST.get('account9')
         debit9 =request.POST.get('debit9')
         credit9 =request.POST.get('credit9')
         short_narration9=request.POST.get('short_narration9')
         dc9=request.POST.get('dc9')

         account10=request.POST.get('account10')
         debit10 =request.POST.get('debit10')
         credit10 =request.POST.get('credit10')
         short_narration10=request.POST.get('short_narration10')
         dc10=request.POST.get('dc10')

         account11=request.POST.get('account11')
         debit11 =request.POST.get('debit11')
         credit11 =request.POST.get('credit11')
         short_narration11=request.POST.get('short_narration11')
         dc11=request.POST.get('dc11')

         account12=request.POST.get('account12')
         debit12 =request.POST.get('debit12')
         credit12 =request.POST.get('credit12')
         short_narration12=request.POST.get('short_narration12')
         dc12=request.POST.get('dc12')

         account13=request.POST.get('account13')
         debit13 =request.POST.get('debit13')
         credit13 =request.POST.get('credit13')
         short_narration13=request.POST.get('short_narration13')
         dc13=request.POST.get('dc13')

         account14=request.POST.get('account14')
         debit14 =request.POST.get('debit14')
         credit14 =request.POST.get('credit14')
         short_narration14=request.POST.get('short_narration14')
         dc14=request.POST.get('dc14')

         account15=request.POST.get('account15')
         debit15 =request.POST.get('debit15')
         credit15 =request.POST.get('credit15')
         short_narration15=request.POST.get('short_narration15')
         dc15=request.POST.get('dc15')



         if vouchers.objects.filter(voucher_no=voucher_no).exists():
            error_message = f"A Voucher No with the number='{voucher_no}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
         else:
            voucher_data = vouchers(date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),short_narration=short_narration,debit=debit,credit=credit,series=series,voucher_no=voucher_no,voucher_type=voucher_type,
                                    account=account,
                                    dc=dc,
                                    
                                    short_narration1=short_narration1,
                                    debit1=debit1,
                                    credit1=credit1,
                                    account1=account1,
                                    dc1=dc1,
                                    
                                    short_narration2=short_narration2,
                                    debit2=debit2,
                                    credit2=credit2,
                                    account2=account2,
                                    dc2=dc2,
                                    
                                    short_narration3=short_narration3,
                                    debit3=debit3,
                                    credit3=credit3,
                                    account3=account3,
                                    dc3=dc3,
                                    
                                    short_narration4=short_narration4,
                                    debit4=debit4,
                                    credit4=credit4,
                                    account4=account4,
                                    dc4=dc4,
                                    
                                    short_narration5=short_narration5,
                                    debit5=debit5,
                                    credit5=credit5,
                                    account5=account5,
                                    dc5=dc5,
                                    
                                    short_narration6=short_narration6,
                                    debit6=debit6,
                                    credit6=credit6,
                                    account6=account6,
                                    dc6=dc6,
                                    
                                    short_narration7=short_narration7,
                                    debit7=debit7,
                                    credit7=credit7,
                                    account7=account7,
                                    dc7=dc7,
                                    
                                    short_narration8=short_narration8,
                                    debit8=debit8,
                                    credit8=credit8,
                                    account8=account8,
                                    dc8=dc8,
                                    
                                    short_narration9=short_narration9,
                                    debit9=debit9,
                                    credit9=credit9,
                                    account9=account9,
                                    dc9=dc9,
                                    
                                    short_narration10=short_narration10,
                                    debit10=debit10,
                                    credit10=credit10,
                                    account10=account10,
                                    dc10=dc10,
                                    
                                    short_narration11=short_narration11,
                                    debit11=debit11,
                                    credit11=credit11,
                                    account11=account11,
                                    dc11=dc11,
                                    
                                    short_narration12=short_narration12,
                                    debit12=debit12,
                                    credit12=credit12,
                                    account12=account12,
                                    dc12=dc12,
                                    
                                    short_narration13=short_narration13,
                                    debit13=debit13,
                                    credit13=credit13,
                                    account13=account13,
                                    dc13=dc13,
                                    
                                    short_narration14=short_narration14,
                                    debit14=debit14,
                                    credit14=credit14,
                                    account14=account14,
                                    dc14=dc14,
                                    
                                    short_narration15=short_narration15,
                                    debit15=debit15,
                                    credit15=credit15,
                                    account15=account15,
                                    dc15=dc15,)
            voucher_data.save()           
            sweetify.success(request, "Vouchers Created Successfully.", timer=6000,button='OK')
    
                    
     data = vouchers.objects.all().order_by('-id')
     
    
     data_combined = list(chain(
        venders.objects.all(),
        Binder_ledger.objects.all(),
        Gather_ledger.objects.all(),
        Stitcher_ledger.objects.all(),
        Paster_ledger.objects.all(),
        Machine_ledger.objects.all(),
        Cutter_ledger.objects.all(),
        Packer_ledger.objects.all(),
        General_ledger.objects.all(),
        Customer_ledger.objects.all()
      ))
     
    
    
    
    
    
    
     return render(request,'vouchers/vouchers.html',{'data':data,'data_combined':data_combined})

@login_required(login_url="")
def edit_voucher(request, id):
    data = vouchers.objects.get(id=id)
    if request.method == 'POST':
        data.series = request.POST.get('series')
        data.voucher_no = request.POST.get('voucher_no')
        data.tax_type = request.POST.get('tax_type')
        data.party = request.POST.get('party')
        data.material_centre = request.POST.get('material_centre')
        # if vouchers.objects.filter(series__iexact=data.series).exists():
        #     error_message = f"A centre with the name '{data.series}' already exists."
        #     return HttpResponseBadRequest(error_message)
        # else:
        #     data.save()
        #     messages.success(request, "Updated Successfully...!!")
        data.save()
        return redirect('/vouchers/')
    
    context = {'data':data,}
    return render(request, 'vouchers/edit_vouchers.html', context)


@login_required(login_url="") 
def delete_voucher(request, id):
    cus = vouchers.objects.get(id=id)
    cus.delete()
    return redirect('/vouchers/')  

def view_voucher(request, id):
    data = vouchers.objects.get(id=id)
    
    
    
    data.save()
    
    context = {
        'data': data,
        
    }

    return render(request, 'vouchers/view_voucher.html',context)




def pasting(request):
    if request.method == 'POST':
        pasting_vender_name = request.POST.get("pasting_vender_name")
        pasting_brand_name1 = request.POST.get("pasting_brand_name1")
        pasting_class_name1 = request.POST.get("pasting_class_name1")
        pasting_medium1 = request.POST.get("pasting_medium1")
        pasting_subject1 = request.POST.get("pasting_subject1")
        # pasting_striching1 = request.POST.get("pasting_striching1")
        pasting_pasting1 = request.POST.get("pasting_pasting1")
        # pasting_cutting1 = request.POST.get("pasting_cutting1")
        pasting_code = request.POST.get("pasting_code")
        pasting_print_date = datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        # formatted_date = b_print_date.strftime('%d-%m-%Y')
        if pasting_brand_name1.lower() == 'spark':
            existing_entries = Pasting.objects.filter(
                
                pasting_brand_name1=pasting_brand_name1,
                pasting_class_name1=pasting_class_name1,
                pasting_medium1=pasting_medium1,
                pasting_subject1=pasting_subject1,
            )

            new_entry_needed = True
            for entry in existing_entries:
                if entry.pasting_vender_name == pasting_vender_name:
                    # Update existing entry
                    entry.pasting_vender_name = pasting_vender_name
                    entry.pasting_code = pasting_code
                    # entry.pasting_striching1 = float(pasting_striching1) + float(entry.pasting_striching1)
                    entry.pasting_pasting1 = float(pasting_pasting1) + float(entry.pasting_pasting1)
                    # entry.pasting_cutting1 = float(pasting_cutting1) + float(entry.pasting_cutting1)
                    entry.pasting_print_date = pasting_print_date
                    entry.save()
                    new_entry_needed = False
                    break
            
            if new_entry_needed:
                # Create new entry
             data1 = Pasting(
                    pasting_brand_name1=pasting_brand_name1,
                    pasting_class_name1=pasting_class_name1,
                    pasting_medium1=pasting_medium1,
                    pasting_subject1=pasting_subject1,
                    # pasting_striching1=pasting_striching1,
                    pasting_pasting1=pasting_pasting1,
                    # pasting_cutting1=pasting_cutting1,
                    pasting_vender_name=pasting_vender_name,
                    pasting_print_date=pasting_print_date,
                    pasting_code=pasting_code,
             )
             data1.save()
        else:
            error_message = f"Brand name '{pasting_brand_name1}' is not allowed for entry creation."
            sweetify.error(request, error_message, timer=8000, button='OK')
    vend = Bindingvender.objects.all()
    book = books.objects.all()
    obj1 = books.objects.values('books_name').distinct().order_by('books_name')
    obj = books.objects.values('code').distinct().order_by('code')
    obj2 = books.objects.values('books_class').distinct().order_by('books_class')

    data11 = Pasting.objects.all().order_by('-id')
    
    context = {
        "data11": data11,
        "vend": vend,
        "book": book,
        "obj1":obj1,
        "obj2":obj2,
        "obj":obj,
    }
    return render(request, 'brands/pasting.html', context)


def delete_pasting(request, id):
    cus = Pasting.objects.get(id=id)
    cus.delete()
    return redirect('/pasting/')    


def edit_pasting(request, id):
    data = Pasting.objects.get(id=id)
    if request.method == 'POST':
        data.tra_from_pasting = request.POST.get('tra_from_pasting')
        data.to_pasting= request.POST.get('to_pasting')
        data.voucher_pasting = request.POST.get('voucher_pasting')
        
        data.pasting_code = request.POST.get('pasting_code')
        data.pasting_vender_name = request.POST.get('pasting_vender_name')
        data.pasting_mobile_no = request.POST.get('pasting_mobile_no')
        data.pasting_adress_name = request.POST.get('pasting_adress_name')
        data.pasting_brand_name1 = request.POST.get('pasting_brand_name1')
        data.pasting_class_name1 = request.POST.get('pasting_class_name1')
        data.pasting_medium1 = request.POST.get('pasting_medium1')
        data.pasting_subject1 = request.POST.get('pasting_subject1')
        data.Material_center = request.POST.get('Material_center')
        # data.pasting_striching1 = request.POST.get('pasting_striching1')
        data.pasting_pasting1 = request.POST.get('pasting_pasting1')
        data.transfer=request.POST.get("transfer")
        data.pending=request.POST.get("transfer_pending")
        transfer = int(data.transfer) if data.transfer else 0


        # data.pasting_cutting1 = request.POST.get('pasting_cutting1')
       
        data.pasting_print_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        
        # Convert gather fields to float (modify as needed based on your model)
        data.tra_from_pasting = data.tra_from_pasting
        print('fghjk',data.tra_from_pasting)
        data.to_pasting= data.to_pasting
        print('fghjk',data.to_pasting)
        data.pending=int(data.pasting_pasting1) - int(transfer)
        print("ppppppppppeeending",data.pending)
        data.transfer_pending = data.pending
        data.Material_center = data.Material_center 
        
        
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("pasting_approval") == "approval_button":
                Pasting.objects.filter(Q(id=id) & Q(pasting_approved_1=False)).update(pasting_approved_1=True)
                return redirect('/pasting/')
        
        # return redirect('/pasting/')
    data.total_transfer_form = data.transfer
    data.total_remaining = data.transfer_pending
    data.save()
    
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3 = Cuttingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data5':data5}
    return render(request, "brands/edit_pasting.html", context)




def cutting_manual(request):
    if request.method == 'POST':
        cutting_vender_name = request.POST.get("cutting_vender_name")
        cutting_brand_name1 = request.POST.get("cutting_brand_name1")
        cutting_class_name1 = request.POST.get("cutting_class_name1")
        cutting_medium1 = request.POST.get("cutting_medium1")
        cutting_subject1 = request.POST.get("cutting_subject1")
        # pasting_striching1 = request.POST.get("pasting_striching1")
        # cutting_pasting1 = request.POST.get("cutting_pasting1")
        cutting_cutting1 = request.POST.get("cutting_cutting1")
        cutting_code = request.POST.get("cutting_code")
        cutting_print_date = datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        # formatted_date = b_print_date.strftime('%d-%m-%Y')

        existing_entries = Manual_cutting.objects.filter(
            
            cutting_brand_name1=cutting_brand_name1,
            cutting_class_name1=cutting_class_name1,
            cutting_medium1=cutting_medium1,
            cutting_subject1=cutting_subject1,
        )

        new_entry_needed = True
        for entry in existing_entries:
            if entry.cutting_vender_name == cutting_vender_name:
                # Update existing entry
                entry.cutting_vender_name = cutting_vender_name
                entry.cutting_code = cutting_code
                # entry.pasting_striching1 = float(pasting_striching1) + float(entry.pasting_striching1)
                # entry.cutting_pasting1 = float(pasting_pasting1) + float(entry.pasting_pasting1)
                entry.cutting_cutting1 = float(cutting_cutting1) + float(entry.cutting_cutting1)
                entry.cutting_print_date = cutting_print_date
                entry.save()
                new_entry_needed = False
                break
        
        if new_entry_needed:
            # Create new entry
         data1 = Manual_cutting(
                cutting_brand_name1=cutting_brand_name1,
                cutting_class_name1=cutting_class_name1,
                cutting_medium1=cutting_medium1,
                cutting_subject1=cutting_subject1,
                # pasting_striching1=pasting_striching1,
                # cutting_pasting1=cutting_pasting1,
                cutting_cutting1=cutting_cutting1,
                cutting_vender_name=cutting_vender_name,
                cutting_print_date=cutting_print_date,
                cutting_code=cutting_code,
         )
         data1.save()

    vend = venders.objects.all()
    book = books.objects.all()
    obj1 = books.objects.values('books_name').distinct().order_by('books_name')
    obj = books.objects.values('code').distinct().order_by('code')
    obj2 = books.objects.values('books_class').distinct().order_by('books_class')

    data11 = Manual_cutting.objects.all().order_by('-id')
    
    context = {
        "data11": data11,
        "vend": vend,
        "book": book,
        "obj1":obj1,
        "obj2":obj2,
        "obj":obj,
    }
    return render(request, 'brands/cutting_manual.html', context)


def delete_manual_cutting(request, id):
    cus = Manual_cutting.objects.get(id=id)
    cus.delete()
    return redirect('/cutting_manual/')    


def edit_manual_cutting(request, id):
    data = Manual_cutting.objects.get(id=id)
    if request.method == 'POST':
        data.tra_from_cutting = request.POST.get('tra_from_cutting')
        data.to_cutting= request.POST.get('to_cutting')
        data.voucher_cutting = request.POST.get('voucher_cutting')
        data.cutting_code = request.POST.get('cutting_code')
        data.cutting_vender_name = request.POST.get('cutting_vender_name')
        data.cutting_mobile_no = request.POST.get('cutting_mobile_no')
        data.cutting_adress_name = request.POST.get('cutting_adress_name')
        data.cutting_brand_name1 = request.POST.get('cutting_brand_name1')
        data.cutting_class_name1 = request.POST.get('cutting_class_name1')
        data.cutting_medium1 = request.POST.get('cutting_medium1')
        data.cutting_subject1 = request.POST.get('cutting_subject1')
       
        data.cutting_cutting1 = request.POST.get('cutting_cutting1')
       
        data.cutting_print_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.transfer=request.POST.get("transfer")
        data.pending=request.POST.get("transfer_pending")
        transfer = int(data.transfer) if data.transfer else 0
        
        # Convert gather fields to float (modify as needed based on your model)
        data.tra_from_cutting = data.tra_from_cutting
        print('fghjk',data.tra_from_cutting)
        data.to_cutting= data.to_cutting
        print('fghjk',data.to_cutting)
        data.pending=int(data.cutting_cutting1) - int(transfer)
        print("ppppppppppeeending",data.pending)
        data.transfer_pending = data.pending
        
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("cutting_approval") == "approval_button":
                Manual_cutting.objects.filter(Q(id=id) & Q(cutting_approved_1=False)).update(cutting_approved_1=True)
                return redirect('/cutting_manual/')
    
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    
    context = {'data': data, 'data2': data2}
    return render(request, "brands/edit_cutting_manual.html", context)





def stiching_transfer_pasting(request):
    data1 = Manual.objects.filter(stiching_approved_1=True, m_code__istartswith='S')
    print("data1",data1)
   
    template=loader.get_template('brands/stiching_transfer_pasting.html')
    context={
      'data1':data1,
    }
    return HttpResponse(template.render(context,request)) 


def pasting_transfer_cutting(request):
    data1 = Pasting.objects.filter(pasting_approved_1=True, pasting_code__istartswith='S')
    print("data1",data1)
   
    template=loader.get_template('brands/pasting_transfer_cutting.html')
    context={
      'data1':data1,
    }
    return HttpResponse(template.render(context,request)) 


def cutting_manual_transfer_packing(request):
    data1 = Manual_cutting.objects.filter(cutting_approved_1=True, cutting_code__istartswith='S')
    print("data1",data1)
   
    template=loader.get_template('brands/cutting_manual_transfer_packing.html')
    context={
      'data1':data1,
    }
    return HttpResponse(template.render(context,request)) 
    





def cover_printing(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        plate_no = request.POST.get('plate_no')
        paper_to_be_used = request.POST.get('paper_to_be_used')  # Convert to float
        width = request.POST.get('width')  # Convert to float
        length = request.POST.get('length')  # Convert to float
        unit = request.POST.get('unit')
        no_of_ups= request.POST.get('no_of_ups')
        subject1=request.POST.get('subject1') 
        subject2= request.POST.get('subject2')
        subject3= request.POST.get('subject3')
        subject4= request.POST.get('subject4')
        subject5= request.POST.get('subject5')
        subject6= request.POST.get('subject6')
        subject7= request.POST.get('subject7')
        subject8= request.POST.get('subject8')
        subject9= request.POST.get('subject9')
        subject10= request.POST.get('subject10')
        subject11= request.POST.get('subject11')
        subject12= request.POST.get('subject12')
        subject13= request.POST.get('subject13')
        subject14= request.POST.get('subject14')
        subject15= request.POST.get('subject15')
        subject16= request.POST.get('subject16')
        plate_rate= request.POST.get('plate_rate')
        printing_rate= request.POST.get('printing_rate')
        tax= request.POST.get('tax')
        lamination= request.POST.get('lamination')
        rate1= request.POST.get('rate1')
        rate2= request.POST.get('rate2')
        bill_company= request.POST.get('bill_company')
        texture= request.POST.get('texture')
        # per_kg_weight = round(float(form_length * form_width * form_gsm) / 20000,2)
        # print('eeeeeeeee', per_kg_weight)
        # wastage_percent = round(float(per_kg_weight) + (float(per_kg_weight)* float(no_of_forms/100)),2)
        # p_raw_material = request.POST.get('p_raw_material')
        # material_center = request.POST.get('material_center')
        # form_received_by = request.POST.get('form_received_by')
        # purchase_quantity = round(float(request.POST.get('purchase_quantity')),2)
        # rimm = round(purchase_quantity / wastage_percent,2)
        # gross = round(rimm * 3.47,2)
        # single_rate = float(request.POST.get('single_rate'))
        # double_rate = float(request.POST.get('double_rate'))
        quantity_cover = request.POST.get('quantity_cover')
        book_data = Cover_Printing(
            
            plate_no=plate_no,
            paper_to_be_used=paper_to_be_used,
            width=width,
            length=length,
            unit=unit,
            no_of_ups= no_of_ups,
            subject1=subject1,
            subject2=subject2,
            subject3=subject3,
            subject4=subject4,
            subject5=subject5,
            subject6=subject6,
            subject7=subject7,
            subject8=subject8,
            subject9=subject9,
            subject10=subject10,
            subject11=subject11,
            subject12=subject12,
            subject13=subject13,
            subject14=subject14,
            subject15=subject15,
            subject16=subject16,
            plate_rate=plate_rate,
            printing_rate=printing_rate,
            tax=tax,
            lamination=lamination,
            rate1=rate1,
            rate2=rate2,
            bill_company=bill_company,
            texture=texture,
            quantity_cover=quantity_cover,

        )
        book_data.save()
    obj=books.objects.all()
    obj1=Company.objects.all()
    data = Cover_Printing.objects.all().order_by('-id')
    obj2 =forms.objects.all()
    return render(request, 'brands/cover_printing.html', {'data': data,'obj':obj,'obj1':obj1,'obj2':obj2})  


def delete_cover(request, id):
    cus = Cover_Printing.objects.get(id=id)
    cus.delete()
    return redirect('/cover_printing/')

@login_required(login_url="")
def edit_cover(request, id):
    data = Cover_Printing.objects.get(id=id)
    if request.method == 'POST':
        # data.from_vend=request.POST.get("from_vend")
        # data.to_vend=request.POST.get("to_vend")
        data.plate_no= request.POST.get('plate_no')
        data.paper_to_be_used= request.POST.get('paper_to_be_used')
        data.width= request.POST.get('width')
        data.length= request.POST.get('length')
        data.no_of_ups= request.POST.get('no_of_ups')
        data.subject1= request.POST.get('subject1')
        data.subject2= request.POST.get('subject2')
        data.subject3= request.POST.get('subject3')
        data.subject4= request.POST.get('subject4')
        data.subject5= request.POST.get('subject5')
        data.subject6= request.POST.get('subject6')
        data.subject7= request.POST.get('subject7')
        data.subject8= request.POST.get('subject8')
        data.subject9= request.POST.get('subject9')
        data.subject10= request.POST.get('subject10')
        data.subject11= request.POST.get('subject11')
        data.subject12= request.POST.get('subject12')
        data.subject13= request.POST.get('subject13')
        data.subject14= request.POST.get('subject14')
        data.subject15= request.POST.get('subject15')
        data.subject16= request.POST.get('subject16')
        data.plate_rate= request.POST.get('plate_rate')
        data.printing_rate= request.POST.get('printing_rate')
        data.bill_company = request.POST.get('bill_company')
        data.quantity_cover = request.POST.get('quantity_cover')
        print("11111111111111",data.quantity_cover)
        # data.from_vend = data.from_vend
        # print('fghjk',data.from_vend)
        # data.to_vend = data.to_vend
        # data.transfer=request.POST.get("transfer")
        # data.pending=request.POST.get("transfer_pending")
        # transfer = int(data.transfer) if data.transfer else 0
        # data.pending=int(data.quantity_cover) - int(transfer)
        # print("ppppppppppeeending",data.pending)
        # data.transfer_pending = data.pending
        
        data.save()
        # if request.method == 'POST':
        #     if request.POST.get("cover_transfer") == "approval_button":
        #         Cover_Printing.objects.filter(Q(id=id) & Q(cover_transfer=False)).update(cover_transfer=True)
        #         return redirect('/cover_printing/')
        #     else:
        #          pass  
        #         # return redirect('/cover_printing/')
    obj=books.objects.all()
    context = {'data':data,'obj':obj,}
    return render(request, 'brands/edit_cover.html', context)




def cover_order(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        material_center = request.POST.get('material_center')
        self_company = request.POST.get('self_company')
        vendor_order = request.POST.get('vendor_order')
        subject_order = request.POST.get('subject_order')  # Convert to float
        date_order = datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')  # Convert to float
        quantity_order =request.POST.get('quantity_order')
        rate_order=request.POST.get('rate_order')
        palte_order= request.POST.get('palte_order')
        lamination_sheet= request.POST.get('lamination_sheet')
        texture_sheet= request.POST.get('texture_sheet')
        printing_sheet= request.POST.get('printing_sheet')
        amount= request.POST.get('amount') 
        
        quantity_order1 =request.POST.get('quantity_order1')
        rate_order1=request.POST.get('rate_order1')
        plate_order1= request.POST.get('plate_order1')
        amount1= request.POST.get('amount1') # Convert to float
        lamination_sheet1= request.POST.get('lamination_sheet1')
        texture_sheet1= request.POST.get('texture_sheet1')
        printing_sheet1= request.POST.get('printing_sheet1')

        quantity_order2 =request.POST.get('quantity_order2')
        rate_order2=request.POST.get('rate_order2')
        plate_order2= request.POST.get('plate_order2')
        amount2= request.POST.get('amount2')
        lamination_sheet2= request.POST.get('lamination_sheet2')
        texture_sheet2= request.POST.get('texture_sheet2')
        printing_sheet2= request.POST.get('printing_sheet2')

        quantity_order3 =request.POST.get('quantity_order3')
        rate_order3=request.POST.get('rate_order3')
        plate_order3= request.POST.get('plate_order3')
        amount3= request.POST.get('amount3') 
        lamination_sheet3= request.POST.get('lamination_sheet3')
        texture_sheet3= request.POST.get('texture_sheet3')
        printing_sheet3= request.POST.get('printing_sheet3')

        quantity_order4 =request.POST.get('quantity_order4')
        rate_order4=request.POST.get('rate_order4')
        plate_order4= request.POST.get('plate_order4')
        amount4= request.POST.get('amount4')
        lamination_sheet4= request.POST.get('lamination_sheet4')
        texture_sheet4= request.POST.get('texture_sheet4')
        printing_sheet4= request.POST.get('printing_sheet4')

        quantity_order5 =request.POST.get('quantity_order5')
        rate_order5=request.POST.get('rate_order5')
        plate_order5= request.POST.get('plate_order5')
        amount5= request.POST.get('amount5')
        lamination_sheet5= request.POST.get('lamination_sheet5')
        texture_sheet5= request.POST.get('texture_sheet5')
        printing_sheet5= request.POST.get('printing_sheet5')

        quantity_order6 =request.POST.get('quantity_order6')
        rate_order6=request.POST.get('rate_order6')
        plate_order6= request.POST.get('plate_order6')
        amount6= request.POST.get('amount6')
        lamination_sheet6= request.POST.get('lamination_sheet6')
        texture_sheet6= request.POST.get('texture_sheet6')
        printing_sheet6= request.POST.get('printing_sheet6')

        quantity_order7 =request.POST.get('quantity_order7')
        rate_order7=request.POST.get('rate_order7')
        plate_order7= request.POST.get('plate_order7')
        amount7= request.POST.get('amount7')
        lamination_sheet7= request.POST.get('lamination_sheet7')
        texture_sheet7= request.POST.get('texture_sheet7')
        printing_sheet7= request.POST.get('printing_sheet7')

        quantity_order8 =request.POST.get('quantity_order8')
        rate_order8=request.POST.get('rate_order8')
        plate_order8= request.POST.get('plate_order8')
        amount8= request.POST.get('amount8')
        lamination_sheet8= request.POST.get('lamination_sheet8')
        texture_sheet8= request.POST.get('texture_sheet8')
        printing_sheet8= request.POST.get('printing_sheet8')

        quantity_order9 =request.POST.get('quantity_order9')
        rate_order9=request.POST.get('rate_order9')
        plate_order9= request.POST.get('plate_order9')
        amount9= request.POST.get('amount9')
        lamination_sheet9= request.POST.get('lamination_sheet9')
        texture_sheet9= request.POST.get('texture_sheet9')
        printing_sheet9= request.POST.get('printing_sheet9')

        quantity_order10 =request.POST.get('quantity_order10')
        rate_order10=request.POST.get('rate_order10')
        plate_order10= request.POST.get('plate_order10')
        amount10= request.POST.get('amount10')
        lamination_sheet10= request.POST.get('lamination_sheet10')
        texture_sheet10= request.POST.get('texture_sheet10')
        printing_sheet10= request.POST.get('printing_sheet10')

        bill_sundries_1 = request.POST.get("bill_sundries_1")
        narration_1= request.POST.get("narration_1")
        amount_1= request.POST.get("amount_1")
        percent_1=request.POST.get("percent_1")

        bill_sundries_2 = request.POST.get("bill_sundries_2")
        narration_2= request.POST.get("narration_2")
        amount_2= request.POST.get("amount_2")
        percent_2=request.POST.get("percent_2")

        bill_sundries_3 = request.POST.get("bill_sundries_3")
        narration_3= request.POST.get("narration_3")
        amount_3= request.POST.get("amount_3")
        percent_3=request.POST.get("percent_3")

        bill_sundries_4 = request.POST.get("bill_sundries_4")
        narration_4= request.POST.get("narration_4")
        amount_4= request.POST.get("amount_4")
        percent_4=request.POST.get("percent_4")

        bill_sundries_5 = request.POST.get("bill_sundries_5")
        narration_5= request.POST.get("narration_5")
        amount_5= request.POST.get("amount_5")
        percent_5=request.POST.get("percent_5")

        
        
        


        


        data = Cover_order(

            material_center = material_center,
            self_company=self_company,
            vendor_order=vendor_order,
            date_order=date_order,
            quantity_order=quantity_order,
            subject_order=subject_order,
            rate_order=rate_order,
            palte_order=palte_order,
            amount=amount,

            quantity_order1=quantity_order1,
            rate_order1=rate_order1,
            plate_order1=plate_order1,
            amount1=amount1,
            quantity_order2=quantity_order2,
            rate_order2=rate_order2,
            plate_order2=plate_order2,
            amount2=amount2,
                 
            quantity_order3=quantity_order3,
            rate_order3=rate_order3,
            plate_order3=plate_order3,
            amount3=amount3,

            quantity_order4=quantity_order4,
            rate_order4=rate_order4,
            plate_order4=plate_order4,
            amount4=amount4,

            quantity_order5=quantity_order5,
            rate_order5=rate_order5,
            plate_order5=plate_order5,
            amount5=amount5,

            quantity_order6=quantity_order6,
            rate_order6=rate_order6,
            plate_order6=plate_order6,
            amount6=amount6,

            quantity_order7=quantity_order7,
            rate_order7=rate_order7,
            plate_order7=plate_order7,
            amount7=amount7,

            quantity_order8=quantity_order8,
            rate_order8=rate_order8,
            plate_order8=plate_order8,
            amount8=amount8,

            quantity_order9=quantity_order9,
            rate_order9=rate_order9,
            plate_order9=plate_order9,
            amount9=amount9,

            quantity_order10=quantity_order10,
            rate_order10=rate_order10,
            plate_order10=plate_order10,
            amount10=amount10,
            lamination_sheet=lamination_sheet,
            texture_sheet=texture_sheet,
            printing_sheet=printing_sheet,
            lamination_sheet1=lamination_sheet1,
            texture_sheet1=texture_sheet1,
            printing_sheet1=printing_sheet1,
            lamination_sheet2=lamination_sheet2,
            texture_sheet2=texture_sheet2,
            printing_sheet2=printing_sheet2,
            lamination_sheet3=lamination_sheet3,
            texture_sheet3=texture_sheet3,
            printing_sheet3=printing_sheet3,
            lamination_sheet4=lamination_sheet4,
            texture_sheet4=texture_sheet4,
            printing_sheet4=printing_sheet4,
            lamination_sheet5=lamination_sheet5,
            texture_sheet5=texture_sheet5,
            printing_sheet5=printing_sheet5,
            lamination_sheet6=lamination_sheet6,
            texture_sheet6=texture_sheet6,
            printing_sheet6=printing_sheet6,
            lamination_sheet7=lamination_sheet7,
            texture_sheet7=texture_sheet7,
            printing_sheet7=printing_sheet7,
            lamination_sheet8=lamination_sheet8,
            texture_sheet8=texture_sheet8,
            printing_sheet8=printing_sheet8,
            lamination_sheet9=lamination_sheet9,
            texture_sheet9=texture_sheet9,
            printing_sheet9=printing_sheet9,
            lamination_sheet10=lamination_sheet10,
            texture_sheet10=texture_sheet10,
            printing_sheet10=printing_sheet10,


            bill_sundries_1=bill_sundries_1,
            narration_1=narration_1,
            amount_1=amount_1,
            percent_1=percent_1,


            bill_sundries_2=bill_sundries_2,
            narration_2=narration_2,
            amount_2=amount_2,
            percent_2=percent_2,


            bill_sundries_3=bill_sundries_3,
            narration_3=narration_3,
            amount_3=amount_3,
            percent_3=percent_3,


            bill_sundries_4=bill_sundries_4,
            narration_4=narration_4,
            amount_4=amount_4,
            percent_4=percent_4,


            bill_sundries_5=bill_sundries_5,
            narration_5=narration_5,
            amount_5=amount_5,
            percent_5=percent_5,

            

        )
        data.save()
    obj=Bindingvender.objects.all()
    obj1=Company.objects.all()
    data = Cover_order.objects.all().order_by('-id')
   
    obj2 = books.objects.all()
    obj3=Cover_Printing.objects.all()
    obj4=venders.objects.all()
    data3=material_centre.objects.all()
    for i in obj3:
        print("gggggggggggg",i.printing_rate)
    
    # for i in data:
        
    #     obj3 = Cover_Printing.objects.all()
    #     for item in obj3:
    #         print("gggggggggggg", item.printing_rate)
    #         obj4 = int(item.printing_rate * i.quantity_order)  # Calculate product for each item
        
    #         print("1111111111111111",obj4)
    return render(request, 'brands/cover_order.html', {'data3':data3,'data':data,'obj':obj,'obj1':obj1,'obj2':obj2,'obj3':obj3,'obj4':obj4,})


def delete_cover_order(request, id):
    cus = Cover_order.objects.get(id=id)
    cus.delete()
    return redirect('/cover_order/')



def edit_cover_order(request, id):
    data = Cover_order.objects.get(id=id)
    if request.method == 'POST':
        data.vendor_order= request.POST.get('vendor_order')
        data.quantity_order= request.POST.get('quantity_order')
        data.subject_order= request.POST.get('subject_order')
        data.rate_order= request.POST.get('rate_order')
        data.palte_order= request.POST.get('palte_order')
        data.date_order= datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.self_company= request.POST.get('self_company')
        data.amount= request.POST.get('amount')

        
        data.save()
        return redirect('/cover_order/')
    
    context = {'data':data,}
    return render(request, 'brands/edit_cover_order.html', context)

def cover_order_bill_view(request,id):
    data = Cover_order.objects.get(id=id)
    

    print("data",data.rate_order2)

    max_orders = 10  # Maximum number of orders to consider

    order_data = []
    total_quantity = 0

    for i in range(1, max_orders + 1):
        quantity_field = 'amount{}'.format(i)
        print("11111111111",quantity_field)
        quantity = getattr(data, quantity_field, None)

        if quantity:
            order_data.append({
                'medium': request.POST.get('midium_{}'.format(i)),
                'subject': request.POST.get('bsubject_{}'.format(i)),
                'quantity': float(quantity),
                'ord_rate3': request.POST.get('ord_rate3{}'.format(i))
            })

            # Calculate the sum of quantities
            total_quantity += float(quantity) 
            print('total_quantity', total_quantity)

    # Update the total_quantity field of the Bindingorder object
    data.binding_order_total = total_quantity + float(data.amount)
    print("1111111111", data.binding_order_total)
    data.save()

    context = {
        'data': data,
        
    }
    return render(request,'brands/cover_order_bill_view.html',context)



def cover_transfer(request):
    # data1 =printing.objects.all().values()
    data1 = Cover_Printing.objects.filter(cover_transfer=True)
    print("data1",data1)
   
    template=loader.get_template('brands/cover_transfer.html')
    context={
      'data1':data1,
    }
    return HttpResponse(template.render(context,request))


def main_group(request):
    if request.method =='POST':
        id = request.POST.get('id')
        group_name = request.POST.get('group_name')
        alias = request.POST.get('alias')
        primary_input = request.POST.get('primary_input')
        type=request.POST.get('type')
        main_group_data = Main_group(group_name=group_name,alias=alias,primary_input=primary_input,type=type)
        main_group_data.save()  
        # if Group.objects.filter(group_type=group_type).exists():
        #     error_message = f"A Group with the type='{group_type}' already exists."
        #     sweetify.error(request, error_message, timer=8000, button='OK')
        # else:
        #     distributor_data = Group(group_type=group_type,main_discount=main_discount,extra_discount=extra_discount)
        #     distributor_data.save()           
        #     sweetify.success(request, "Group Created Successfully.", timer=6000 ,button='OK')
                    
    data = Main_group.objects.all()

    context= {
        'data':data,
    }
    return render(request,'group/main_group.html',context) 

@login_required(login_url="")
def delete_main_group(request, id):
    cus = Main_group.objects.get(id=id)
    cus.delete()
    return redirect('/main_group/')

def item_creation(request):
    if request.method =='POST':
        id = request.POST.get('id')
        brand_trade_name = request.POST.get('brand_trade_name')
        # brand_subtitle = request.POST.get('brand_subtitle')
        class1=request.POST.get('class1')
        size=request.POST.get('size')
        subject_name_author = request.POST.get('subject_name_author')
        code = request.POST.get('code')
        from_size = request.POST.get('from_size')
        item_binding_form_size = request.POST.get('item_binding_form_size')
        no_of_pages=request.POST.get('no_of_pages')
        unit = request.POST.get('unit')
        no_of_forms = request.POST.get('no_of_forms')
        binding_rate = request.POST.get('binding_rate')
        opening_stock_qty=request.POST.get('opening_stock_qty')
        opening_stock_value=request.POST.get('opening_stock_value') 
        bundle_pack = request.POST.get('bundle_pack')
        tax_information_rate = request.POST.get('tax_information_rate')
        tax_information_rate_central=request.POST.get('tax_information_rate_central')
        binding_no_of_form=request.POST.get("binding_no_of_form")
        item_material_center=request.POST.get("item_material_center")

        main_group_data = Item_creation(brand_trade_name=brand_trade_name,item_material_center=item_material_center,class1=class1,size=size,subject_name_author=subject_name_author,code=code,from_size=from_size,item_binding_form_size=item_binding_form_size,no_of_pages=no_of_pages,unit=unit,no_of_forms=no_of_forms,binding_rate=binding_rate,opening_stock_qty=opening_stock_qty,opening_stock_value=opening_stock_value,bundle_pack=bundle_pack,tax_information_rate=tax_information_rate,tax_information_rate_central=tax_information_rate_central,binding_no_of_form=binding_no_of_form)
        main_group_data.save()  
    data55= Item_creation.objects.all()
    data56=Product_group.objects.all()
    data57=classes.objects.all()
    data58=books.objects.all()
    data59=material_centre.objects.all()

    context= {
        'data55':data55,
        'data56':data56,
        'data57':data57,
        'data58':data58,
        'data59':data59,
    }

    return render(request,'group/item_creation.html' ,context) 
    
def delete_item_creation(request, id):
    cus = Item_creation.objects.get(id=id)
    cus.delete()
    return redirect('/item_creation/')


def edit_item_creation(request, id):
    data = Item_creation.objects.get(id=id)
    if request.method == 'POST':
        data.brand_trade_name= request.POST.get('brand_trade_name')
        # data.brand_subtitle= request.POST.get('brand_subtitle')
        data.class1= request.POST.get('class1')
        data.size= request.POST.get('size')
        data.subject_name_author= request.POST.get('subject_name_author')
        # data.date_order= datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.code= request.POST.get('code')
        data.from_size= request.POST.get('from_size')
        data.no_of_pages= request.POST.get('no_of_pages')
        data.unit= request.POST.get('unit')
        data.no_of_forms= request.POST.get('no_of_forms')
        data.binding_rate= request.POST.get('binding_rate')
        # data.date_order= datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.opening_stock_qty= request.POST.get('opening_stock_qty')
        data.opening_stock_value= request.POST.get('opening_stock_value')
        data.bundle_pack= request.POST.get('bundle_pack')
        data.tax_information_rate= request.POST.get('tax_information_rate')
        data.tax_information_rate_central= request.POST.get('tax_information_rate_central')
        data.item_material_center=request.POST.get("item_material_center")



        
        data.save()
        return redirect('/item_creation/')
    data59=material_centre.objects.all()
    context = {'data':data,'data59':data59,}
    return render(request, 'group/edit_item_creation.html', context)

def database_creation(request):
    if request.method =='POST':
        id = request.POST.get('id')
        district_d = request.POST.get('district_d')
        tehsil = request.POST.get('tehsil')
        town=request.POST.get('town')
        pincode=request.POST.get('pincode')
        
        # tahsil = request.POST.get('subject_name_author')
        database_data=database(district_d=district_d,tehsil=tehsil,town=town,pincode=pincode,)
        database_data.save()
    data52= database.objects.all()
    data=District.objects.all()
    data1=District1.objects.all()

    context= {
        'data52':data52,
        'data':data,
        'data1':data1,
    }

    return render(request, 'group/database.html',context)
def edit_database(request, id):
    data = database.objects.get(id=id)
    if request.method == 'POST':
        data.district_d= request.POST.get('district_d')
        data.tehsil= request.POST.get('tehsil')
        data.town= request.POST.get('town')
        data.pincode= request.POST.get('pincode')
        data.save()
        return redirect('/database/')
    data1=District1.objects.all()
    context = {'data':data,'data1':data1,}
    return render(request, 'group/edit_database.html', context)


def delete_database_creation(request,id):
    cus = database.objects.get(id=id)
    cus.delete()
    return redirect('/database/')


def tour_group(request):
    if request.method =='POST':
        
        
        zone=request.POST.get('zone')
        tour_group=request.POST.get('tour_group')
        
        database_data=Tour_group(zone=zone,tour_group=tour_group,)
        database_data.save()
    data52= Tour_group.objects.all()
    

    context= {
        'data52':data52,
        
    }

    return render(request, 'group/tour_group.html',context)

def delete_tour_group(request,id):
    cus = Tour_group.objects.get(id=id)
    cus.delete()
    return redirect('/tour_group/')



def district(request):
    if request.method == "POST":
        id = request.POST.get('id')
        district_master = request.POST.get('district_master').capitalize()
        # tehsil_master = request.POST.get('tehsil_master')
        
        
        if District.objects.filter(district_master__iexact=district_master).exists():
            error_message = f"A brand with the name '{district_master}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            # return HttpResponseBadRequest(error_message)
        else:
            # brand_code = generate_brand_code()
            
            database_data = District(district_master=district_master)
            database_data.save()
            sweetify.success(request, "District and Tehsil Created Successfully.", timer=6000, button='OK')
        
    data52= District.objects.all()
    
    context= {
        'data52':data52,
    }
    
    return render(request, 'group/district.html',context)
def delete_district(request,id):
    cus = District.objects.get(id=id)
    cus.delete()
    return redirect('/district/')

def Tehsil(request):
    all_district = District.objects.all()

    if request.method == 'POST':
        
        district_name = request.POST.get('district_master', '')
        tehsil_name = request.POST.get('tehsil_master', '').capitalize()

        
        existing_tehsil = District1.objects.filter(tehsil_master__iexact=tehsil_name).first()

        if existing_tehsil:
            error_message = f"A Tehsil with the name '{tehsil_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
        else:
            Tehsil_data = District1(district_master=district_name, tehsil_master=tehsil_name)
            Tehsil_data.save()
            sweetify.success(request, "District and Tehsil Created Successfully.", timer=6000, button='OK')

    data= District1.objects.all()

    return render(request,'group/add_tehsil.html',{'all_district':all_district,'data':data})


def delete_tehsil(request,id):
    cus = District1.objects.get(id=id)
    cus.delete()
    return redirect('/Tehsil/')

import json
def transactions_operation(request):
    if request.method =='POST':
        id = request.POST.get('id')

        company_type = request.POST.get('company_type')
        t_series = request.POST.get('t_series')
        party = request.POST.get('party')
        vehno=request.POST.get('vehno')
        delivered_at = request.POST.get('delivered_at')
        narration=request.POST.get('narration')

        Particular_1= request.POST.get('Particular_1')
        roll_quantity_1=request.POST.get('roll_quantity_1')
        roll_unit_1 = request.POST.get('roll_unit_1')
        roll_rate_1=request.POST.get('roll_rate_1')
        roll_amount_1=request.POST.get('roll_amount_1')
        ex_no_1=request.POST.get('ex_no_1')
        weight_1=request.POST.get('weight_1')
        p_weight_1=request.POST.get('p_weight_1')
        rim_1=request.POST.get('rim_1')
        sheet_1=request.POST.get('sheet_1')

        Particular= request.POST.get('Particular')
        roll_quantity=request.POST.get('roll_quantity')
        roll_unit = request.POST.get('roll_unit')
        roll_rate=request.POST.get('roll_rate')
        roll_amount=request.POST.get('roll_amount')
        ex_no=request.POST.get('ex_no')
        weight =request.POST.get('weight')
        p_weight=request.POST.get('p_weight')
        rim=request.POST.get('rim')
        sheet=request.POST.get('sheet')
        
        cloned_data_list = []
        for i in range(1, 11):  # Assuming a maximum of 10 clone fields
            Particular_2 = request.POST.get(f'Particular{i}')
            print("22222222",Particular_2)
            roll_quantity_2 = request.POST.get(f'roll_quantity{i}')
            print("dfghj",roll_quantity_2)
            
            roll_unit_2 = request.POST.get(f'roll_unit{i}')
            print("dfghj",roll_unit_2)
            
            roll_rate_2 = request.POST.get(f'roll_rate{i}')
            print("dfghj",roll_rate_2)

            roll_amount_2 = request.POST.get(f'roll_amount{i}')
            print("dfghj",roll_amount_2)

            ex_no_2 = request.POST.get(f'ex_no{i}')
            print("dfghj",ex_no_2)

            weight_2 = request.POST.get(f'weight{i}')
            print("dfghj",weight_2)

            p_weight_2 = request.POST.get(f'p_weight{i}')
            print("dfghj",p_weight_2)

            rim_2 = request.POST.get(f'rim{i}')
            print("dfghj",rim_2)

            sheet_2 = request.POST.get(f'sheet{i}')
            print("dfghj",sheet_2)

            print('6666666666666666666666666666',Particular_2,roll_quantity_2,roll_unit_2,roll_rate_2,roll_amount_2,ex_no_2,weight_2,p_weight_2,rim_2,sheet_2)
            if Particular_2 and roll_quantity_2 and roll_unit_2 and roll_rate_2 and roll_amount_2 and ex_no_2 and weight_2 and p_weight_2 and rim_2 and sheet_2:
                cloned_data_list.append({
                    'Particular': Particular_2,
                    'roll_quantity': roll_quantity_2,
                    'roll_unit': roll_unit_2,
                    'roll_rate':roll_rate_2,
                    'roll_amount':roll_amount_2,
                    'ex_no': ex_no_2,
                    'weight': weight_2,
                    'p_weight': p_weight_2,
                    'rim':rim_2,
                    'sheet':sheet_2,
                })

        database_data = transaction_operation.objects.create(
            company_type =company_type,
            t_series = t_series,
            party = party,
            vehno = vehno,
            delivered_at = delivered_at,
            narration = narration,

            Particular_1 = Particular_1,
            roll_quantity_1 = roll_quantity_1,
            roll_unit_1 =roll_unit_1,
            roll_rate_1 = roll_rate_1,
            roll_amount_1 =roll_amount_1,
            ex_no_1 = ex_no_1,
            weight_1 =weight_1,
            p_weight_1 = p_weight_1,
            rim_1 = rim_1,
            sheet_1 = sheet_1,

            Particular = Particular,
            roll_quantity = roll_quantity,
            roll_unit =roll_unit,
            roll_rate = roll_rate,
            roll_amount =roll_amount,
            ex_no = ex_no,
            weight =weight,
            p_weight = p_weight,
            rim = rim,
            sheet = sheet,

            dynamic_data=json.dumps(cloned_data_list),
        )
        database_data.save()
        return redirect('/transactions_operation/')
    data52= transaction_operation.objects.all()
    data53=material_centre.objects.all()
    data55=forms.objects.all()

    context= {
        'data52':data52,
        'data53':data53,
        'data55':data55,
    }
    return render(request,'group/transactions.html',context)

def transaction_view(request,id):
    data = transaction_operation.objects.filter(id=id).first()
    dynamic_data = json.loads(data.dynamic_data) if data and data.dynamic_data else []
    total_amount = 0
    total_weight = 0
    total_p_weight = 0
    total_rim = 0
    total_sheet=0
    for item in dynamic_data:
        total_amount += float(item.get('roll_amount', 0))
        total_weight += float(item.get('weight', 0))
        total_p_weight += float(item.get('p_weight', 0))
        total_sheet += float(item.get('sheet', 0))
        total_rim += float(item.get('rim', 0))
    # total_amount = sum(float(item.get(f'pr_voucher_amount{i}', 0)) for i in range(1, 11) for item in dynamic_data)
        print("sssssssssss",total_amount)

    total_amount += float(data.roll_amount_1)
    print("kkkkkkkkkkkkk",total_amount)

    total_weight += float(data.weight_1) 
    print("jjjjjjjjjjjjjjjj",total_weight)

    total_p_weight += float(data.p_weight_1) 
    print("jjjjjjjjjjjjjjjj",total_p_weight)

    total_rim += float(data.rim_1) 
    print("jjjjjjjjjjjjjjjj",total_rim)

    total_sheet += float(data.sheet_1) 
    print("jjjjjjjjjjjjjjjj",total_sheet)
    

    context = {
     'data':data,
     'dynamic_data':dynamic_data,
     'total_amount':total_amount,
     'total_weight':total_weight,
     'total_p_weight':total_p_weight,
     'total_rim':total_rim,
     'total_sheet':total_sheet,
    }
    return render(request,'group/transactions_view.html',context)


def delete_transaction(request,id):
    cus = transaction_operation.objects.get(id=id)
    cus.delete()
    return redirect('/transactions_operation/')

def edit_transaction(request, id):
    data = transaction_operation.objects.get(id=id)
    if request.method == 'POST':
        data.district= request.POST.get('district')
        data.tehsil= request.POST.get('tehsil')
        data.town= request.POST.get('town')
        data.pincode= request.POST.get('pincode')
        data.save()
        return redirect('/transactions_operation/')
    
    context = {'data':data,}
    return render(request, 'group/edit_transaction.html', context)





def binder_ledger(request):
    if request.method=="POST":

      name= request.POST.get("name")
      alias= request.POST.get("alias")
      prefix= request.POST.get("prefix")
      adress= request.POST.get("adress")
      contact_no= request.POST.get("contact_no")
      opening_balence= request.POST.get("opening_balence")
      whatsapp= request.POST.get("whatsapp")
      debit_binder= request.POST.get("debit_binder")
      email=request.POST.get("email")
      vender_group=request.POST.get("vender_group")
      vender_gst=request.POST.get("vender_gst")
      if Binder_ledger.objects.filter(name__iexact=name).exists():
            error_message = f"A name with the  '{name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
            
            
            

        data1 = Binder_ledger(
                
                name=name,
                alias=alias,
                prefix=prefix,
                adress=adress,
                contact_no=contact_no,
                whatsapp=whatsapp,
                opening_balence=opening_balence,
                debit_binder=debit_binder,
                email=email,
                vender_group=vender_group,
                vender_gst=vender_gst,


                
                

        )
        data1.save()
        sweetify.success(request, "Binder Created Successfully.", timer=6000, button='OK')
    
    data=Binder_ledger.objects.all().order_by('-id')
    data2=Main_group.objects.all()
    return render(request, 'brands/binder_ledger.html', {'data':data,'data2':data2})


@login_required(login_url="") 
def delete_binder(request, id):
    cus = Binder_ledger.objects.get(id=id)
    cus.delete()
    return redirect('/binder_ledger/')

@login_required(login_url="")
def edit_binder(request, id):
    data = Binder_ledger.objects.get(id=id)
    if request.method == 'POST':
            data.name = request.POST.get('name')
            data.debit_binder= request.POST.get('debit_binder')
            data.adress = request.POST.get('adress')
            data.alias = request.POST.get('alias')
            data.contact_no = request.POST.get('contact_no')
            data.whatsapp = request.POST.get('whatsapp')
            data.opening_balence = request.POST.get('opening_balence')
            data.email = request.POST.get('email')
            data.vender_group = request.POST.get('vender_group')
            data.vender_gst = request.POST.get('vender_gst')
            
            data.save()
           
            return redirect('/binder_ledger/')
    
    print("datanew", data.name)
    data2=Main_group.objects.all()
    context = {'data':data,'data2':data2,}
    return render(request, 'brands/edit_ledger.html', context)






def gather_ledger(request):
    if request.method=="POST":

      g_name= request.POST.get("g_name")
      g_alias= request.POST.get("g_alias")
      g_prefix= request.POST.get("g_prefix")
      g_adress= request.POST.get("g_adress")
      g_contact_no= request.POST.get("g_contact_no")
      g_opening_balence= request.POST.get("g_opening_balence")
      g_whatsapp= request.POST.get("g_whatsapp")
      gathering_rate= request.POST.get("gathering_rate")
      debit_gather= request.POST.get("debit_gather")
      g_email=request.POST.get("g_email")
      g_vender_group=request.POST.get("g_vender_group")
      g_vender_gst=request.POST.get("g_vender_gst")
      if Gather_ledger.objects.filter(g_name__iexact=g_name).exists():
            error_message = f"A name with the  '{g_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
            
            
            

        data1 = Gather_ledger(
                
                g_name=g_name,
                g_alias=g_alias,
                g_prefix=g_prefix,
                g_adress=g_adress,
                g_contact_no=g_contact_no,
                g_whatsapp=g_whatsapp,
                g_opening_balence=g_opening_balence,
                gathering_rate=gathering_rate,
                debit_gather=debit_gather,
                g_email=g_email,
                g_vender_group=g_vender_group,
                g_vender_gst=g_vender_gst,
                

        )
        data1.save()
        sweetify.success(request, "Gather Created Successfully.", timer=6000, button='OK')
    
    data=Gather_ledger.objects.all().order_by('-id')
    data2=Main_group.objects.all()
    
    return render(request, 'brands/gather_ledger.html', {'data':data,'data2':data2,})


@login_required(login_url="") 
def delete_gather_ledger(request, id):
    cus = Gather_ledger.objects.get(id=id)
    cus.delete()
    return redirect('/gather_ledger/')

@login_required(login_url="")
def edit_gather_ledger(request, id):
    data = Gather_ledger.objects.get(id=id)
    if request.method == 'POST':
            data.g_name = request.POST.get('g_name')
            data.g_adress = request.POST.get('g_adress')
            data.g_alias = request.POST.get('g_alias')
            data.g_contact_no = request.POST.get('g_contact_no')
            data.g_whatsapp = request.POST.get('g_whatsapp')
            data.g_opening_balence = request.POST.get('g_opening_balence')
            data.gathering_rate = request.POST.get('gathering_rate')
            data.debit_gather = request.POST.get('debit_gather')
            data.g_email = request.POST.get('g_email')
            data.g_vender_group = request.POST.get('g_vender_group')
            data.g_vender_gst = request.POST.get('g_vender_gst')

            data.save()
           
            return redirect('/gather_ledger/')
    
    print("datanew", data.g_name)
    data2=Main_group.objects.all()
    context = {'data':data,'data2':data2,}
    
    return render(request, 'brands/edit_gather_ledger.html', context)




def stitcher_ledger(request):
    if request.method=="POST":

      s_name= request.POST.get("s_name")
      s_alias= request.POST.get("s_alias")
      s_prefix= request.POST.get("s_prefix")
      s_adress= request.POST.get("s_adress")
      s_contact_no= request.POST.get("s_contact_no")
      s_opening_balence= request.POST.get("s_opening_balence")
      s_whatsapp= request.POST.get("s_whatsapp")
      stitching_rate= request.POST.get("stitching_rate")
      debit_stitcher = request.POST.get('debit_stitcher')
      s_email=request.POST.get("s_email")
      s_vender_group=request.POST.get("s_vender_group")
      s_vender_gst=request.POST.get("s_vender_gst")
      if Stitcher_ledger.objects.filter(s_name__iexact=s_name).exists():
            error_message = f"A name with the  '{s_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
            
            
            

        data1 = Stitcher_ledger(
                
                s_name=s_name,
                s_alias=s_alias,
                s_prefix=s_prefix,
                s_adress=s_adress,
                s_contact_no=s_contact_no,
                s_whatsapp=s_whatsapp,
                s_opening_balence=s_opening_balence,
                stitching_rate=stitching_rate,
                debit_stitcher=debit_stitcher,
                s_email=s_email,
                s_vender_group=s_vender_group,
                s_vender_gst=s_vender_gst,
                
                

        )
        data1.save()
        sweetify.success(request, "Stitching Created Successfully.", timer=6000, button='OK')
    
    data=Stitcher_ledger.objects.all().order_by('-id')
    data2=Main_group.objects.all()
    return render(request, 'brands/stitcher_ledger.html', {'data':data,'data2':data2})


@login_required(login_url="") 
def delete_stitcher_ledger(request, id):
    cus = Stitcher_ledger.objects.get(id=id)
    cus.delete()
    return redirect('/stitcher_ledger/')

@login_required(login_url="")
def edit_stitcher_ledger(request, id):
    data = Stitcher_ledger.objects.get(id=id)
    if request.method == 'POST':
            data.s_name = request.POST.get('s_name')
            data.s_adress = request.POST.get('s_adress')
            data.s_alias = request.POST.get('s_alias')
            data.s_contact_no = request.POST.get('s_contact_no')
            data.s_whatsapp = request.POST.get('s_whatsapp')
            data.s_opening_balence = request.POST.get('s_opening_balence')
            data.stitching_rate = request.POST.get('stitching_rate')
            data.debit_stitcher = request.POST.get('debit_stitcher')
            data.s_email = request.POST.get('s_email')
            data.s_vender_group = request.POST.get('s_vender_group')
            data.s_vender_gst = request.POST.get('s_vender_gst')

            data.save()
           
            return redirect('/stitcher_ledger/')
    
    print("datanew", data.s_name)
    data2=Main_group.objects.all()
    context = {'data':data,'data2':data2,}
    
    return render(request, 'brands/edit_stitcher_ledger.html', context)




def paster_ledger(request):
    if request.method=="POST":

      p_name= request.POST.get("p_name")
      p_alias= request.POST.get("p_alias")
      p_prefix= request.POST.get("p_prefix")
      p_adress= request.POST.get("p_adress")
      p_contact_no= request.POST.get("p_contact_no")
      p_opening_balence= request.POST.get("p_opening_balence")
      p_whatsapp= request.POST.get("p_whatsapp")
      pasting_rate= request.POST.get("pasting_rate")
      debit_paster = request.POST.get('debit_paster')
      p_email=request.POST.get("p_email")
      p_vender_group=request.POST.get("p_vender_group")
      p_vender_gst=request.POST.get("p_vender_gst")

      if Paster_ledger.objects.filter(p_name__iexact=p_name).exists():
            error_message = f"A name with the  '{p_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
            
            
            

        data1 = Paster_ledger(
                
                p_name=p_name,
                p_alias=p_alias,
                p_prefix=p_prefix,
                p_adress=p_adress,
                p_contact_no=p_contact_no,
                p_whatsapp=p_whatsapp,
                p_opening_balence=p_opening_balence,
                pasting_rate=pasting_rate,
                debit_paster=debit_paster,
                p_email=p_email,
                p_vender_group=p_vender_group,
                p_vender_gst=p_vender_gst,


                
                

        )
        data1.save()
        sweetify.success(request, "Stitching Created Successfully.", timer=6000, button='OK')
    
    data=Paster_ledger.objects.all().order_by('-id')
    data2=Main_group.objects.all()
    return render(request, 'brands/paster_ledger.html', {'data':data,'data2':data2,})


@login_required(login_url="") 
def delete_paster_ledger(request, id):
    cus = Paster_ledger.objects.get(id=id)
    cus.delete()
    return redirect('/paster_ledger/')

@login_required(login_url="")
def edit_paster_ledger(request, id):
    data = Paster_ledger.objects.get(id=id)
    if request.method == 'POST':
            data.p_name = request.POST.get('p_name')
            data.p_adress = request.POST.get('p_adress')
            data.p_alias = request.POST.get('p_alias')
            data.p_contact_no = request.POST.get('p_contact_no')
            data.p_whatsapp = request.POST.get('p_whatsapp')
            data.p_opening_balence = request.POST.get('p_opening_balence')
            data.pasting_rate = request.POST.get('pasting_rate')
            data.debit_paster = request.POST.get('debit_paster')
            data.p_email = request.POST.get('p_email')
            data.p_vender_group = request.POST.get('p_vender_group')
            data.p_vender_gst = request.POST.get('p_vender_gst')

            data.save()
           
            return redirect('/paster_ledger/')
    
    print("datanew", data.p_name)
    data2=Main_group.objects.all()
    context = {'data':data,'data2':data2,}
    return render(request, 'brands/edit_paster_ledger.html', context)




def machine_ledger(request):
    if request.method=="POST":

      m_name= request.POST.get("m_name")
      m_alias= request.POST.get("m_alias")
      m_prefix= request.POST.get("m_prefix")
      m_adress= request.POST.get("m_adress")
      m_contact_no= request.POST.get("m_contact_no")
      m_opening_balence= request.POST.get("m_opening_balence")
      m_whatsapp= request.POST.get("m_whatsapp")
      machine_rate= request.POST.get("machine_rate")
      debit_machine=request.POST.get('debit_machine')
      m_email=request.POST.get("m_email")
      m_vender_group=request.POST.get("m_vender_group")
      m_vender_gst=request.POST.get("m_vender_gst")
      if Machine_ledger.objects.filter(m_name__iexact=m_name).exists():
            error_message = f"A name with the  '{m_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
            
            
            

        data1 = Machine_ledger(
                
                m_name=m_name,
                m_alias=m_alias,
                m_prefix=m_prefix,
                m_adress=m_adress,
                m_contact_no=m_contact_no,
                m_whatsapp=m_whatsapp,
                m_opening_balence=m_opening_balence,
                debit_machine=debit_machine,
                m_email=m_email,
                m_vender_group=m_vender_group,
                m_vender_gst=m_vender_gst,
                machine_rate=machine_rate,
                
                

        )
        data1.save()
        sweetify.success(request, "Machine Binding Created Successfully.", timer=6000, button='OK')
    
    data=Machine_ledger.objects.all().order_by('-id')
    data2=Main_group.objects.all()
    
    return render(request, 'brands/machine_leadger.html', {'data':data,'data2':data2,})


@login_required(login_url="") 
def delete_machine_ledger(request, id):
    cus = Machine_ledger.objects.get(id=id)
    cus.delete()
    return redirect('/machine_ledger/')

@login_required(login_url="")
def edit_machine_ledger(request, id):
    data = Machine_ledger.objects.get(id=id)
    if request.method == 'POST':
            data.m_name = request.POST.get('m_name')
            data.m_adress = request.POST.get('m_adress')
            data.m_alias = request.POST.get('m_alias')
            data.m_contact_no = request.POST.get('m_contact_no')
            data.m_whatsapp = request.POST.get('m_whatsapp')
            data.m_opening_balence = request.POST.get('m_opening_balence')
            data.machine_rate = request.POST.get('machine_rate')
            data.debit_machine=request.POST.get('debit_machine')
            data.m_email = request.POST.get('m_email')
            data.m_vender_group = request.POST.get('m_vender_group')
            data.m_vender_gst = request.POST.get('m_vender_gst')

            data.save()
           
            return redirect('/machine_ledger/')
    
    print("datanew", data.m_name)
    data2=Main_group.objects.all()
    context = {'data':data,'data2':data2,}
    
    return render(request, 'brands/edit_machine_leadger.html', context)


def cutter_ledger(request):
    if request.method=="POST":

      c_name= request.POST.get("c_name")
      c_alias= request.POST.get("c_alias")
      c_prefix= request.POST.get("c_prefix")
      c_adress= request.POST.get("c_adress")
      c_contact_no= request.POST.get("c_contact_no")
      c_opening_balence= request.POST.get("c_opening_balence")
      c_whatsapp= request.POST.get("c_whatsapp")
      cutting_rate= request.POST.get("cutting_rate")
      cutting_rate_sheet= request.POST.get('cutting_rate_sheet')
      debit_cutter=request.POST.get('debit_cutter')
      c_email=request.POST.get("c_email")
      c_vender_group=request.POST.get("c_vender_group")
      c_vender_gst=request.POST.get("c_vender_gst")
      if Cutter_ledger.objects.filter(c_name__iexact=c_name).exists():
            error_message = f"A name with the  '{c_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
            
            
            

        data1 = Cutter_ledger(
                
                c_name=c_name,
                c_alias=c_alias,
                c_prefix=c_prefix,
                c_adress=c_adress,
                c_contact_no=c_contact_no,
                c_whatsapp=c_whatsapp,
                c_opening_balence=c_opening_balence,
                cutting_rate=cutting_rate,
                debit_cutter=debit_cutter,
                cutting_rate_sheet=cutting_rate_sheet,
                c_email=c_email,
                c_vender_group=c_vender_group,
                c_vender_gst=c_vender_gst,
                
                

        )
        data1.save()
        sweetify.success(request, "Stitching Created Successfully.", timer=6000, button='OK')
    
    data=Cutter_ledger.objects.all().order_by('-id')
    data2=Main_group.objects.all()
    return render(request, 'brands/cutter_ledger.html', {'data':data,'data2':data2,})


@login_required(login_url="") 
def delete_cutter_ledger(request, id):
    cus = Cutter_ledger.objects.get(id=id)
    cus.delete()
    return redirect('/cutter_ledger/')

@login_required(login_url="")
def edit_cutter_ledger(request, id):
    data = Cutter_ledger .objects.get(id=id)
    if request.method == 'POST':
            data.c_name = request.POST.get('c_name')
            data.c_adress = request.POST.get('c_adress')
            data.c_alias = request.POST.get('c_alias')
            data.c_contact_no = request.POST.get('c_contact_no')
            data.c_whatsapp = request.POST.get('c_whatsapp')
            data.c_opening_balence = request.POST.get('c_opening_balence')
            data.cutting_rate = request.POST.get('cutting_rate')
            data.cutting_rate_sheet= request.POST.get('cutting_rate_sheet')
            data.debit_cutter= request.POST.get('debit_cutter')
            data.c_email = request.POST.get('c_email')
            data.c_vender_group = request.POST.get('c_vender_group')
            data.c_vender_gst = request.POST.get('c_vender_gst')

            data.save()
           
            return redirect('/cutter_ledger/')
    
    print("datanew", data.c_name)
    data2=Main_group.objects.all()
    context = {'data':data,'data2':data2}
    return render(request, 'brands/edit_cutter_ledger.html', context)





def packer_ledger(request):
    if request.method=="POST":

      pack_name= request.POST.get("pack_name")
      pack_alias= request.POST.get("pack_alias")
      pack_prefix= request.POST.get("pack_prefix")
      pack_adress= request.POST.get("pack_adress")
      pack_contact_no= request.POST.get("pack_contact_no")
      pack_opening_balence= request.POST.get("pack_opening_balence")
      pack_whatsapp= request.POST.get("pack_whatsapp")
      packing_rate= request.POST.get("packing_rate")
      bitti_rate= request.POST.get("bitti_rate") 
      debit_packer=request.POST.get('debit_packer')
      pack_email=request.POST.get("pack_email")
      pack_vender_group=request.POST.get("pack_vender_group")
      pack_vender_gst=request.POST.get("pack_vender_gst")
      if Packer_ledger.objects.filter(pack_name__iexact=pack_name).exists():
            error_message = f"A name with the  '{pack_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
            
            
            

        data1 = Packer_ledger(
                
                pack_name= pack_name,
                pack_alias=pack_alias,
                pack_prefix=pack_prefix,
                pack_adress=pack_adress,
                pack_contact_no=pack_contact_no,
                pack_whatsapp=pack_whatsapp,
                pack_opening_balence=pack_opening_balence,
                packing_rate=packing_rate,
                bitti_rate=bitti_rate,
                debit_packer=debit_packer,
                pack_email=pack_email,
                pack_vender_group=pack_vender_group,
                pack_vender_gst=pack_vender_gst,
                
                

        )
        data1.save()
        sweetify.success(request, "Packing Created Successfully.", timer=6000, button='OK')
    
    data=Packer_ledger.objects.all().order_by('-id')
    data2=Main_group.objects.all()
    return render(request, 'brands/packer_ledger.html', {'data':data,'data2':data2,})


@login_required(login_url="") 
def delete_packer_ledger(request, id):
    cus = Packer_ledger.objects.get(id=id)
    cus.delete()
    return redirect('/packer_ledger/')

@login_required(login_url="")
def edit_packer_ledger(request, id):
    data = Packer_ledger .objects.get(id=id)
    if request.method == 'POST':
            data.pack_name = request.POST.get('pack_name')
            data.pack_adress = request.POST.get('pack_adress')
            data.pack_alias = request.POST.get('pack_alias')
            data.pack_contact_no = request.POST.get('pack_contact_no')
            data.pack_whatsapp = request.POST.get('pack_whatsapp')
            data.pack_opening_balence = request.POST.get('pack_opening_balence')
            data.packing_rate = request.POST.get('packing_rate')
            data.bitti_rate = request.POST.get('bitti_rate')
            data.debit_packer= request.POST.get('debit_packer')
            data.pack_email = request.POST.get('pack_email')
            data.pack_vender_group = request.POST.get('pack_vender_group')
            data.pack_vender_gst = request.POST.get('pack_vender_gst')

            data.save()
           
            return redirect('/packer_ledger/')
    
    print("datanew", data.pack_name)
    data2=Main_group.objects.all()
    context = {'data':data,'data2':data2,}
    return render(request, 'brands/edit_packer_ledger.html', context)




def general_ledger(request):
    if request.method=="POST":

      g_name= request.POST.get("g_name")
      g_alias= request.POST.get("g_alias")
      g_prefix= request.POST.get("g_prefix")
      g_adress= request.POST.get("g_adress")
      g_contact_no= request.POST.get("g_contact_no")
      g_opening_balence= request.POST.get("g_opening_balence")
      g_whatsapp= request.POST.get("g_whatsapp")
      previous_balence= request.POST.get("previous_balence")
      act_material_center= request.POST.get("act_material_center") 
      fax= request.POST.get("fax") 
      eamil= request.POST.get("eamil") 
      tel_no= request.POST.get("tel_no") 
      print_name= request.POST.get("print_name") 
      group= request.POST.get("group") 
      debit_general= request.POST.get('debit_general')
      g_vender_gst=request.POST.get('g_vender_gst')

      if General_ledger.objects.filter(g_name__iexact=g_name).exists():
            error_message = f"A name with the  '{g_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
            
            
            

        data1 = General_ledger(
                
                g_name= g_name,
                g_alias=g_alias,
                g_prefix=g_prefix,
                g_adress=g_adress,
                g_contact_no=g_contact_no,
                g_whatsapp=g_whatsapp,
                g_opening_balence=g_opening_balence,
                previous_balence=previous_balence,
                act_material_center=act_material_center,
                fax=fax,
                eamil=eamil,
                tel_no=tel_no,
                print_name=print_name,
                group=group,
                debit_general=debit_general,
                g_vender_gst=g_vender_gst,
                

        )
        data1.save()
        sweetify.success(request, "General Account Created Successfully.", timer=6000, button='OK')
    
    data=General_ledger.objects.all().order_by('-id')
    data1=Main_group.objects.all()
    data2=Main_group.objects.all()

    context= {
        'data':data,
        'data1':data1,
        'data2':data2,
    }

    return render(request, 'brands/general_ledger.html', context)


@login_required(login_url="") 
def delete_general_ledger(request, id):
    cus = General_ledger.objects.get(id=id)
    cus.delete()
    return redirect('/general_ledger/')

@login_required(login_url="")
def edit_general_ledger(request, id):
    data = General_ledger .objects.get(id=id)
    if request.method == 'POST':
            data.g_name = request.POST.get('g_name')
            data.g_adress = request.POST.get('g_adress')
            data.g_alias = request.POST.get('g_alias')
            data.g_contact_no = request.POST.get('g_contact_no')
            data.g_whatsapp = request.POST.get('g_whatsapp')
            data.g_opening_balence = request.POST.get('g_opening_balence')
            data.previous_balence = request.POST.get('previous_balence')
            data.act_material_center = request.POST.get('act_material_center')
            data.fax = request.POST.get('fax')
            data.eamil = request.POST.get('eamil')
            data.tel_no = request.POST.get('tel_no')
            data.print_name = request.POST.get('print_name')
            data.group = request.POST.get('group')
            data.debit_general= request.POST.get('debit_general')
            data.g_vender_gst=request.POST.get('g_vender_gst')

            data.save()
           
            return redirect('/general_ledger/')
    
    print("datanew", data.g_name)
    data2=Main_group.objects.all()
    context = {'data':data,'data2':data2,}
    return render(request, 'brands/edit_general.html', context)



def customer_ledger(request):
    if request.method=="POST":

      cust_name= request.POST.get("cust_name")
      cust_alias= request.POST.get("cust_alias")
      cust_prefix= request.POST.get("cust_prefix")
      cust_adress= request.POST.get("cust_adress")
      cust_contact_no= request.POST.get("cust_contact_no")
      cust_opening_balence= request.POST.get("cust_opening_balence")
      cust_whatsapp= request.POST.get("cust_whatsapp")
      code= request.POST.get("code")
      district= request.POST.get("district") 
      zone= request.POST.get("zone") 
      cust_eamil= request.POST.get("cust_eamil") 
      tour_group= request.POST.get("tour_group") 
      area= request.POST.get("area") 
      group= request.POST.get("group") 
      tahsil= request.POST.get("tahsil") 
      price_category= request.POST.get("price_category") 
      pin= request.POST.get("pin")
      debit_customer= request.POST.get('debit_customer')
      c_vender_gst=request.POST.get('c_vender_gst')

      if Customer_ledger.objects.filter(cust_name__iexact=cust_name).exists():
            error_message = f"A name with the  '{cust_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
          data1 = Customer_ledger(
                    
                    cust_name= cust_name,
                    cust_alias=cust_alias,
                    cust_prefix=cust_prefix,
                    cust_adress=cust_adress,
                    cust_contact_no=cust_contact_no,
                    cust_whatsapp=cust_whatsapp,
                    cust_opening_balence=cust_opening_balence,
                    code=code,
                    district=district,
                    zone=zone,
                    cust_eamil=cust_eamil,
                    tour_group=tour_group,
                    area=area,
                    group=group,
                    price_category=price_category,
                    tahsil=tahsil,
                    pin=pin,
                    debit_customer=debit_customer,
                    c_vender_gst=c_vender_gst,

          )
          data1.save()
          sweetify.success(request, "General Account Created Successfully.", timer=6000, button='OK')
        
    data=Customer_ledger.objects.all().order_by('-id')
    obj2=database.objects.all()
    obj=Group.objects.all()
    obj1=Category.objects.all()
    data09=Main_group.objects.all()
    
    return render(request, 'brands/customer_ledger.html', {'data':data,'obj2':obj2,'data09':data09,'obj1':obj1})


@login_required(login_url="") 
def delete_customer_ledger(request, id):
    cus = Customer_ledger.objects.get(id=id)
    cus.delete()
    return redirect('/customer_ledger/')

@login_required(login_url="")
def edit_customer_ledger(request, id):
    data = Customer_ledger .objects.get(id=id)
    if request.method == 'POST':
            data.cust_name = request.POST.get('cust_name')
            data.cust_adress = request.POST.get('cust_adress')
            data.cust_alias = request.POST.get('cust_alias')
            data.cust_contact_no = request.POST.get('cust_contact_no')
            data.cust_whatsapp = request.POST.get('cust_whatsapp')
            data.cust_opening_balence = request.POST.get('cust_opening_balence')
            data.code = request.POST.get('code')
            data.district = request.POST.get('district')
            data.zone = request.POST.get('zone')
            data.cust_eamil = request.POST.get('cust_eamil')
            data.tour_group = request.POST.get('tour_group')
            data.area = request.POST.get('area')
            data.group = request.POST.get('group')
            data.price_category = request.POST.get('price_category')
            data.tahsil = request.POST.get('tahsil')
            data.pin =request.POST.get('pin')
            data.debit_customer= request.POST.get('debit_customer')
            data.c_vender_gst=request.POST.get('c_vender_gst')
            data.save()
           
            return redirect('/customer_ledger/')
    
    print("datanew", data.cust_name)
    data09=Main_group.objects.all()
    context = {'data':data,'data09':data09,}
    return render(request, 'brands/edit_customer_ledger.html', context)


# from brand_app.models import cover_scrap_voucher

def cover_scrap(request):
    cover_printing_data = Cover_Printing.objects.all()
    if request.method=="POST":
      
      for record in cover_printing_data:
            # Iterate through cover_scrap_voucher particulars and compare with Cover_Printing subjects
            for i in range(1, 11):  # Assuming 10 entries for cover_scrap_particular
                cover_scrap_particular = request.POST.get(f"cover_scrap_particular{i}")
                print("cover_scrap_particular",cover_scrap_particular)
                cover_printing_subject = getattr(record, f"subject{i}")
                print()

                # If there is a match between cover_scrap_particular and Cover_Printing subject
                if cover_scrap_particular == cover_printing_subject:
                    # Fetch quantity from cover_scrap_voucher
                    quantity = request.POST.get(f"cover_scrap_qty{i}")

                    # Update quantity_cover field in the Cover_Printing model
                    setattr(record, 'quantity_cover', quantity)
                    quantity_cover_after_match = record.quantity_cover
                    print("lllllllllllllllllllllllllllllllllllllllll",quantity_cover_after_match)

                    print(f"Match found for {cover_scrap_particular}. Quantity: {quantity} - Updated quantity_cover: {quantity_cover_after_match}")


      cover_date= request.POST.get("cover_date")
      cover_voucher_no= request.POST.get("cover_voucher_no")

      cover_scrap_particular1=request.POST.get("cover_scrap_particular1")
      cover_scrap_qty1=request.POST.get("cover_scrap_qty1")
      cover_scrap_unit1=request.POST.get("cover_scrap_unit1")
      cover_scrap_rate1=request.POST.get("cover_scrap_rate1")

      cover_scrap_particular2=request.POST.get("cover_scrap_particular2")
      cover_scrap_qty2=request.POST.get("cover_scrap_qty2")
      cover_scrap_unit2=request.POST.get("cover_scrap_unit2")
      cover_scrap_rate2=request.POST.get("cover_scrap_rate2")

      cover_scrap_particular3=request.POST.get("cover_scrap_particular3")
      cover_scrap_qty3=request.POST.get("cover_scrap_qty3")
      cover_scrap_unit3=request.POST.get("cover_scrap_unit3")
      cover_scrap_rate3=request.POST.get("cover_scrap_rate3")

      cover_scrap_particular4=request.POST.get("cover_scrap_particular4")
      cover_scrap_qty4=request.POST.get("cover_scrap_qty4")
      cover_scrap_unit4=request.POST.get("cover_scrap_unit4")
      cover_scrap_rate4=request.POST.get("cover_scrap_rate4")

      cover_scrap_particular5=request.POST.get("cover_scrap_particular5")
      cover_scrap_qty5=request.POST.get("cover_scrap_qty5")
      cover_scrap_unit5=request.POST.get("cover_scrap_unit5")
      cover_scrap_rate5=request.POST.get("cover_scrap_rate5")

      cover_scrap_particular6=request.POST.get("cover_scrap_particular6")
      cover_scrap_qty6=request.POST.get("cover_scrap_qty6")
      cover_scrap_unit6=request.POST.get("cover_scrap_unit6")
      cover_scrap_rate6=request.POST.get("cover_scrap_rate6")

      cover_scrap_particular7=request.POST.get("cover_scrap_particular7")
      cover_scrap_qty7=request.POST.get("cover_scrap_qty7")
      cover_scrap_unit7=request.POST.get("cover_scrap_unit7")
      cover_scrap_rate7=request.POST.get("cover_scrap_rate7")

      cover_scrap_particular8=request.POST.get("cover_scrap_particular8")
      cover_scrap_qty8=request.POST.get("cover_scrap_qty8")
      cover_scrap_unit8=request.POST.get("cover_scrap_unit8")
      cover_scrap_rate8=request.POST.get("cover_scrap_rate8")

      cover_scrap_particular9=request.POST.get("cover_scrap_particular9")
      cover_scrap_qty9=request.POST.get("cover_scrap_qty9")
      cover_scrap_unit9=request.POST.get("cover_scrap_unit9")
      cover_scrap_rate9=request.POST.get("cover_scrap_rate9")

      cover_scrap_particular10=request.POST.get("cover_scrap_particular10")
      cover_scrap_qty10=request.POST.get("cover_scrap_qty10")
      cover_scrap_unit10=request.POST.get("cover_scrap_unit10")
      cover_scrap_rate10=request.POST.get("cover_scrap_rate10")
      cover_scrap_total_quantity=request.POST.get("cover_scrap_total_quantity")




      cover_subject= request.POST.get("cover_subject")
      cover_qty= request.POST.get("cover_qty")
      print("3333333",cover_subject)
      cover_unit= request.POST.get("cover_unit")
      cover_rate= request.POST.get("cover_rate")
      cover_bill_material= request.POST.get("cover_bill_material")
      cover_bill_material_name= request.POST.get("cover_bill_material_name")
      cover_alias= request.POST.get("cover_alias")
      cover_item_produce= request.POST.get("cover_item_produce")
      cover_quantity= request.POST.get("cover_quantity")
      cover_unit1= request.POST.get("cover_unit1")
      cover_expenses= request.POST.get("cover_expenses")
      item_generated= request.POST.get("item_generated")
      item_consumed= request.POST.get("item_consumed")
      

      main_group_data =cover_scrap_voucher(cover_date=cover_date,cover_voucher_no=cover_voucher_no,
                                           cover_scrap_particular1=cover_scrap_particular1,cover_scrap_qty1=cover_scrap_qty1,
                                           cover_scrap_unit1=cover_scrap_unit1,cover_scrap_rate1=cover_scrap_rate1,
                                           cover_scrap_particular2=cover_scrap_particular2,cover_scrap_qty2=cover_scrap_qty2,
                                           cover_scrap_unit2=cover_scrap_unit2,cover_scrap_rate2=cover_scrap_rate2,
                                           cover_scrap_particular3=cover_scrap_particular3,cover_scrap_qty3=cover_scrap_qty3,
                                           cover_scrap_unit3=cover_scrap_unit3,cover_scrap_rate3=cover_scrap_rate3,
                                           cover_scrap_particular4=cover_scrap_particular4,cover_scrap_qty4=cover_scrap_qty4,
                                           cover_scrap_unit4=cover_scrap_unit4,cover_scrap_rate4=cover_scrap_rate4,
                                           cover_scrap_particular5=cover_scrap_particular5,cover_scrap_qty5=cover_scrap_qty5,
                                           cover_scrap_unit5=cover_scrap_unit5,cover_scrap_rate5=cover_scrap_rate5,
                                           cover_scrap_particular6=cover_scrap_particular6,cover_scrap_qty6=cover_scrap_qty6,
                                           cover_scrap_unit6=cover_scrap_unit6,cover_scrap_rate6=cover_scrap_rate6,
                                           cover_scrap_particular7=cover_scrap_particular7,cover_scrap_qty7=cover_scrap_qty7,
                                           cover_scrap_unit7=cover_scrap_unit7,cover_scrap_rate7=cover_scrap_rate7,
                                           cover_scrap_particular8=cover_scrap_particular8,cover_scrap_qty8=cover_scrap_qty8,
                                           cover_scrap_unit8=cover_scrap_unit8,cover_scrap_rate8=cover_scrap_rate8,
                                           cover_scrap_particular9=cover_scrap_particular9,cover_scrap_qty9=cover_scrap_qty9,
                                           cover_scrap_unit9=cover_scrap_unit9,cover_scrap_rate9=cover_scrap_rate9,
                                           cover_scrap_particular10=cover_scrap_particular10,cover_scrap_qty10=cover_scrap_qty10,
                                           cover_scrap_unit10=cover_scrap_unit10,cover_scrap_rate10=cover_scrap_rate10,
                                           cover_scrap_total_quantity=cover_scrap_total_quantity,
                                           cover_subject=cover_subject,cover_qty=cover_qty,cover_unit=cover_unit,cover_rate=cover_rate,
                                            cover_bill_material=cover_bill_material,cover_bill_material_name=cover_bill_material_name,
                                            cover_alias=cover_alias,cover_item_produce=cover_item_produce,cover_quantity=cover_quantity,
                                            cover_unit1=cover_unit1,cover_expenses=cover_expenses,item_generated=item_generated,item_consumed=item_consumed)
      main_group_data.save()  

    data25 = cover_scrap_voucher.objects.all()
    data27=BOM.objects.all()
    data26=books.objects.all()

    context= {
        'data25':data25,
        'data26':data26,
        'data27':data27,
    }

    return render(request,'group/cover_scrap_voucher.html',context)

def delete_cover_scrap_voucher(request,id):
    cus = cover_scrap_voucher.objects.get(id=id)
    cus.delete()
    return redirect('/cover_scrap/')

def edit_cover_scrap_voucher(request,id):
    data = cover_scrap_voucher.objects.get(id=id)
    if request.method == 'POST':
            data.cover_voucher_no = request.POST.get('cover_voucher_no')

            data.cover_scrap_particular1=request.POST.get("cover_scrap_particular1")
            data.cover_scrap_qty1=request.POST.get("cover_scrap_qty1")
            data.cover_scrap_unit1=request.POST.get("cover_scrap_unit1")
            data.cover_scrap_rate1=request.POST.get("cover_scrap_rate1")

            data.cover_scrap_particular2=request.POST.get("cover_scrap_particular2")
            data.cover_scrap_qty2=request.POST.get("cover_scrap_qty2")
            data.cover_scrap_unit2=request.POST.get("cover_scrap_unit2")
            data.cover_scrap_rate2=request.POST.get("cover_scrap_rate2")

            data.cover_scrap_particular3=request.POST.get("cover_scrap_particular3")
            data.cover_scrap_qty3=request.POST.get("cover_scrap_qty3")
            data.cover_scrap_unit3=request.POST.get("cover_scrap_unit3")
            data.cover_scrap_rate3=request.POST.get("cover_scrap_rate3")

            data.cover_scrap_particular4=request.POST.get("cover_scrap_particular4")
            data.cover_scrap_qty4=request.POST.get("cover_scrap_qty4")
            data.cover_scrap_unit4=request.POST.get("cover_scrap_unit4")
            data.cover_scrap_rate4=request.POST.get("cover_scrap_rate4")

            data.cover_scrap_particular5=request.POST.get("cover_scrap_particular5")
            data.cover_scrap_qty5=request.POST.get("cover_scrap_qty5")
            data.cover_scrap_unit5=request.POST.get("cover_scrap_unit5")
            data.cover_scrap_rate5=request.POST.get("cover_scrap_rate5")

            data.cover_scrap_particular6=request.POST.get("cover_scrap_particular6")
            data.cover_scrap_qty6=request.POST.get("cover_scrap_qty6")
            data.cover_scrap_unit6=request.POST.get("cover_scrap_unit6")
            data.cover_scrap_rate6=request.POST.get("cover_scrap_rate6")

            data.cover_scrap_particular7=request.POST.get("cover_scrap_particular7")
            data.cover_scrap_qty7=request.POST.get("cover_scrap_qty7")
            data.cover_scrap_unit7=request.POST.get("cover_scrap_unit7")
            data.cover_scrap_rate7=request.POST.get("cover_scrap_rate7")

            data.cover_scrap_particular8=request.POST.get("cover_scrap_particular8")
            data.cover_scrap_qty8=request.POST.get("cover_scrap_qty8")
            data.cover_scrap_unit8=request.POST.get("cover_scrap_unit8")
            data.cover_scrap_rate8=request.POST.get("cover_scrap_rate8")

            data.cover_scrap_particular9=request.POST.get("cover_scrap_particular9")
            data.cover_scrap_qty9=request.POST.get("cover_scrap_qty9")
            data.cover_scrap_unit9=request.POST.get("cover_scrap_unit9")
            data.cover_scrap_rate9=request.POST.get("cover_scrap_rate9")

            data.cover_scrap_particular10=request.POST.get("cover_scrap_particular10")
            data.cover_scrap_qty10=request.POST.get("cover_scrap_qty10")
            data.cover_scrap_unit10=request.POST.get("cover_scrap_unit10")
            data.cover_scrap_rate10=request.POST.get("cover_scrap_rate10")
            data.cover_scrap_total_quantity=request.POST.get("cover_scrap_total_quantity")


            data.cover_subject = request.POST.get('cover_subject')
            data.cover_qty = request.POST.get('cover_qty')
            data.cover_unit = request.POST.get('cover_unit')
            data.cover_rate= request.POST.get("cover_rate")
            data.cover_bill_material= request.POST.get("cover_bill_material")
            data.cover_bill_material_name= request.POST.get("cover_bill_material_name")
            data.cover_alias= request.POST.get("cover_alias")
            data.cover_item_produce = request.POST.get('cover_item_produce')
            data.cover_quantity= request.POST.get("cover_quantity")
            data.cover_unit1= request.POST.get("cover_unit1")
            data.cover_expenses= request.POST.get("cover_expenses")
            data.item_generated= request.POST.get("item_generated")
            data.item_consumed= request.POST.get("item_consumed")
            data.save()

            return redirect('/cover_scrap/')
    data26=books.objects.all()
    context = {'data':data,'data26':data26}
    return render(request, 'group/edit_cover_scrap.html', context)





def other_voucher(request):

    if request.method=="POST":
      contes= request.POST.get("contes")
      account_dr= request.POST.get("account_dr")
      account_cr= request.POST.get("account_cr")
      Debit_credit= request.POST.get("Debit_credit")
      account= request.POST.get("account")
      debit= request.POST.get("debit")
      credit= request.POST.get("credit")
      inst_type= request.POST.get("inst_type")
      inst_no= request.POST.get("inst_no")

      other_vouchers_data =other_vouchers(contes=contes,account_dr=account_dr,account_cr=account_cr,Debit_credit=Debit_credit,account=account,debit=debit,
                                            credit=credit,inst_type=inst_type,
                                            inst_no=inst_no)
      other_vouchers_data.save()  

    data25=other_vouchers.objects.all()

    context={
          'data25':data25,
      }


    return render(request,'group/contes.html',context)


def delete_other_voucher(request,id):
    cus = other_vouchers.objects.get(id=id)
    cus.delete()
    return redirect('/other_voucher/')

def edit_other_voucher(request,id):
    data = other_vouchers.objects.get(id=id)
    if request.method == 'POST':
            data.contes = request.POST.get('contes')
            data.account_dr = request.POST.get('account_dr')
            data.account_cr = request.POST.get('account_cr')
            data.Debit_credit = request.POST.get('Debit_credit')
            data.account= request.POST.get("account")
            data.debit= request.POST.get("debit")
            data.credit= request.POST.get("credit")
            data.inst_type= request.POST.get("inst_type")
            data.inst_no = request.POST.get('inst_no')
           
            data.save()

            return redirect('/other_voucher/')
    context = {'data':data,}
    return render(request, 'group/edit_other_voucher.html', context)


def by_product(request):
    if request.method=="POST":
      pname= request.POST.get("pname")
      palias= request.POST.get("palias")
      p_print_name= request.POST.get("p_print_name")
      opening_stock_p= request.POST.get("opening_stock_p")
      punit= request.POST.get("punit")
      p_opening_stock_value= request.POST.get("p_opening_stock_value")
      p_gsm= request.POST.get("p_gsm")
      p_p_width= request.POST.get("p_p_width")
      punit33= request.POST.get("punit33")
      punit22= request.POST.get("punit22")
      punit2= request.POST.get("punit2")
      punit3= request.POST.get("punit3")
      punit1_conversion= request.POST.get("punit1_conversion")
      pcunit1= request.POST.get("pcunit1")
      punit2_conversion= request.POST.get("punit2_conversion")
      pcunit2= request.POST.get("pcunit2")
      sale_price= request.POST.get("sale_price")
      purchase_price= request.POST.get("purchase_price")
      mrp= request.POST.get("mrp")
      min_sale_price= request.POST.get("min_sale_price")
      sale_value_price= request.POST.get("sale_value_price")
      product_tax_local= request.POST.get("product_tax_local")
      product_tax_central= request.POST.get("product_tax_central")
      by_material_center= request.POST.get("by_material_center")
      product_gst= request.POST.get("product_gst")
      product_gst=float(product_gst) / 100
      product_item_description= request.POST.get("product_item_description")

      product_by_data=product_by(pname=pname,       
                                    palias=palias,
                                    p_print_name=p_print_name,
                                    opening_stock_p=opening_stock_p,
                                    punit=punit,
                                    p_opening_stock_value=p_opening_stock_value,
                                    p_gsm=p_gsm,
                                    punit33=punit33,
                                    
                                    punit22=punit22,
                                    punit2=punit2,
                                    punit3=punit3,
                                    punit1_conversion=punit1_conversion,
                                    pcunit1=pcunit1,
                                    punit2_conversion=punit2_conversion,
                                    pcunit2=pcunit2,
                                    sale_price=sale_price,
                                    purchase_price=purchase_price,
                                    mrp=mrp,
                                    min_sale_price=min_sale_price,
                                    sale_value_price=sale_value_price,
                                    product_tax_local=product_tax_local,
                                    product_tax_central=product_tax_central,
                                    product_gst=product_gst,
                                    product_item_description=product_item_description,
                                    by_material_center=by_material_center,)
      
      product_by_data.save()  

    data25=product_by.objects.all()
    data1=material_centre.objects.all()

    context={
          'data25':data25,
          'data1':data1,
      }

    return render(request, 'group/by_product.html',context)



def delete_by_product(request,id):
    cus = product_by.objects.get(id=id)
    cus.delete()
    return redirect('/by_product/')

def edit_by_product(request,id):
    data = product_by.objects.get(id=id)
    if request.method == 'POST':
            data.pname= request.POST.get("pname")
            data.palias= request.POST.get("palias")
            data.p_print_name= request.POST.get("p_print_name")
            data.opening_stock_p= request.POST.get("opening_stock_p")
            # data.punit= request.POST.get("punit")
            data.p_opening_stock_value= request.POST.get("p_opening_stock_value")
            
            data.punit1= request.POST.get("punit1")
            data.punit2= request.POST.get("punit2")
            data.punit3= request.POST.get("punit3")
            data.punit1_conversion= request.POST.get("punit1_conversion")
            # data.pcunit1= request.POST.get("pcunit1")
            data.punit2_conversion= request.POST.get("punit2_conversion")
            # data.pcunit2= request.POST.get("pcunit2")
            # data.punit33= request.POST.get("punit33")
            # data.punit22= request.POST.get("punit22")
            data.sale_price= request.POST.get("sale_price")
            data.purchase_price= request.POST.get("purchase_price")
            data.mrp= request.POST.get("mrp")
            data.min_sale_price= request.POST.get("min_sale_price")
            data.sale_value_price= request.POST.get("sale_value_price")
            data.product_tax_local= request.POST.get("product_tax_local")
            data.product_tax_central= request.POST.get("product_tax_central")
            data.product_gst= request.POST.get("product_gst")
            data.product_item_description= request.POST.get("product_item_description")
            data.by_material_center= request.POST.get("by_material_center")

            data.save()

            return redirect('/by_product/')
    data1=material_centre.objects.all()
    context = {'data':data,'data1':data1,}
    return render(request, 'group/edit_by_product.html', context)






def product_group(request):
    if request.method=="POST":

      brand_name= request.POST.get("brand_name")
      sub_brand_name= request.POST.get("sub_brand_name")
      brand_subtitle= request.POST.get("brand_subtitle")
      categories= request.POST.get("categories")
      main_discount= request.POST.get("main_discount")
      extra_discount= request.POST.get("extra_discount")
      size= request.POST.get("size")
      
      if Product_group.objects.filter(brand_name__iexact=brand_name).exists():
            error_message = f"A name with the  '{brand_name}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
      else:
          data1 = Product_group(
                    
                    brand_name= brand_name,
                    sub_brand_name=sub_brand_name,
                    brand_subtitle=brand_subtitle,
                    categories=categories,
                    main_discount=main_discount,
                    extra_discount=extra_discount,
                    size=size,
                   
                    

          )
          data1.save()
          sweetify.success(request, " Created Successfully.", timer=6000, button='OK')
        
    data=Product_group.objects.all().order_by('-id')
    obj2=brands.objects.all()
    obj=Category.objects.all()
    obj1=Subbrand.objects.all()
    
    return render(request, 'group/product_group.html', {'data':data,'obj2':obj2,'obj':obj,'obj1':obj1,})


@login_required(login_url="") 
def delete_product_group(request, id):
    cus = Product_group.objects.get(id=id)
    cus.delete()
    return redirect('/product_group/')

@login_required(login_url="")
def edit_product_group(request, id):
    data = Product_group .objects.get(id=id)
    if request.method == 'POST':
            # data.brand_name = request.POST.get('brand_name')
            # data.sub_brand_name = request.POST.get('sub_brand_name')
            data.brand_subtitle = request.POST.get('brand_subtitle')
            # data.categories = request.POST.get('categories')
            data.main_discount = request.POST.get('main_discount')
            data.extra_discount = request.POST.get('extra_discount')
            data.size = request.POST.get('size')
            
            data.save()
           
            return redirect('/product_group/')
    
    print("datanew", data.brand_name)
    context = {'data':data,}
    return render(request, 'group/edit_product_group.html', context)


def categories(request):
   if request.method=="POST":
    category= request.POST.get("category")
    if Category.objects.filter(category__iexact=category).exists():
            error_message = f"A name with the  '{category}' already exists."
            sweetify.error(request, error_message, timer=8000, button='OK')
            
    else:
          data1 = Category(
                    
                    category= category,
          )
          data1.save()
          sweetify.success(request, " Created Successfully.", timer=6000, button='OK')
        
   data=Category.objects.all().order_by('-id')
   obj2=brands.objects.all()
   return render(request, 'group/category.html', {'data':data,'obj2':obj2,})


def delete_categories(request,id):
    cus = Category.objects.get(id=id)
    cus.delete()
    return redirect('/categories/')



def form_printing(request):
    if request.method== "POST":
        form_pages=request.POST.get("form_pages")
        form_vender=request.POST.get("form_vender")
        form_code=request.POST.get("form_code")
        form_brand=request.POST.get("form_brand")
        form_class=request.POST.get("form_class")
        form_medium=request.POST.get("form_medium")
        
        form_subject=request.POST.get("form_subject")
        form_quantity=request.POST.get("form_quantity")
        form1=request.POST.get("form1")
        form2=request.POST.get("form2")
        form3=request.POST.get("form3")
        form4=request.POST.get("form4")
        form5=request.POST.get("form5")
        form6=request.POST.get("form6")
        form7=request.POST.get("form7")
        form8=request.POST.get("form8")
        form9=request.POST.get("form9")
        form10=request.POST.get("form10")
        # transfer_1=request.POST.get("transfer_1")
        # transfer_2=request.POST.get("transfer_2")
        # transfer_3=request.POST.get("transfer_3")
        # transfer_4=request.POST.get("transfer_4")
        # transfer_5=request.POST.get("transfer_5")
        # # transfer5=request.POST.get("tra_from_6")
        # transfer_6=request.POST.get("transfer_6")
        # transfer_7=request.POST.get("transfer_7")
        # transfer_8=request.POST.get("transfer_8")
        # transfer_9=request.POST.get("transfer_9")
        # transfer_10=request.POST.get("transfer_10")
        form_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        total_forms=request.POST.get("total_forms")
        received_form_values = [
        float(request.POST.get(f'form{i}', '0')) for i in range(1, 11)
        ]
        total_forms = sum(received_form_values)
        print('total',total_forms)

        
        existing_entries = Form_Printing.objects.filter(
        form_brand=form_brand,
        form_class=form_class,
        form_medium=form_medium,
        form_subject=form_subject,
        # m1_no=m1_no,
        # a1_name=a1_name,
        form_pages=form_pages,
        # p_print_date = p_print_date,
        )


        # Extract the values of additional boxes
        additional_box_values = [
        int(request.POST.get(f'p_additional_box_{i}', '0'))
        for i in range(1, 11)
        ]

    # Calculate the total received forms, including the additional boxes
        received_form_values = [
            int(request.POST.get(f'form{i}', '0')) + additional_box_values[i - 1]
            for i in range(1, 11)
        ]
        total_forms = sum(received_form_values)

        existing_entries = Form_Printing.objects.filter(
            form_brand=form_brand,
            form_class=form_class,
            form_medium=form_medium,
            form_subject=form_subject,
            form_pages=form_pages,
        )

        new_entry_needed = True
        for entry in existing_entries:
            if entry.form_vender == form_vender:
                entry.form_date = form_date
                for i in range(1, 11):
                    entry_field_name = f'form{i}'
                    current_value = getattr(entry, entry_field_name, 0)
                    if current_value is not None:
                        new_value = int(current_value) + received_form_values[i - 1]
                    else:
                        new_value = received_form_values[i - 1]
                    setattr(entry, entry_field_name, new_value)
                entry.form_vender = form_vender
                entry.form_code = form_code

                if entry.total_forms is not None:
                    entry.total_forms = float(entry.total_forms) + float(total_forms)
                else:
                    entry.total_forms = float(total_forms)
                entry.save()
                new_entry_needed = False
                break

        
    
        if new_entry_needed:
        #  form_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
         data = Form_Printing(
            form_vender=form_vender,
            form_code=form_code,
            form_brand=form_brand,
            form_class=form_class,
            form_pages=form_pages,
           
            form_medium=form_medium,
            form_subject=form_subject,
            form_quantity=form_quantity,
            form1=form1,
            form2=form2,
            form3=form3,
            form4=form4,
            form5=form5,
            form6=form6,
            form7=form7,
            form8=form8,
            form9=form9,
            form10=form10,
            # transfer_1=transfer_1,
            # transfer_2=transfer_2,
            # transfer_3=transfer_3,
            # transfer_4=transfer_4,
            # transfer_5=transfer_5,
            # transfer_6=transfer_6,
            # transfer_7=transfer_7,
            # transfer_8=transfer_8,
            # transfer_9=transfer_9,
            # transfer_10=transfer_10,
            form_date = form_date,
            total_forms =total_forms,
            

         )
         data.save()
    obj=Covervender.objects.all()
    obj1=Company.objects.all()
    data = Cover_order.objects.all().order_by('-id')
    obj2 = books.objects.all()
    obj3=Form_Printing.objects.all()
    obj4= venders.objects.all()
    
    return render(request, 'brands/form_printing.html', {'data':data,'obj':obj,'obj1':obj1,'obj2':obj2,'obj3':obj3,'obj4':obj4,})

def delete_form_printing(request,id):
  cus = Form_Printing.objects.get(id=id)
  cus.delete()
  return redirect('/form_printing/')    

def edit_form_printing(request, id):
     data = Form_Printing.objects.get(id=id)
     if request.method == 'POST':
        data.form_vender=request.POST.get("form_vender")
       
        data.form_code=request.POST.get("form_code")
        data.form_brand=request.POST.get("form_brand")
        data.form_class=request.POST.get("form_class")
        data.form_medium=request.POST.get("form_medium")
        
        data.form_subject=request.POST.get("form_subject")
        data.form_quantity=request.POST.get("form_quantity")
        
        data.form_pages=request.POST.get("form_pages")
        print("ggggggggggggg",data.form_pages)
        data.form1=request.POST.get("form1")
        print("ggggggggggggg",data.form1)
        data.form2=request.POST.get("form2")
        data.form3=request.POST.get("form3")
        data.form4=request.POST.get("form4")
        data.form5=request.POST.get("form5")
        data.form6=request.POST.get("form6")
        data.form7=request.POST.get("form7")
        data.form8=request.POST.get("form8")
        data.form9=request.POST.get("form9")
        data.form10=request.POST.get("form10")
        data.total_forms=request.POST.get("total_forms")
        print('total',data.total_forms)
        

        data.form_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
         
        # data.p_print_date=request.POST.get('p_print_date')
        data.total_forms = ""
        # Convert gather fields to float (modify as needed based on your model)
        received_form_values = [
        int(getattr(data, f'form{i}') or 0) if getattr(data, f'form{i}') is not None and getattr(data, f'form{i}') != 'None' else 0 for i in range(1, 11)
        ]

        data.total_forms = sum(received_form_values)
        print('total', data.total_forms)    
        
        data.save()
        print("ggggggggggggg",data.form1)
        messages.success(request, "Updated Successfully...!!")
        return redirect('/form_printing/')
    
     data2 = venders.objects.all()  # Correct the model name based on your actual model
    
     context = {'data': data, 'data2': data2}
     return render(request, "brands/edit_form_printing.html", context)


def cover_rawmaterial(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        cmname = request.POST.get('cmname')
        cmalias = request.POST.get('cmalias')
        cprint_name = request.POST.get('cprint_name')
        cqunit = request.POST.get('cqunit')
        copening_stock_value = request.POST.get('copening_stock_value')
        cgsm = request.POST.get('cgsm')
        cp_width = request.POST.get('cp_width')
        cp_length = request.POST.get('cp_length')
        cunit1 = request.POST.get('cunit1')
        cqunit1 = request.POST.get('cqunit1')
        cunit2 = request.POST.get('cunit2')
        cqunit2 = request.POST.get('cqunit2')
        cunit3 = request.POST.get('cunit3')
        cqunit3 = request.POST.get('cqunit3')
        cunit1_conversion = request.POST.get('cunit1_conversion')
        cmunit1 = request.POST.get('cmunit1')
        cunit2_conversion = request.POST.get('cunit2_conversion')
        cmunit2 = request.POST.get('cmunit2')
        opening_stock_rm = request.POST.get('opening_stock_rm')
        p_date = request.POST.get('p_date')
        gross = request.POST.get('gross')
        cwaste_percent=request.POST.get('cwaste_percent')
        cWaste=request.POST.get('cWaste')
        opening_stock1= request.POST.get('opening_stock1')
        opening_unit1 = request.POST.get('opening_unit1')
        opening_unit2 = request.POST.get('opening_unit2')
        opening_stock2 = request.POST.get('opening_stock2')
        conversion_factor1= request.POST.get('conversion_factor1')
        conversion_unit1= request.POST.get('conversion_unit1')
        conversion_unit2= request.POST.get('conversion_unit2')

        conversion_factor2= request.POST.get('conversion_factor2')
        unit4= request.POST.get('unit4')
        cover_material = request.POST.get('cover_material')


        # Rest of your code
        books_data = Cover_rawmaterial(
            p_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'),
            cmname=cmname,
            cmalias=cmalias,
            cprint_name=cprint_name,
            cqunit=cqunit,
            copening_stock_value=copening_stock_value,
            cgsm=cgsm,
            
            cp_width=cp_width,
            cp_length=cp_length,
            cunit1=cunit1,
            cqunit1=cqunit1,
            cunit2=cunit2,
            cqunit2=cqunit2,
            cunit3=cunit3,
            cqunit3=cqunit3,
            cunit1_conversion=cunit1_conversion,
            cmunit1=cmunit1,
            cunit2_conversion=cunit2_conversion,
            cmunit2=cmunit2,
            opening_stock_rm=opening_stock_rm,
            
            gross=gross,
            cwaste_percent=cwaste_percent,
            cWaste=cWaste,
            opening_stock1=opening_stock1,
            opening_unit1=opening_unit1,
            opening_unit2=opening_unit2,
            opening_stock2=opening_stock2,
            conversion_factor1=conversion_factor1,
            conversion_unit1=conversion_unit1,
            conversion_unit2=conversion_unit2,
            conversion_factor2=conversion_factor2,
            unit4=unit4,
            cover_material=cover_material,

        )
        books_data.save()

    data = Cover_rawmaterial.objects.all()
    data1= master_unit.objects.all()
    center = material_centre.objects.all()
    return render(request, 'form/cover_rawmaterial.html', {'data': data, 'center': center,'data1':data1,})

@login_required(login_url="")
def edit_cover_rawmaterial(request,id):
    data=Cover_rawmaterial.objects.get(id=id)
    if request.method=='POST':
        id=request.POST.get('id')
        data.cmname =request.POST.get('cmname')
        data.cmalias =request.POST.get('cmalias')
        data.cprint_name =request.POST.get('cprint_name')
        data.copening_stock_value =request.POST.get('copening_stock_value')
        
        data.cover_material =request.POST.get('cover_material')
       
        # data.gsm_paper=request.POST.get('gsm_paper')
        data.cgsm = request.POST.get('cgsm')
        data.cp_width = request.POST.get('cp_width')
        data.cp_length = request.POST.get('cp_length')
        data.cunit1 = request.POST.get('cunit1')
        data.cqunit1 = request.POST.get('cqunit1')
        data.cunit2 = request.POST.get('cunit2')
        data.cqunit2 = request.POST.get('cqunit2')
        data.cunit3 = request.POST.get('cunit3')
        print("22222222222222",data.cunit3)
        data.cqunit3 = request.POST.get('cqunit3')
        data.cunit1_conversion = request.POST.get('cunit1_conversion')
        data.cmunit1 = request.POST.get('cmunit1')
        data.cunit2_conversion = request.POST.get('cunit2_conversion')
        data.cmunit2 = request.POST.get('cmunit2')
        data.opening_stock_rm = request.POST.get('opening_stock_rm')
        
        
        data.cWaste = request.POST.get('cWaste')
        data.cwaste_percent = request.POST.get('cwaste_percent')
        data.opening_stock1= request.POST.get('opening_stock1')
        data.opening_unit1 = request.POST.get('opening_unit1')
        data.opening_unit2 = request.POST.get('opening_unit2')
        data.opening_stock2 = request.POST.get('opening_stock2')
        data.conversion_factor1= request.POST.get('conversion_factor1')
        data.conversion_unit1= request.POST.get('conversion_unit1')
        data.conversion_unit2= request.POST.get('conversion_unit2')

        data.conversion_factor2= request.POST.get('conversion_factor2')
        data.unit4= request.POST.get('unit4')
        
        
            
        data.save()
        
        messages.success(request,'Updated Successfully...!')
        return redirect('/cover_rawmaterial/')
    data2= master_unit.objects.all()
    center = material_centre.objects.all()
    context={'data':data,'data2':data2,'center':center}
    return render(request,'form/edit_cover_rawmaterial.html',context)




@login_required(login_url="")
def delete_cover_rawmaterial(request, id):
    cus = Cover_rawmaterial.objects.get(id=id)
    cus.delete()
    return redirect('/cover_rawmaterial/')



def unit_master(request):
    if request.method=='POST':
        unit_name_master=request.POST.get('unit_name_master')
        unit_Alias=request.POST.get('unit_Alias')
        unit_print_name=request.POST.get('unit_print_name')
        unit=master_unit(unit_name_master=unit_name_master,unit_Alias=unit_Alias,unit_print_name=unit_print_name)
        unit.save()
    data=master_unit.objects.all()
    context={
        'data':data,
    }


    return render(request, 'classes/unit_master.html',context)

def edit_unit_master(request,id):
    data=master_unit.objects.get(id=id)
    if request.method=='POST':
        data.unit_name_master=request.POST.get('unit_name_master')
        data.unit_Alias=request.POST.get('unit_Alias')
        data.unit_print_name=request.POST.get('unit_print_name')
        data.save()

        return redirect('/unit_master/')
    context={'data':data}
    return render(request,'classes/edit_unit_master.html',context)


@login_required(login_url="")
def delete_unit_master(request, id):
    cus = master_unit.objects.get(id=id)
    cus.delete()
    return redirect('/unit_master/')



def unit_conversion(request):
    if request.method=='POST':
        main_unit=request.POST.get('main_unit')
        sub_unit=request.POST.get('sub_unit')
        Con_factor=request.POST.get('Con_factor')
        factor=conversion_unit(main_unit=main_unit,sub_unit=sub_unit,Con_factor=Con_factor)
        factor.save()
    data=conversion_unit.objects.all()

    context={
        'data':data,
    }
        
    return render(request,'classes/unit_conversion.html',context)


@login_required(login_url="")
def delete_unit_conversion(request, id):
    cus = conversion_unit.objects.get(id=id)
    cus.delete()
    return redirect('/unit_conversion/')


def edit_unit_conversion(request,id):
    data=conversion_unit.objects.get(id=id)
    if request.method=='POST':
        data.main_unit=request.POST.get('main_unit')
        data.sub_unit=request.POST.get('sub_unit')
        data.Con_factor=request.POST.get('Con_factor')
        data.save()

        return redirect('/unit_conversion/')
    context={'data':data}
    return render(request,'classes/edit_unit_conversion.html',context)



def item_masters(request):

    if request.method=='POST':
        item_name=request.POST.get('item_name')
        item_alias=request.POST.get('item_alias')
        item_print_name=request.POST.get('item_print_name')
        item_group=request.POST.get('item_group')
        factor=item_master(item_group=item_group,item_name=item_name,item_alias=item_alias,item_print_name=item_print_name)
        factor.save()
    data=item_master.objects.all()

    context={
        'data':data,
    }
    return render(request,'brands/item_master.html',context)




def pending_printing(request,id):
    data = printing.objects.get(id=id)
   
    if request.method == 'POST':
        data.tra_from_1=request.POST.get("tra_from_1")
        data.to_1=request.POST.get("to_1")
        data.transfer1=request.POST.get("transfer1")
        data.transfer2=request.POST.get("transfer2")
        data.transfer3=request.POST.get("transfer3")
        data.transfer4=request.POST.get("transfer4")
        data.transfer5=request.POST.get("transfer5")
        data.transfer6=request.POST.get("transfer6")
        data.transfer7=request.POST.get("transfer7")
        data.transfer8=request.POST.get("transfer8")
        data.transfer9=request.POST.get("transfer9")
        data.transfer10=request.POST.get("transfer10")
          
        total_form1 =int(data.transfer_total_form1 or '0')
        total_form2 =int(data.transfer_total_form2 or '0')
        total_form3 =int(data.transfer_total_form3 or '0')
        total_form4 =int(data.transfer_total_form4 or '0')
        total_form5 =int(data.transfer_total_form5 or '0')
        total_form6 =int(data.transfer_total_form6 or '0')
        total_form7 =int(data.transfer_total_form7 or '0')
        total_form8 =int(data.transfer_total_form8 or '0')
        total_form9 =int(data.transfer_total_form9 or '0')
        total_form10=int(data.transfer_total_form10 or '0')
        
        data.transfer_total_form1 =total_form1 + int(data.transfer1 or '0')
        data.transfer_total_form2 =total_form2 + int(data.transfer2 or '0')
        data.transfer_total_form3 =total_form3 + int(data.transfer3 or '0')
        data.transfer_total_form4 =total_form4 + int(data.transfer4 or '0')
        data.transfer_total_form5 =total_form5 + int(data.transfer5 or '0')
        data.transfer_total_form6 =total_form6 + int(data.transfer6 or '0')
        data.transfer_total_form7 =total_form7 + int(data.transfer7 or '0')
        data.transfer_total_form8 =total_form8 + int(data.transfer8 or '0')
        data.transfer_total_form9 =total_form9 + int(data.transfer9 or '0')
        data.transfer_total_form10=total_form10 + int(data.transfer10 or '0')
        
        data.p_stock=request.POST.get("p_stock")
        data.p_vender_name=request.POST.get("p_vender_name")
        data.p_code=request.POST.get("p_code")
        data.p_vender_address =request.POST.get("p_vender_address")
        data.p_vender_mob=request.POST.get("p_vender_mob")
        data.p_brand_name_1=request.POST.get("p_brand_name_1")
        data.p_class_name_1=request.POST.get("p_class_name_1")
        data.p_medium_1=request.POST.get("p_medium_1")    
        data.p_subject_1=request.POST.get("p_subject_1")
        data.Material_center = request.POST.get('Material_center')
        data.p_quantity_1=request.POST.get("p_quantity_1")
        data.p_rim_1=request.POST.get("p_rim_1")
        data.p_pages_1=request.POST.get("p_pages_1")
        data.p_single_1=request.POST.get("p_single_1")
        data.p_double_1=request.POST.get("p_double_1")
        data.p_four_1=request.POST.get("p_four_1")
        data.p_print_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.transfer=request.POST.get("transfer")
        data.pending=request.POST.get("transfer_pending")
        transfer = int(data.transfer) if data.transfer else 0
        data.total_recived_forms=request.POST.get('total_recived_forms')
        data.p_print_date=request.POST.get('p_print_date')
        total_received_forms = ""
        received_form_values = [
            int(getattr(data, f'p_printing_received_form_{i}') or 0) if getattr(data, f'p_printing_received_form_{i}') is not None else 0 for i in range(1, 11)
        ]
        data.total_received_forms = sum(received_form_values)
        data.tra_from_1 = data.tra_from_1   
        data.to_1 = data.to_1 
        data.transfer1 = data.transfer1  
        data.Material_center=data.Material_center
            
        current_value1 = int(data.transfer_pending1 or '0')
        current_value2 = int(data.transfer_pending2 or '0')
        current_value3 = int(data.transfer_pending3 or '0')
        current_value4 = int(data.transfer_pending4 or '0')
        current_value5 = int(data.transfer_pending5 or '0')
        current_value6 = int(data.transfer_pending6 or '0')
        current_value7 = int(data.transfer_pending7 or '0')
        current_value8 = int(data.transfer_pending8 or '0')
        current_value9 = int(data.transfer_pending9 or '0')
        current_value10 = int(data.transfer_pending10 or '0')
        
        data.transfer_pending1 = current_value1 - int(data.transfer1 or '0')
        data.transfer_pending2 = current_value2 - int(data.transfer2 or '0')
        data.transfer_pending3 = current_value3 - int(data.transfer3 or '0')
        data.transfer_pending4 = current_value4 - int(data.transfer4 or '0')
        data.transfer_pending5 = current_value5 - int(data.transfer5 or '0')
        data.transfer_pending6 = current_value6 - int(data.transfer6 or '0')
        data.transfer_pending7 = current_value7 - int(data.transfer7 or '0')
        data.transfer_pending8 = current_value8 - int(data.transfer8 or '0')
        data.transfer_pending9 = current_value9 - int(data.transfer9 or '0')
        data.transfer_pending10 = current_value10 - int(data.transfer10 or '0')
       
        total_data_transfer = ((data.transfer_pending1)+(data.transfer_pending2)+(data.transfer_pending3)+(data.transfer_pending4)+
        (data.transfer_pending5)+(data.transfer_pending6)+(data.transfer_pending7)+(data.transfer_pending8)+(data.transfer_pending9)+(data.transfer_pending10))
       
        
        total_pending_transfer = int(data.total_transfer_form or '0')
        
        transfer_data = (int(data.transfer1 or '0') + int(data.transfer2 or '0') + int(data.transfer3 or '0') +
                 int(data.transfer4 or '0') + int(data.transfer5 or '0') + int(data.transfer6 or '0') +
                 int(data.transfer7 or '0') + int(data.transfer8 or '0') + int(data.transfer9 or '0') +
                 int(data.transfer10 or '0'))
    
        data.total_transfer_form = total_pending_transfer + int(transfer_data)
      
        transfer = ""
        
        transfer_form_values = [
            int(getattr(data, f'transfer{i}') or 0) if getattr(data, f'transfer{i}') is not None else 0 for i in range(1, 11)
        ]

        data.transfer = sum(transfer_form_values)
        transfer_pending = ""
       
        transfer_form_values = [
            int(getattr(data, f'transfer_pending{i}') or 0) if getattr(data, f'transfer_pending{i}') is not None else 0 for i in range(1, 11)
        ]

        data.transfer_pending = sum(transfer_form_values)
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("print_approval") == "approval_button":
                printing.objects.filter(Q(id=id) & Q(print_approved_1=False)).update(print_approved_1=True)
                return redirect('/printing-data/')
                

      
    data2 = venders.objects.all()
    data3=Bindingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data5':data5}
    
    return render(request, 'brands/pending_printing.html', context)



def pending_gathering(request,id):
    data = Gathering.objects.get(id=id)
    
    if request.method == 'POST':
        data.tra_from_2=request.POST.get("tra_from_2")
        data.to_2=request.POST.get("to_2")
        print("1234",data.to_2)
        data.code_g = request.POST.get('code_g')
        data.v1_name = request.POST.get('v1_name')
        data.m1_no = request.POST.get('m1_no')
        data.a1_name = request.POST.get('a1_name')
        data.Material_center = request.POST.get('Material_center')
        data.date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.transfer=request.POST.get('transfer')
        data.pending=request.POST.get("transfer_pending")
        data.transfer1 = request.POST.get("transfer1")
        transfer = int(data.transfer) if data.transfer else 0
        
        total_form =int(data.total_transfer_form or '0')
        data.total_transfer_form =total_form + int(data.transfer1 or '0')
        
        print('total_transfer_form',data.total_transfer_form, data.transfer1, data.transfer)
        
        total_quantity = int(data.g_quantity or '0')
        data.total_remaining=total_quantity - int(data.total_transfer_form or 0)
        print('data.total_remaining',data.total_remaining,data.pending,total_quantity)
        
        # Convert gather fields to float (modify as needed based on your model)
        gather_values =[
        
            int(getattr(data, f'gather{i}') or 0) if getattr(data, f'gather{i}') is not None else 0 for i in range(1, 11)
        ]
        total_gather = sum(gather_values)
        data.total_gather = total_gather
        data.tra_from_2 = data.tra_from_2
        print('fghjk',data.tra_from_2)
        data.Material_center=data.Material_center
        data.to_2 = data.to_2
        print('fghjk',data.to_2)
        data.pending=int(data.g_quantity) - int(transfer)
        print("ppppppppppeeending",data.pending)
        data.transfer_pending = data.total_remaining
        data.save()
        messages.success(request, "Updated Successfully...!!")
        if request.method == 'POST':
            if request.POST.get("gather_approved_1") == "approval_button":
                Gathering.objects.filter(Q(id=id) & Q( gather_approved_1=False)).update( gather_approved_1=True)
                return redirect('/gathering/')
              
        # return redirect('/gathering/')
    
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3=Bindingvender.objects.all()
    data4=Gather_ledger.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data4':data4,'data5':data5}
    return render(request, 'brands/pending_gathering.html',context)



def pending_machine(request, id):
    data = Binding.objects.get(id=id)
    if request.method == 'POST':
        data.tra_from_b=request.POST.get("tra_from_b")
        data.to_b=request.POST.get("to_b")
        data.b_code=request.POST.get("b_code")
        data.voucher_b=request.POST.get("voucher_b")
        data.b_v1_name = request.POST.get('b_v1_name')
        data.b_m1_no = request.POST.get('b_m1_no')
        data.b_a1_name = request.POST.get('b_a1_name')
        data.Material_center = request.POST.get('Material_center')
        data.b_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.total_binding=request.POST.get('total_binding')
        data.transfer=request.POST.get("transfer")
        transfer1=request.POST.get("transfer1")
        data.pending=request.POST.get("transfer_pending")
        transfer = int(data.transfer) if data.transfer else 0
        data.tra_from_b = data.tra_from_b
        data.to_b = data.to_b
        data.Material_center=data.Material_center
        total_form =int(data.total_transfer_form or '0')
        data.total_transfer_form =total_form + int(transfer1 or '0')
        total_quantity = int(data.total_binding or '0')
        data.total_remaining=total_quantity - int(data.total_transfer_form or 0)
        data.pending=data.total_remaining
        data.transfer_pending = data.total_remaining
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("binding_approval") == "approval_button":
                Binding.objects.filter(Q(id=id) & Q(binding_approved_1=False)).update(binding_approved_1=True)
                return redirect('/machine_binding/')
      
    data2 = venders.objects.all() 
    data3=Bindingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data,'data3':data3, 'data2': data2,'data5':data5}
    return render(request, "brands/pending_machine.html", context)


def pending_manual(request, id):
    data = Manual.objects.get(id=id)
    if request.method == 'POST':
        
        data.tra_from_m = request.POST.get('tra_from_m')
        data.to_m= request.POST.get('to_m')
        data.voucher_m = request.POST.get('voucher_m')
        data.m_code = request.POST.get('m_code')
        data.b_vender_name = request.POST.get('b_vender_name')
        data.b_mobile_no = request.POST.get('b_mobile_no')
        data.b_adress_name = request.POST.get('b_adress_name')
        data.m_brand_name1 = request.POST.get('m_brand_name1')
        data.m_class_name1 = request.POST.get('m_class_name1')
        data.m_medium1 = request.POST.get('m_medium1')
        data.m_subject1 = request.POST.get('m_subject1')
        data.m_striching1 = request.POST.get('m_striching1')
        data.transfer=request.POST.get("transfer")
        data.pending=request.POST.get("transfer_pending")
        data.Material_center = request.POST.get('Material_center')
        transfer1=request.POST.get("transfer1")
        
        transfer = int(data.transfer) if data.transfer else 0
        # data.m_pasting1 = request.POST.get('m_pasting1')
        # data.m_cutting1 = request.POST.get('m_cutting1')
        total_form =int(data.total_transfer_form or '0')
        data.total_transfer_form =total_form + int(transfer1 or '0')
        total_stiching = int(data.m_striching1 or '0')
        data.total_remaining=total_stiching - int(data.total_transfer_form or 0)
        print('data.total_remaining',data.total_remaining,total_stiching,data.total_transfer_form)
       
        data.b_print_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        
        # Convert gather fields to float (modify as needed based on your model)
        data.tra_from_m = data.tra_from_m
        data.Material_center =data.Material_center
        print('fghjk',data.tra_from_m)
        data.to_m = data.to_m
        print('fghjk',data.to_m)
        data.pending=data.total_remaining
        print("ppppppppppeeending",data.pending)
        data.transfer_pending = data.total_remaining
        
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("stiching_approval") == "approval_button":
                Manual.objects.filter(Q(id=id) & Q(stiching_approved_1=False)).update(stiching_approved_1=True)
                return redirect('/manual-info/')
               
        
       
        
    
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3=Bindingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data5':data5}
    return render(request, "brands/pending_manual.html", context)



def pending_pasting(request, id):
    data = Pasting.objects.get(id=id)
    if request.method == 'POST':
        data.tra_from_pasting = request.POST.get('tra_from_pasting')
        data.to_pasting= request.POST.get('to_pasting')
        data.voucher_pasting = request.POST.get('voucher_pasting')
        
        data.pasting_code = request.POST.get('pasting_code')
        data.pasting_vender_name = request.POST.get('pasting_vender_name')
        data.pasting_mobile_no = request.POST.get('pasting_mobile_no')
        data.pasting_adress_name = request.POST.get('pasting_adress_name')
        data.pasting_brand_name1 = request.POST.get('pasting_brand_name1')
        data.pasting_class_name1 = request.POST.get('pasting_class_name1')
        data.pasting_medium1 = request.POST.get('pasting_medium1')
        data.pasting_subject1 = request.POST.get('pasting_subject1')
        data.Material_center = request.POST.get('Material_center')
        # data.pasting_striching1 = request.POST.get('pasting_striching1')
        data.pasting_pasting1 = request.POST.get('pasting_pasting1')
        data.transfer=request.POST.get("transfer")
        transfer1=request.POST.get("transfer1")
        data.pending=request.POST.get("transfer_pending")
        transfer = int(data.transfer) if data.transfer else 0
        
        total_form =int(data.total_transfer_form or '0')
        data.total_transfer_form =total_form + int(transfer1 or '0')
        total_pasting = int(data.pasting_pasting1 or '0')
        data.total_remaining=total_pasting - int(data.total_transfer_form or 0)
        print('data.total_remaining',data.total_remaining,total_pasting,data.total_transfer_form)
       


        # data.pasting_cutting1 = request.POST.get('pasting_cutting1')
       
        data.pasting_print_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        
        # Convert gather fields to float (modify as needed based on your model)
        data.tra_from_pasting = data.tra_from_pasting
       
        data.to_pasting= data.to_pasting
      
        data.pending=data.total_remaining
      
        data.transfer_pending = data.total_remaining
        data.Material_center = data.Material_center 
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("pasting_approval") == "approval_button":
                Pasting.objects.filter(Q(id=id) & Q(pasting_approved_1=False)).update(pasting_approved_1=True)
                return redirect('/pasting/')
        
        # return redirect('/pasting/')
   
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3 = Cuttingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data5':data5}
    return render(request, "brands/pending_pasting.html", context)



def pending_cutting(request, id):
    data = Cuttting.objects.get(id=id)
    if request.method == 'POST':
        data.c_code = request.POST.get('c_code')
        data.tra_from_c = request.POST.get('tra_from_c')
        data.to_c = request.POST.get('to_c')
        data.voucher_c = request.POST.get('voucher_c')
        data.c_brand_name = request.POST.get('c_brand_name')
        data.c_v1_name = request.POST.get('c_v1_name')
        data.c_m1_no = request.POST.get('c_m1_no')
        data.c_a1_name = request.POST.get('c_a1_name')
        data.c_brand_name = request.POST.get('c_brand_name')
        data.c_class_name = request.POST.get('c_class_name')
        data.c_medium = request.POST.get('c_medium')
        data.c_subject = request.POST.get('c_subject')
        data.Material_center = request.POST.get('Material_center')
        
        data.c_date=datetime.now().strftime('%d-%m-%Y %I:%M:%S %p')
        data.c_quantity_5=request.POST.get('c_quantity_5')

        data.transfer=request.POST.get("transfer")
        transfer1=request.POST.get("transfer1")
        data.pending=request.POST.get("transfer_pending")

        transfer = int(data.transfer) if data.transfer else 0
        
        total_form =int(data.total_transfer_form or '0')
        data.total_transfer_form =total_form + int(transfer1 or '0')
        total_cutting = int(data.c_quantity_5 or '0')
        data.total_remaining=total_cutting - int(data.total_transfer_form or 0)
        # print('data.total_remaining',data.total_remaining,total_cutting,data.total_transfer_form)
     
        data.pending=data.total_remaining 
        data.transfer_pending = data.total_remaining
        data.tra_from_c = data.tra_from_c
        data.Material_center=data.Material_center 
        data.to_c = data.to_c
        
        data.save()
        messages.success(request, "Updated Successfully...!!")
        
        if request.method == 'POST':
            if request.POST.get("machine_cutting_approval") == "approval_button":
                Cuttting.objects.filter(Q(id=id) & Q(machine_cutting_approved_1=False)).update(machine_cutting_approved_1=True)
                return redirect('/cutting_info/')
            # else:
            #     pass    
        
        
        # return redirect('/cutting_info/')
    
    data2 = venders.objects.all()  # Correct the model name based on your actual model
    data3 = Packingvender.objects.all()
    data5=material_centre.objects.all()
    
    context = {'data': data, 'data2': data2,'data3':data3,'data5':data5}
    return render(request, "brands/pending_cutting.html", context)